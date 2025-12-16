import os
from openpyxl import Workbook
import logging
# import logging.handlers
from logging.handlers import TimedRotatingFileHandler

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import chardet
import shutil
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
import datetime
import base64



class Arquivos():
    """Manipula arquivos."""

    def __init__(self):
        self.module_logger = self.criar_logger_diario("./log", nome_arquivo_log="manipula_diretorio_arquivos.log", logger=logging.getLogger(__name__))

    def arquivos_existem(self, *caminhos):
        """
        Verifica se todos os arquivos fornecidos existem.

        :param *caminhos: Um ou mais caminhos de arquivos a serem verificados.
        :type *caminhos: str

        :returns: True se todos os arquivos existem, False caso contrário.
        :rtype: bool
        """
        return all(os.path.isfile(caminho) for caminho in caminhos)

    def criar_arquivo(self, caminho):
        """
        Cria um arquivo vazio caso ele não exista no caminho especificado.

        Cria o(s) diretorio(s) caso não exista.

        :param caminho: Caminho completo do arquivo a ser criado.
        :type caminho: str

        :returns: True se o arquivo foi criado, False caso contrário.
        :rtype: bool
        """
        if not self.arquivos_existem(caminho):
            # Garante que o diretório do arquivo exista
            os.makedirs(os.path.dirname(caminho), exist_ok=True)
            with open(caminho, 'w') as f:
                pass  # Cria um arquivo vazio
            # print(f"✅ Arquivo criado: {caminho}")
            self.module_logger.info(f"✅ Arquivo criado: {caminho}")
            return True
        else:
            # print(f"ℹ️ Arquivo já existe: {caminho}")
            self.module_logger.warning(f"ℹ️ Arquivo já existe: {caminho}")
            return False

    def criar_xlsx_com_colunas(self, caminho_arquivo, colunas=None):
        """
        Cria um arquivo .xlsx com uma planilha contendo as colunas se forem especificadas.

        Sobrescreve o arquivo se ele já existir.

        :param caminho_arquivo: Caminho e nome do arquivo Excel a ser criado.
        :type caminho_arquivo: str
        :param colunas: Lista com os nomes das colunas. (ex: ["Nome", "Idade", "Email"]).
        :type colunas: list

        :returns: True se o caminho for criado com sucesso, False caso contrario
        :rtype: bool
        """
        if not self.arquivos_existem(caminho_arquivo):
            self.criar_arquivo(caminho_arquivo)
        try:
            wb = Workbook()
            ws = wb.active

            if colunas:
                ws.append(colunas)  # Insere os nomes das colunas como primeira linha

            wb.save(caminho_arquivo)
            # print(f"Arquivo '{caminho_arquivo}' criado com as colunas: {colunas}")
            self.module_logger.info(f"Arquivo '{caminho_arquivo}' criado com as colunas: {colunas}")
            return True
        except Exception as erro:
            # print('❌ OCORREU UM ERRO:', erro)
            self.module_logger.error(f"❌ OCORREU UM ERRO: {erro}")
            return False

    def inserir_tabela_xlsx(self, lista_de_dicionarios, caminho_arquivo, dic_nome_colunas=None):
        """
        Monta tabela e salva em uma planilha xlsx.

        :param lista_de_dicionarios: Lista contendo dicionários com os dados.
        :type lista_de_dicionarios: list
        :param caminho_arquivo: Caminho completo do arquivo Excel (.xlsx) a ser formatado.
        :type caminho_arquivo: str
        :param dic_nome_colunas: dicionario contendo as colunas de lista_de_dicionarios como chaves e o nome da coluna que deseja como valor.
        :type dic_nome_colunas: dict
        """
        try:
            self.module_logger.info("Estruturando dados em tabela...")
            df = pd.DataFrame(lista_de_dicionarios)

            # Converte campos datetime para string, se houver
            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]' or df[col].apply(lambda x: str(type(x))).str.contains("datetime").any():
                    df[col] = df[col].dt.strftime('%d/%m/%Y')

            if dic_nome_colunas is not None:
                #  Apenas as colunas indicadas são renomeadas.
                # inplace=True altera o df diretamente.
                df.rename(columns = dic_nome_colunas,
                          inplace=True)

                # substituir todos os nomes das colunas:
                # df.columns = ['CPF', 'APELIDO', 'DATA HORA CRIAÇÃO', 'DATA HORA MODIFICAÇÃO', 'DATA ADMISSÃO', 'STAATV RM', 'NOME', 'LOG', 'CÓDIGO SAP']

            df.to_excel(caminho_arquivo, index=False)
            self.module_logger.info(f"✅ Relatório salvo com sucesso em: {caminho_arquivo}")

        except Exception as e:
            self.module_logger.error(f"❌ Erro ao estruturar dados em tabela: {e}")

    def formatar_xlsx(self, caminho_arquivo, aba="Sheet1"):
        """
        Aplica estilo visual ao cabeçalho de uma planilha Excel e ajusta automaticamente a largura das colunas.

        :param caminho_arquivo: Caminho completo do arquivo Excel (.xlsx) a ser formatado.
        :type caminho_arquivo: str
        :param aba: Nome da aba onde aplicar o estilo. Padrão é 'Sheet1'.
        :type aba: str

        Estilos aplicados:
        - Cabeçalho com fundo azul (cor: #0070C0)
        - Texto do cabeçalho em branco, negrito e centralizado
        - Largura das colunas ajustada com base no maior conteúdo da coluna
        """
        try:
            self.module_logger.info("Formatando o estilo do relatorio...")

            wb = load_workbook(caminho_arquivo)

            try:
                ws = wb[aba]
            except:
                ws = wb["Sheet"]

            # Definição do estilo azul
            cabecalho_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            fonte_branca = Font(color="FFFFFF", bold=True)
            alinhamento_central = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primeira linha = cabeçalho
                cell.fill = cabecalho_fill
                cell.font = fonte_branca
                cell.alignment = alinhamento_central

            wb.save(caminho_arquivo)
            self.module_logger.info("✅ Formatação aplicada")

        except Exception as e:
            self.module_logger.error(f"❌ Erro ao formatar o estilo do relatorio: {e}")


        try:
            self.module_logger.info("Ajustando a largura das colunas do relatorio...")

            wb = load_workbook(caminho_arquivo)
            try:
                ws = wb[aba]
            except:
                ws = wb["Sheet"]

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 4  # margem extra

            wb.save(caminho_arquivo)

            self.module_logger.info("✅ Estilo e largura aplicados")
        except Exception as e:
            self.module_logger.error(f"❌ Erro ao aplicar largura: {e}")

        self.module_logger.info(f"✅ Relatório formatado e salvo com sucesso em: {caminho_arquivo}")

    def criar_logger_diario(self, diretorio_log, nome_arquivo_log="log", logger=logging.getLogger()):
        """
        Cria e configura um logger que grava logs com rotação diária, criando um novo arquivo a cada dia.

        :param diretorio_log: Caminho para o diretório onde os logs serão salvos.
        :type diretorio_log: str
        :param nome_arquivo_log: Nome base do arquivo de log (ex: "log" -> log_2025-06-20.log).
        :type nome_arquivo_log: str
        :param logger: Logger nomeado para __name__.
        :type logger: logging.Logger

        :returns: Logger configurado com rotação diária.
        :rtype: logging.Logger
        """
        # Caminho base para logs rotacionados: log/log_
        dir_arq_log = os.path.join(diretorio_log, nome_arquivo_log)

        try:
            # self.criar_arquivo(dir_arq_log)
            os.makedirs(diretorio_log, exist_ok=True)

            logger.setLevel(logging.DEBUG)
        except Exception as e:
            raise Exception(f"❌ Rootlogger não pode ser iniciado.: {e}")

        try:
            if not logger.handlers:
                formatter = logging.Formatter(
                    "%(name)s - %(levelname)s - %(asctime)s - %(message)s"
                )

                # Handler com rotação diária
                handler = TimedRotatingFileHandler(
                    filename=dir_arq_log,
                    when='midnight',             # rotação diária
                    interval=1,                  # a cada 1 dia
                    backupCount=30,              # mantém últimos 30 arquivos
                    encoding='utf-8',
                    utc=False
                )
                handler.suffix = "%Y-%m-%d.log"  # formato do nome final do arquivo
                handler.setFormatter(formatter)

                # Também loga no console
                console_handler = logging.StreamHandler()
                console_handler.setFormatter(formatter)

                logger.addHandler(handler)
                logger.addHandler(console_handler)

            return logger
        except Exception as e:
            raise Exception(f"O handler do arquivo de log não pode ser configurado. {e}")

    def mover_arquivo(self, origem, destino_pasta, sobrescrever=False):
        """Move um arquivo para outro diretorio.

        :param origem: Caminho completo do arquivo.
        :type origem: str
        :param destino_pasta: Caminho para onde se quer mover o arquivo
        :type destino_pasta: str
        :param sobrescrever: Define se deve sobrescrever um arquivo existente no destino.
        :type sobrescrever: bool
        :raises FileNotFoundError: Se o arquivo de origem ou o diretório de destino não forem encontrados.
        """
        nome_arquivo = os.path.basename(origem)
        destino_arquivo = os.path.join(destino_pasta, nome_arquivo)


        if not self.arquivos_existem(origem):
            self.module_logger.warning("❌ Arquivo não encontrado, verifique se o caminho está correto.")
            raise FileNotFoundError("❌ Arquivo não encontrado, verifique se o caminho está correto.")

        if not os.path.isdir(destino_pasta):
            self.module_logger.warning("❌ Diretório de destino não encontrado.")
            raise FileNotFoundError("❌ Diretório de destino não encontrado.")

        if os.path.exists(destino_arquivo):
            if not sobrescrever:
                self.module_logger.warning("⚠️ Já existe um arquivo com o mesmo nome no destino.")
                raise FileExistsError("⚠️ Já existe um arquivo com o mesmo nome no destino.")
            else:
                self.module_logger.info("♻️ Arquivo existente será sobrescrito.")

        shutil.move(origem, destino_arquivo)
        self.module_logger.info(f"✅ Arquivo movido com sucesso para {destino_arquivo}")

    def copiar_arquivo(self, origem, destino_pasta, sobrescrever=False):
        """
        Copia um arquivo para outro diretório.

        :param origem: Caminho completo do arquivo de origem.
        :type origem: str
        :param destino_pasta: Caminho para o diretório de destino.
        :type destino_pasta: str
        :param sobrescrever: Define se deve sobrescrever um arquivo existente no destino.
        :type sobrescrever: bool
        :raises FileNotFoundError: Se o arquivo de origem ou o diretório de destino não forem encontrados.
        """
        nome_arquivo = os.path.basename(origem)
        destino_arquivo = os.path.join(destino_pasta, nome_arquivo)

        if not self.arquivos_existem(origem):
            self.module_logger.warning("❌ Arquivo não encontrado, verifique se o caminho está correto.")
            raise FileNotFoundError("❌ Arquivo não encontrado, verifique se o caminho está correto.")

        if not os.path.isdir(destino_pasta):
            self.module_logger.warning("❌ Diretório de destino não encontrado.")
            raise FileNotFoundError("❌ Diretório de destino não encontrado.")

        if os.path.exists(destino_arquivo):
            if not sobrescrever:
                self.module_logger.warning("⚠️ Já existe um arquivo com o mesmo nome no destino.")
                raise FileExistsError("⚠️ Já existe um arquivo com o mesmo nome no destino.")
            else:
                self.module_logger.info("♻️ Arquivo existente será sobrescrito.")

        shutil.copy2(origem, destino_arquivo)
        self.module_logger.info(f"✅ Arquivo copiado com sucesso para {destino_arquivo}")
    
    def detectar_encoding(self, arquivo_csv, num_bytes=1024):
        """
        Detecta o encoding (codificação de caracteres) de um arquivo de texto ou CSV.

        Esta função utiliza a biblioteca `chardet` para analisar os primeiros bytes do arquivo
        e estimar a codificação mais provável (como 'utf-8', 'ISO-8859-1', 'cp1252', etc.).
        É útil para evitar erros de decodificação ao ler arquivos com `pandas` ou `open()`.

        :param arquivo_csv: Caminho para o arquivo que será analisado.
        :type arquivo_csv: str
        :param num_bytes: Quantidade de bytes lidos do início do arquivo para análise (padrão: 1024).
        :type num_bytes: int

        :return: O encoding detectado (ex: 'utf-8', 'latin1').
        :rtype: str
        """
        with open(arquivo_csv, 'rb') as f:
            resultado = chardet.detect(f.read(num_bytes))
            return resultado['encoding']

    def ler_csv(self, arquivo_csv,
                linha_cabecalho: Union[int, None] = 0,
                linha_inicio=0,
                colunas_desejadas=None,
                tipo_colunas=None,
                delimiter=';',
                quantidade_linhas_lidas=None,
                strict: bool = True,
                encoding=None,):
        """
        Lê um CSV usando pandas com suporte a encoding, linha inicial, colunas filtradas,
        delimitador customizado e quantidade limitada de linhas.

        :param arquivo_csv: Caminho do CSV.
        :type arquivo_csv: str
        :param linha_cabecalho: Número da linha usada como cabeçalho (default=0).
        :type linha_cabecalho: int or None
        :param linha_inicio: Quantidade de linhas para 'pular'.
        :type linha_inicio: int
        :param colunas_desejadas: Lista de colunas a manter. Se None, retorna todas.
        :type colunas_desejadas: list or None
        :param tipo_colunas: Dict mapeando colunas para tipos desejados (dtype do pandas).
        :type tipo_colunas: dict
        :param delimiter: Separador do CSV. Padrão: ';'.
        :type delimiter: str
        :param quantidade_linhas_lidas: Número máximo de linhas de dados a ler (ignora após esse limite).
        :type quantidade_linhas_lidas: int or None
        :param strict: Se True, lança erro caso colunas_desejadas não existam.
        :type strict: bool
        :param encoding: Codificação de caracteres.
        :type encoding: str

        :return: Lista de dicionários (cada linha como dict).
        :rtype: list
        """
        if linha_inicio < 0:
            raise ValueError("⚠️ 'linha_inicio' deve ser maior ou igual 0.")
        if quantidade_linhas_lidas is not None and quantidade_linhas_lidas <= 0:
            raise ValueError("⚠️ 'quantidade_linhas_lidas' deve ser maior que 0.")

        try:
            # Pula linhas apenas depois do header
            if linha_cabecalho is None:
                skiprows = linha_inicio if linha_inicio > 0 else None
            else:
                skiprows = range(linha_cabecalho + 1, linha_cabecalho + 1 + linha_inicio) if linha_inicio > 0 else None

            if not encoding:
                encoding = self.detectar_encoding(arquivo_csv)

            # pandas assume que a primeira linha tem os nomes das colunas
            df = pd.read_csv(arquivo_csv,
                             delimiter=delimiter,
                             header=linha_cabecalho,
                             dtype=tipo_colunas,
                             skiprows=skiprows,  # Pula as primeiras n linhas do arquivo ao carregá-lo.
                            #  decimal=","
                             nrows=quantidade_linhas_lidas,
                             encoding=encoding,
                             )

            # Remove linhas onde todos os valores são NaN
            df = df.dropna(how='all')

            # substituindo valores None para Nan e em seguida substituindo Nan para vazio
            df = df.replace([None], np.nan).fillna("")

            # print(df.head(5))
            # print(df.shape[0])  # quantidade de linhas (sem cabeçalho)
            # print(df.shape[1])  # quantidade de colunas
            
            if colunas_desejadas:
                colunas_invalidas = [c for c in colunas_desejadas if c not in df.columns]
                if colunas_invalidas:
                    msg = f"⚠️ Colunas não encontradas no CSV: {colunas_invalidas}"
                    self.module_logger.warning(msg)
                    if strict:
                        raise KeyError(msg)
                    else:
                        colunas_desejadas = [c for c in colunas_desejadas if c in df.columns]
                df = df[colunas_desejadas]
            
            self.module_logger.info(f"CSV carregado com {df.shape[0]} linhas e {df.shape[1]} colunas.")

            # print(df.head(5).to_dict(orient='dict'))  # {'CONTRATO': [...], 'MATRICULA': [...]}
            # print(df.head(5).to_dict(orient='list'))  # mesmo que acima, mas com listas ordenadas
            # print(df.head(5).to_dict(orient='series'))  # valores como pd.Series
            return df.to_dict(orient='records')

        except FileNotFoundError:
            erro = f"❌ Arquivo '{arquivo_csv}' não foi encontrado!"
            self.module_logger.error(erro)
            raise
        except Exception as e:
            self.module_logger.error(f"❌ Erro ao ler CSV '{arquivo_csv}': {e}")
            raise
    
    def obter_arquivos(self, diretorio: str, extensao: Optional[str] = None) -> List[str]:
        """
        Lista caminhos absolutos de arquivos diretamente contidos no diretório especificado,
        com suporte a filtro por extensão (sem buscar em subpastas).

        :param diretorio: Caminho do diretório base.
        :type diretorio: str
        :param extensao: (Optional) Ex.: ".txt" para filtrar arquivos por extensão.
        :type extensao: str

        :return: Lista de caminhos absolutos dos arquivos encontrados.
        :rtype: List[str]

        :raises: ValueError Se o caminho não existir ou não for um diretório.
        """
        path = Path(diretorio)

        self.module_logger.info(f"Verificando diretório: {diretorio}")
        if not path.exists():
            self.module_logger.info(f"⚠️ O diretório '{diretorio}' não existe.")
            raise ValueError(f"⚠️ O diretório '{diretorio}' não existe.")
        if not path.is_dir():
            self.module_logger.info(f"⚠️ O caminho '{diretorio}' não é um diretório.")
            raise ValueError(f"⚠️ O caminho '{diretorio}' não é um diretório.")

        self.module_logger.info("Iniciando varredura...")
        arquivos_encontrados = []

        try:
            # for arquivo in path.rglob('*'):  # obtem tudo inclusive subpastas
            for arquivo in path.iterdir():  # somente o conteúdo direto do diretório
                if not arquivo.is_file():
                    continue
                if extensao and arquivo.suffix.lower() != extensao.lower():
                    continue
                try:
                    caminho_resolvido = str(arquivo.resolve(strict=False))
                    arquivos_encontrados.append(caminho_resolvido)
                    self.module_logger.info(f"Arquivo encontrado: {caminho_resolvido}")
                except Exception as e:
                    self.module_logger.error(f"❌ Erro ao resolver caminho de '{arquivo}': {e}")
        except Exception as e:
            self.module_logger.error(f"❌ Erro durante a varredura de arquivos: {e}.")
            raise Exception(f"❌ Erro durante a varredura de arquivos: {e}")

        self.module_logger.info(f"Total de arquivos encontrados: {len(arquivos_encontrados)}")
        return arquivos_encontrados

    def ler_txt(self, arquivo_txt, encoding='utf-8', remover_quebra_linha=True):
        """
        Lê um arquivo .txt e retorna uma lista com as linhas do arquivo.

        :param caminho_arquivo_txt: Caminho completo ou relativo para o arquivo .txt
        :type caminho_arquivo_txt: str
        :param encoding: Codificação do arquivo (default: utf-8)
        :type encoding: str
        :param remover_quebra_linha: Se True, remove os \n das linhas
        :type remover_quebra_linha: bool

        :return: Lista de linhas (strings)
        :rtype: list

        :raises: IOError em caso de erro de leitura
        """
        encoding = self.detectar_encoding(arquivo_txt)

        try:
            with open(arquivo_txt, 'r', encoding=encoding) as arquivo:
                if remover_quebra_linha:
                    return [linha.rstrip('\n') for linha in arquivo]
                else:
                    return arquivo.readlines()
        except Exception as e:
            self.module_logger.error(f"❌ Erro ao ler o arquivo '{arquivo_txt}': {e}")
            raise IOError(f"❌ Erro ao ler o arquivo '{arquivo_txt}': {e}")

    def obter_infomacoes_arquivo(self, caminho_arquivo):
        """
        Retorna informações detalhadas sobre um arquivo.
        
        :param caminho_arquivo: Caminho completo do arquivo
        :type caminho_arquivo: str

        :return: dicionário com informações extraídas
        :rtype: dict

        :raises FileNotFoundError: se o arquivo não existir
        """
        caminho = Path(caminho_arquivo)

        if not caminho.is_file():
            self.module_logger.error(f"❌ Arquivo não encontrado: {caminho_arquivo}")
            raise FileNotFoundError(f"❌ Arquivo não encontrado: {caminho_arquivo}")

        stat = caminho.stat()

        return {
            "data_criacao": datetime.datetime.fromtimestamp(stat.st_ctime),
            "data_ultima_gravacao": datetime.datetime.fromtimestamp(stat.st_mtime),
            "nome_com_extensao": caminho.name,
            "nome_sem_extensao": caminho.stem,
            "nome_completo": str(caminho.resolve()),
            "extensao": caminho.suffix.lstrip('.'),
            "diretorio": str(caminho.parent),
            "tamanho_bytes": stat.st_size,
        }

    def alterar_valor_celula_vazia_xlsx(self, caminho_arquivo, dic_coluna_valor):
        """
        Altera a celula de determinada coluna quando o valor é vazio
        
        :param caminho_arquivo: Caminho completo do arquivo
        :type caminho_arquivo: str

        :param dic_coluna_valor: Dicionario com a coluna que se deseja alterar como chave e o valor que se deseja adicionar como valor
        :type dic_coluna_valor: dict
        """
        try:
            # Carregar o arquivo Excel
            df = pd.read_excel(caminho_arquivo)

            # Substituir valores vazios na coluna
            for coluna, valor in dic_coluna_valor.items():
                df[coluna] = df[coluna].fillna(valor)

            # Substituir valores 2 por 'LANÇADO' na coluna 'Status'
            # df['Status'] = df['Status'].replace('LANÇADO', 'TESTE')

            # Salvar o resultado em um novo arquivo
            df.to_excel(caminho_arquivo, index=False)
            self.module_logger.error(f"✅ Valores das colunas {dic_coluna_valor.keys()} alteradas")
        except Exception as erro:
            self.module_logger.error(f"❌ Erro ao alterar o valor das colunas vazia: {erro}")
            raise Exception(f"❌ Erro ao alterar o valor da coluna vazia: {erro}")
    
    def alterar_valor_celula_xlsx_com_condicao(self, caminho_arquivo, list_coluna_valores):
        """
        Altera o valor da celula se a propria tiver um valor específico.

        :param caminho_arquivo: Caminho completo do arquivo
        :type caminho_arquivo: str

        :param list_coluna_valores: Uma lista com os 3 item obrigatorio [coluna, valor_atual, valor_alterado] respctivamente nessa ordem.
        :type list_coluna_valores: list
        """
        try:
            # Carregar o arquivo Excel
            df = pd.read_excel(caminho_arquivo)

            coluna = list_coluna_valores[0]
            valor_atual = str(list_coluna_valores[1])
            valor_alterado = str(list_coluna_valores[2])

            # Substituir valores
            df[coluna] = df[coluna].replace(valor_atual, valor_alterado)

            # Salvar o resultado em um novo arquivo
            df.to_excel(caminho_arquivo, index=False)
            self.module_logger.error(f"✅ Coluna: {coluna} | Valor atual: {valor_atual} | Alterado para: {valor_alterado}")
        except Exception as erro:
            self.module_logger.error(f"❌ Erro ao alterar o valor da coluna: {erro}")
            raise Exception(f"❌ Erro ao alterar o valor da coluna: {erro}")

    def ler_xlsx(self,
                arquivo_xlsx: str,
                sheet_name: str = 0,
                linha_cabecalho: Union[int, None] = 0,
                linha_inicio: int = 0,
                colunas_desejadas: Optional[List[str]] = None,
                tipo_colunas: Optional[dict] = None,
                quantidade_linhas_lidas: Optional[int] = None,
                strict: bool = True) -> List[Dict[str, Any]]:
        """
        Lê um arquivo Excel (.xlsx) usando pandas com suporte a filtro de colunas,
        tipos de dados, linha inicial e quantidade máxima de linhas.

        :param arquivo_xlsx: Caminho do arquivo Excel.
        :type arquivo_xlsx: str
        :param sheet_name: Nome ou índice da planilha (default=0, primeira planilha).
        :type sheet_name: str or int
        :param linha_cabecalho: Número da linha usada como cabeçalho (default=0).
        :type linha_cabecalho: int or None
        :param linha_inicio: Número de linhas a pular no início (default=0).
        :type linha_inicio: int
        :param colunas_desejadas: Lista de colunas a manter (ou None para manter todas).
        :type colunas_desejadas: list or None
        :param tipo_colunas: Dict mapeando colunas para tipos desejados (dtype do pandas).
        :type tipo_colunas: dict
        :param quantidade_linhas_lidas: Número máximo de linhas de dados a ler.
        :type quantidade_linhas_lidas: int
        :param strict: Se True, lança erro caso colunas_desejadas não existam.
        :type strict: bool

        :return: Lista de dicionários (cada linha como dict).
        :rtype: list
        """

        if linha_inicio < 0:
            raise ValueError("⚠️ 'linha_inicio' deve ser maior ou igual 0.")
        if quantidade_linhas_lidas is not None and quantidade_linhas_lidas <= 0:
            raise ValueError("⚠️ 'quantidade_linhas_lidas' deve ser maior que 0.")

        try:
            # Pula linhas apenas depois do header
            if linha_cabecalho is None:
                skiprows = linha_inicio if linha_inicio > 0 else None
            else:
                skiprows = range(linha_cabecalho + 1, linha_cabecalho + 1 + linha_inicio) if linha_inicio > 0 else None

            df = pd.read_excel(
                arquivo_xlsx,
                sheet_name=sheet_name,
                header=linha_cabecalho,
                dtype=tipo_colunas,
                skiprows=skiprows,
                nrows=quantidade_linhas_lidas,
                engine="openpyxl"  # engine padrão para .xlsx
            )

            df = df.dropna(how="all").fillna("")

            if colunas_desejadas:
                colunas_invalidas = [c for c in colunas_desejadas if c not in df.columns]
                if colunas_invalidas:
                    msg = f"⚠️ Colunas não encontradas no XLSX: {colunas_invalidas}"
                    self.module_logger.warning(msg)
                    if strict:
                        raise KeyError(msg)
                    else:
                        colunas_desejadas = [c for c in colunas_desejadas if c in df.columns]
                df = df[colunas_desejadas]

            self.module_logger.info(
                f"Excel carregado da planilha '{sheet_name}' com {df.shape[0]} linhas e {df.shape[1]} colunas."
            )

            return df.to_dict(orient="records")

        except FileNotFoundError:
            erro = f"❌ Arquivo '{arquivo_xlsx}' não foi encontrado!"
            self.module_logger.error(erro)
            raise

        except Exception as e:
            self.module_logger.error(f"❌ Erro ao ler XLSX '{arquivo_xlsx}': {e}")
            raise

    def converter_para_base64(self, file_path: str) -> str:
        """
        Converte um arquivo em uma string Base64.

        Esta função lê o conteúdo binário de um arquivo e o converte
        para uma representação Base64, útil para transmissão de dados
        binários em formatos de texto (como JSON ou XML).

        :param file_path: Caminho absoluto ou relativo do arquivo a ser convertido.
        :type file_path: str
        :return: Representação Base64 do conteúdo do arquivo.
        :rtype: str
        :raises FileNotFoundError: Se o arquivo especificado não existir.
        :raises IOError: Se ocorrer um erro ao ler o arquivo.
        """
        path = Path(file_path)
        
        if not path.is_file():
            raise FileNotFoundError(f"⚠️  Arquivo não encontrado: {file_path}")

        try:
            with path.open("rb") as file:
                encoded_bytes = base64.b64encode(file.read())
                self.module_logger.info(f"✅  Arquivo convertido para base64")
                return encoded_bytes.decode("utf-8")
        except IOError as e:
            raise IOError(f"❌  Erro ao ler o arquivo '{file_path}': {e}")



class Diretorios():
    """Manipula diretórios."""

    def __init__(self):
        obj_arq = Arquivos()
        self.module_logger = obj_arq.criar_logger_diario("./log", nome_arquivo_log="manipula_diretorio_arquivos.log", logger=logging.getLogger(__name__))

    def diretorio_existe(self, caminho):
        """
        Verifica se um diretório existe no caminho especificado.

        :param caminho: Caminho do diretório a ser verificado.
        :type caminho: str

        :returns: True se o diretório existe, False caso contrário.
        :rtype: bool
        """
        return os.path.isdir(caminho)

    def criar_diretorio(self, caminho):
        """
        Cria o diretório especificado caso ele não exista.

        :param caminho: Caminho do diretório a ser verificado ou criado.
        :type caminho: str

        :returns: True se o diretório foi criado, False caso contrário.
        :rtype: bool
        """
        try:
            if not self.diretorio_existe(caminho):
                os.makedirs(caminho)
                # print(f"✅ Diretório criado: {caminho}")
                self.module_logger.info(f"✅ Diretório criado: {caminho}")
                return True
            else:
                # print(f"ℹ️ Diretório já existe: {caminho}")
                self.module_logger.warning(f"ℹ️ Diretório já existe: {caminho}")
                return False
        except OSError as erro:
            self.module_logger.warning(f"❌ {erro}")
            raise OSError(f"❌ {erro}")
        except Exception as erro:
            self.module_logger.warning(f"❌ {erro}")
            raise Exception(f"❌ {erro}")

    def limpar_diretorio(self, caminho: Union[str, Path]) -> bool:
        """
        Remove todos os arquivos e subpastas dentro do diretório informado,
        mantendo o diretório principal.

        :param caminho: Caminho para o diretório que será limpo.
        :type caminho: str

        :return: True se tudo ocorreu bem, False caso contrário.
        :rtype: bool
        """
        try:
            pasta = Path(caminho)

            if not pasta.exists():
                self.module_logger.warning(f"⚠️  Diretório não existe: {pasta}")
                return False

            if not pasta.is_dir():
                self.module_logger.warning(f"⚠️  O caminho informado não é um diretório: {pasta}")
                return False

            for item in pasta.iterdir():
                try:
                    if item.is_file() or item.is_symlink():
                        item.unlink()  # remove arquivo
                    elif item.is_dir():
                        shutil.rmtree(item)  # remove pasta recursivamente

                except Exception as e:
                    self.module_logger.error(f"❌ Erro ao remover '{item}': {e}")

            self.module_logger.info(f"✅ Diretório limpo com sucesso: {pasta}")
            return True

        except Exception as erro:
            self.module_logger.error(f"❌ Erro inesperado ao limpar diretório: {erro}")
            return False
