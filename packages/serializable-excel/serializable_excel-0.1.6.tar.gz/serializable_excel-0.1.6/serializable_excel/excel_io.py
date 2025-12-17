"""
Functions for reading and writing Excel files using openpyxl.
"""

from io import BytesIO
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook, load_workbook
try:  # разные версии openpyxl по-разному экспортируют классы ячеек
    from openpyxl.cell.cell import EmptyCell, MergedCell  # type: ignore
except ImportError:  # fallback для старых/новых версий
    from openpyxl.cell.cell import MergedCell  # type: ignore

    class EmptyCell:  # minimal stub for isinstance checks
        pass
from openpyxl.worksheet.worksheet import Worksheet

from serializable_excel.colors import CellStyle, CellStyleApplier

if TYPE_CHECKING:
    from serializable_excel.excel_types import ExcelType


def read_excel_headers(worksheet: Worksheet, header_row: int = 1) -> Dict[int, str]:
    """
    Read column headers from an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        header_row: Row number containing headers (1-indexed)

    Returns:
        Dictionary mapping column index (1-indexed) to header name
    """
    headers = {}
    for cell in worksheet[header_row]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()
    return headers


def read_excel_rows(
    worksheet: Worksheet, start_row: int = 2, max_row: Optional[int] = None
) -> List[Dict[int, Any]]:
    """
    Read data rows from an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        start_row: First row to read (1-indexed, typically 2 to skip header)
        max_row: Last row to read (None for all rows)

    Returns:
        List of dictionaries mapping column index (1-indexed) to cell value
    """
    rows = []
    end_row = max_row if max_row is not None else worksheet.max_row

    for row_num in range(start_row, end_row + 1):
        row_data = {}
        row = worksheet[row_num]
        for cell in row:
            # EmptyCell/MergedCell в openpyxl 3.1 могут не иметь .column
            if isinstance(cell, (EmptyCell, MergedCell)):
                continue
            col = getattr(cell, "column", None)
            if col is None:
                continue
            if col in row_data or cell.value is None:
                continue
            row_data[col] = cell.value
        # Only add non-empty rows
        if row_data:
            rows.append(row_data)

    return rows


def load_workbook_from_source(
    source: Union[str, bytes, BytesIO],
    read_only: bool = True,
    data_only: bool = True,
) -> Workbook:
    """
    Load a workbook from a file path, bytes, or BytesIO.

    Args:
        source: File path, bytes, or BytesIO object
        read_only: Open workbook in read-only mode
        data_only: Load only cell values, not formulas

    Returns:
        openpyxl Workbook object
    """
    if isinstance(source, bytes):
        source = BytesIO(source)
    return load_workbook(source, read_only=read_only, data_only=data_only)


def _build_workbook(
    headers: Dict[str, int],
    data_rows: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
    cell_colors: Optional[Dict[Tuple[int, int], CellStyle]] = None,
    column_types: Optional[Dict[str, "ExcelType"]] = None,
) -> Workbook:
    """
    Build a workbook with headers and data.

    Args:
        headers: Dictionary mapping header names to column indices (1-indexed)
        data_rows: List of dictionaries mapping header names to values
        sheet_name: Name of the worksheet
        cell_colors: Dictionary mapping (row, col) tuples to CellStyle objects
        column_types: Dictionary mapping header names to ExcelType objects

    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for header, col_idx in headers.items():
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(data_rows, start=2):
        for header, col_idx in headers.items():
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col_idx)

            # Apply Excel type formatting if specified (sets value and format)
            if column_types is not None and header in column_types:
                excel_type = column_types[header]
                if excel_type is not None:
                    excel_type.apply_format(cell, value)
                else:
                    cell.value = value
            else:
                cell.value = value

            # Apply cell color if specified
            if cell_colors is not None:
                style = cell_colors.get((row_idx, col_idx))
                if style is not None:
                    CellStyleApplier.apply(cell, style)

    return wb


def write_excel(
    headers: Dict[str, int],
    data_rows: List[Dict[str, Any]],
    file_path: str,
    sheet_name: str = "Sheet1",
    cell_colors: Optional[Dict[Tuple[int, int], CellStyle]] = None,
    column_types: Optional[Dict[str, "ExcelType"]] = None,
) -> None:
    """
    Write data to an Excel file.

    Args:
        headers: Dictionary mapping header names to column indices (1-indexed)
        data_rows: List of dictionaries mapping header names to values
        file_path: Path where to save the Excel file
        sheet_name: Name of the worksheet
        cell_colors: Dictionary mapping (row, col) tuples to CellStyle objects
        column_types: Dictionary mapping header names to ExcelType objects
    """
    wb = _build_workbook(headers, data_rows, sheet_name, cell_colors, column_types)
    wb.save(file_path)


def write_excel_to_bytes(
    headers: Dict[str, int],
    data_rows: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
    cell_colors: Optional[Dict[Tuple[int, int], CellStyle]] = None,
    column_types: Optional[Dict[str, "ExcelType"]] = None,
) -> bytes:
    """
    Write data to Excel format and return as bytes.

    Args:
        headers: Dictionary mapping header names to column indices (1-indexed)
        data_rows: List of dictionaries mapping header names to values
        sheet_name: Name of the worksheet
        cell_colors: Dictionary mapping (row, col) tuples to CellStyle objects
        column_types: Dictionary mapping header names to ExcelType objects

    Returns:
        Excel file content as bytes
    """
    wb = _build_workbook(headers, data_rows, sheet_name, cell_colors, column_types)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
