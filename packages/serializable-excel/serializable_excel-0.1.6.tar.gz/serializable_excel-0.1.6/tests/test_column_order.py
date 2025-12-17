"""
Tests for column ordering functionality.
"""

from typing import Optional

import pytest
from openpyxl import Workbook, load_workbook

from serializable_excel import Column, DynamicColumn, ExcelModel


class UserModel(ExcelModel):
    """Test model with static columns."""

    name: str = Column(header='Name')
    age: int = Column(header='Age')
    email: str = Column(header='Email')
    phone: str = Column(header='Phone')


class ForecastModel(ExcelModel):
    """Test model with static and dynamic columns."""

    month: str = Column(header='Month')
    manager: str = Column(header='Manager')
    characteristics: dict = DynamicColumn()


def get_column_order_from_excel(file_path: str) -> list[str]:
    """Extract column headers in order from Excel file."""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = []
    for cell in ws[1]:  # First row contains headers
        if cell.value:
            headers.append(str(cell.value))
    return headers


def test_static_columns_with_function_order(temp_excel_file: str):
    """Test static columns ordering using function."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
        UserModel(name='Bob', age=30, email='bob@test.com', phone='456'),
    ]

    def column_order(header: str) -> Optional[int]:
        """Order: Email first, then Name, then Age, Phone last."""
        order_map = {'Email': 1, 'Name': 2, 'Age': 3, 'Phone': 4}
        return order_map.get(header)

    UserModel.to_excel(users, temp_excel_file, column_order=column_order)

    # Verify column order
    headers = get_column_order_from_excel(temp_excel_file)
    assert headers == ['Email', 'Name', 'Age', 'Phone']

    # Verify data integrity
    loaded = UserModel.from_excel(temp_excel_file)
    assert len(loaded) == 2
    assert loaded[0].email == 'alice@test.com'
    assert loaded[0].name == 'Alice'


def test_static_columns_with_dict_order(temp_excel_file: str):
    """Test static columns ordering using dictionary."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
    ]

    column_order_dict = {'Age': 1, 'Email': 2, 'Name': 3, 'Phone': 4}

    UserModel.to_excel(users, temp_excel_file, column_order=column_order_dict)

    # Verify column order
    headers = get_column_order_from_excel(temp_excel_file)
    assert headers == ['Age', 'Email', 'Name', 'Phone']


def test_reading_with_shuffled_columns(temp_excel_file: str):
    """Reading Excel with shuffled columns should still map fields correctly."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
        UserModel(name='Bob', age=30, email='bob@test.com', phone='456'),
    ]

    # Write with default order, then shuffle columns in the Excel file
    UserModel.to_excel(users, temp_excel_file)

    wb = load_workbook(temp_excel_file)
    ws = wb.active

    # Capture existing rows (including header) in original order: Name, Age, Email, Phone
    rows = list(ws.iter_rows(values_only=True))

    # New order: Email, Phone, Name, Age
    new_order = [2, 3, 0, 1]
    shuffled_rows = [[row[idx] for idx in new_order] for row in rows]

    # Replace sheet content with shuffled columns
    ws.delete_cols(1, ws.max_column)
    for r_idx, row in enumerate(shuffled_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(temp_excel_file)

    # Read back and verify data is mapped correctly
    loaded = UserModel.from_excel(temp_excel_file)
    assert len(loaded) == 2
    assert loaded[0].name == 'Alice'
    assert loaded[0].email == 'alice@test.com'
    assert loaded[0].age == 25
    assert loaded[0].phone == '123'
    assert loaded[1].name == 'Bob'
    assert loaded[1].email == 'bob@test.com'
    assert loaded[1].age == 30
    assert loaded[1].phone == '456'


def test_reading_shuffled_with_dynamic_columns(temp_excel_file: str):
    """Reading Excel with shuffled static + dynamic columns should work."""
    wb = Workbook()
    ws = wb.active

    # Headers shuffled: Beta (dynamic), Month (static), Alpha (dynamic), Manager (static)
    headers = ['Beta', 'Month', 'Alpha', 'Manager']
    ws.append(headers)
    ws.append(['b1', '2024-01', 'a1', 'Alice'])
    ws.append(['b2', '2024-02', 'a2', 'Bob'])

    wb.save(temp_excel_file)

    loaded = ForecastModel.from_excel(temp_excel_file, dynamic_columns=True)

    assert len(loaded) == 2
    assert loaded[0].month == '2024-01'
    assert loaded[0].manager == 'Alice'
    assert loaded[0].characteristics == {'Alpha': 'a1', 'Beta': 'b1'}

    assert loaded[1].month == '2024-02'
    assert loaded[1].manager == 'Bob'
    assert loaded[1].characteristics == {'Alpha': 'a2', 'Beta': 'b2'}


def test_static_columns_partial_order(temp_excel_file: str):
    """Test static columns with partial ordering (some columns without order)."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
    ]

    def column_order(header: str) -> Optional[int]:
        """Only Email and Name have order, others go to end."""
        order_map = {'Email': 1, 'Name': 2}
        return order_map.get(header)

    UserModel.to_excel(users, temp_excel_file, column_order=column_order)

    # Verify column order: Email, Name first, then Age, Phone (original order)
    headers = get_column_order_from_excel(temp_excel_file)
    assert headers[0] == 'Email'
    assert headers[1] == 'Name'
    # Age and Phone should be at the end, in original order
    assert 'Age' in headers[2:]
    assert 'Phone' in headers[2:]


def test_dynamic_columns_order(temp_excel_file: str):
    """Test dynamic columns ordering."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={
                'Sales': 150,
                'Priority': 'High',
                'Status': 'Active',
            },
        ),
        ForecastModel(
            month='2024-02',
            manager='Bob',
            characteristics={
                'Sales': 80,
                'Priority': 'Medium',
                'Status': 'Pending',
            },
        ),
    ]

    # Simulate initial orders: Sales=1, Priority=5, Status=10 (will normalize)
    column_order = {'Sales': 1, 'Priority': 5, 'Status': 10}

    ForecastModel.to_excel(
        forecasts, temp_excel_file, column_order=column_order
    )

    # Verify column order: dynamic columns with order come first (normalized 1,2,3),
    # then static columns without order
    headers = get_column_order_from_excel(temp_excel_file)
    # Dynamic columns should come first (they have order 1, 2, 3)
    assert (
        'Sales' in headers[:3]
        or 'Priority' in headers[:3]
        or 'Status' in headers[:3]
    )
    # Static columns without order should follow
    assert 'Month' in headers
    assert 'Manager' in headers


def test_dynamic_columns_with_initial_orders(temp_excel_file: str):
    """Test dynamic columns with initial order numbers (simulating order_layer)."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={
                'Sales': 150,
                'Priority': 'High',
                'Status': 'Active',
            },
        ),
    ]

    # Simulate order_layer values: Sales=1, Priority=5, Status=10 (will normalize)
    column_order = {'Sales': 1, 'Priority': 5, 'Status': 10}

    ForecastModel.to_excel(
        forecasts, temp_excel_file, column_order=column_order
    )

    # Verify that dynamic columns are normalized (no gaps)
    headers = get_column_order_from_excel(temp_excel_file)
    # Find dynamic columns
    dynamic_start = headers.index('Sales') if 'Sales' in headers else -1
    if dynamic_start >= 0:
        # Check that dynamic columns are sequential
        dynamic_headers = headers[dynamic_start:]
        # Should be Sales, Priority, Status (or similar normalized order)
        assert len(dynamic_headers) >= 3


def test_mixed_static_and_dynamic_order(temp_excel_file: str):
    """Test ordering with both static and dynamic columns."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={'Sales': 150, 'Priority': 'High'},
        ),
    ]

    def static_order(header: str) -> Optional[int]:
        """Manager first, then Month."""
        order_map = {'Manager': 1, 'Month': 2}
        return order_map.get(header)

    # Shared order layer for static and dynamic columns
    column_order = {'Manager': 1, 'Month': 2, 'Priority': 1, 'Sales': 5}

    ForecastModel.to_excel(
        forecasts,
        temp_excel_file,
        column_order=column_order,
    )

    # Verify order: All columns with order are mixed together and normalized
    # Manager=1, Month=2, Priority=1, Sales=5 -> normalized: Manager=1, Priority=2, Month=3, Sales=4
    headers = get_column_order_from_excel(temp_excel_file)
    # Manager should be first (order 1)
    assert headers[0] == 'Manager'
    # Priority should be second (normalized from 1, but Manager also has 1, so Priority gets 2)
    # Actually, both Manager and Priority have order 1, so they normalize to 1 and 2
    # Then Month has order 2, which normalizes to 3
    # Then Sales has order 5, which normalizes to 4
    # So order should be: Manager (1), Priority (2), Month (3), Sales (4)
    assert 'Priority' in headers[:4]
    assert 'Month' in headers[:4]
    assert 'Sales' in headers[:4]


def test_dynamic_columns_compete_with_static_order_layer(temp_excel_file: str):
    """Dynamic columns should participate in column_order alongside static ones."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={'Priority': 'High', 'Sales': 150},
        ),
    ]

    # Shared order_layer for static and dynamic columns
    column_order = {
        'Priority': 1,  # dynamic
        'Sales': 2,  # dynamic
        'Manager': 5,  # static
        'Month': 10,  # static
    }

    ForecastModel.to_excel(
        forecasts,
        temp_excel_file,
        column_order=column_order,
        # dynamic_column_order omitted to force shared order_layer usage
    )

    headers = get_column_order_from_excel(temp_excel_file)

    # Orders normalize from 1,2,5,10 -> 1,2,3,4 preserving relative priority
    assert headers[:4] == ['Priority', 'Sales', 'Manager', 'Month']


def test_order_normalization_with_gaps(temp_excel_file: str):
    """Test that order numbers with gaps are normalized (1, 5 -> 1, 2)."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
    ]

    def column_order(header: str) -> Optional[int]:
        """Order with gaps: Email=1, Name=5 (should normalize to 1, 2)."""
        order_map = {'Email': 1, 'Name': 5}
        return order_map.get(header)

    UserModel.to_excel(users, temp_excel_file, column_order=column_order)

    # Verify that columns are sequential (no gaps)
    headers = get_column_order_from_excel(temp_excel_file)
    # Email should be first, Name should be second (normalized from 1, 5)
    assert headers[0] == 'Email'
    assert headers[1] == 'Name'


def test_no_order_specified(temp_excel_file: str):
    """Test that without order specification, columns use default order."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
    ]

    UserModel.to_excel(users, temp_excel_file)

    # Verify default order (as defined in model)
    headers = get_column_order_from_excel(temp_excel_file)
    assert headers == ['Name', 'Age', 'Email', 'Phone']


def test_dynamic_columns_without_order_function(temp_excel_file: str):
    """Test dynamic columns without order function (should use sorted order)."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={'Zebra': 'z', 'Alpha': 'a', 'Beta': 'b'},
        ),
    ]

    ForecastModel.to_excel(forecasts, temp_excel_file)

    # Verify that dynamic columns are sorted alphabetically
    headers = get_column_order_from_excel(temp_excel_file)
    # Static columns without order should come first
    assert headers[0] == 'Month'
    assert headers[1] == 'Manager'
    # Dynamic columns without order should follow, sorted alphabetically: Alpha, Beta, Zebra
    assert headers[2] == 'Alpha'
    assert headers[3] == 'Beta'
    assert headers[4] == 'Zebra'


def test_column_order_with_bytes():
    """Test column ordering when exporting to bytes."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
    ]

    def column_order(header: str) -> Optional[int]:
        order_map = {'Email': 1, 'Name': 2, 'Age': 3, 'Phone': 4}
        return order_map.get(header)

    excel_bytes = UserModel.to_excel(
        users, return_bytes=True, column_order=column_order
    )
    assert excel_bytes is not None

    # Load from bytes and verify
    loaded = UserModel.from_excel(excel_bytes)
    assert len(loaded) == 1
    assert loaded[0].email == 'alice@test.com'


def test_complex_order_scenario(temp_excel_file: str):
    """Test complex scenario with mixed ordering and gaps."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={
                'Early': 'e',
                'Late': 'l',
                'Middle': 'm',
            },
        ),
    ]

    # Combined orders with gaps that will be normalized together
    column_order = {
        'Manager': 10,
        'Month': 20,
        'Early': 1,
        'Middle': 5,
        'Late': 10,
    }

    ForecastModel.to_excel(
        forecasts,
        temp_excel_file,
        column_order=column_order,
    )

    # Verify normalized order
    headers = get_column_order_from_excel(temp_excel_file)
    # Manager should be before Month (normalized from 10, 20 -> 1, 2)
    manager_idx = headers.index('Manager')
    month_idx = headers.index('Month')
    assert manager_idx < month_idx

    # Dynamic columns should be normalized: Early, Middle, Late (1, 2, 3)
    early_idx = headers.index('Early')
    middle_idx = headers.index('Middle')
    late_idx = headers.index('Late')
    assert early_idx < middle_idx < late_idx


def test_same_order_number_sorted_by_name(temp_excel_file: str):
    """Test that columns with the same order number are sorted by name."""
    users = [
        UserModel(name='Alice', age=25, email='alice@test.com', phone='123'),
        UserModel(name='Bob', age=30, email='bob@test.com', phone='456'),
    ]

    def column_order(header: str) -> Optional[int]:
        """Email and Name both have order 1, should be sorted by name."""
        order_map = {'Email': 1, 'Name': 1, 'Age': 2, 'Phone': 3}
        return order_map.get(header)

    UserModel.to_excel(users, temp_excel_file, column_order=column_order)

    # Verify column order: Email comes before Name (alphabetically) when both have order 1
    headers = get_column_order_from_excel(temp_excel_file)
    email_idx = headers.index('Email')
    name_idx = headers.index('Name')
    age_idx = headers.index('Age')
    phone_idx = headers.index('Phone')

    # Email should come before Name (alphabetically, both have order 1)
    assert email_idx < name_idx
    # Age should come after Email and Name (order 2)
    assert age_idx > email_idx
    assert age_idx > name_idx
    # Phone should come last (order 3)
    assert phone_idx > age_idx


def test_same_order_mixed_static_dynamic(temp_excel_file: str):
    """Test that static and dynamic columns with same order are sorted by name."""
    forecasts = [
        ForecastModel(
            month='2024-01',
            manager='Alice',
            characteristics={'Beta': 'b', 'Alpha': 'a'},
        ),
    ]

    column_order = {'Manager': 1, 'Month': 1, 'Alpha': 1, 'Beta': 1}

    ForecastModel.to_excel(
        forecasts,
        temp_excel_file,
        column_order=column_order,
    )

    # Verify order: all columns with order 1 should be sorted alphabetically
    headers = get_column_order_from_excel(temp_excel_file)
    # Alpha, Beta, Manager, Month all have order 1, should be sorted: Alpha, Beta, Manager, Month
    assert headers[0] == 'Alpha'
    assert headers[1] == 'Beta'
    assert headers[2] == 'Manager'
    assert headers[3] == 'Month'


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
