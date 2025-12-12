import os
import json
import time
import random
import dateutil.parser
from datetime import datetime
import requests
import logging
logger = logging.getLogger(__name__)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import (
    DoughnutChart,
    ScatterChart,
    BarChart,
    Reference,
    Series,
    LineChart,
    PieChart
)
from openpyxl.chart.series import DataPoint
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from decimal import Decimal
from fractions import Fraction

from analytics_sdk.utilities import (
    BASE_API_URL,
    APP_ID,
    API_KEY,
    API_SECRET,
    API_RETRY,
    API_TIME_STACKS,
    APP_DISPLAY_NAME
)

chart_colors = ["0077C8", "e35420", "32a3df", "673ab7", "9c27b0", "F5BE0A","e30255","a002e3","0079e3","05e3a8","f5d255","f36686","c661e9","24a5ed","62ecbf","F7E487","fb9fb5","e39df0","79ccef","91F4D5","f8ba7e","f3a5ea","cd91eb","87E4F0","cdf38d","f78f4a","eb6dd6","9352e5","5ad1e9","afeb5a","f55a22","e302be","6834E3","02bee3","8ae305"]

class ExcelRenderer:
    ROW_START = 3
    global res
    res = ''
    def __init__(self, analysis_run, tenantId, additional_props, excel_data, report_gen_start_time, report_gen_completed_time):
        """
        :param analysis_run: AnalysisRun
        """
        self.wb = Workbook()
        self.analysis_run = analysis_run
        self.tenantId = tenantId
        self.excel_data = excel_data
        self.report_gen_start_time = report_gen_start_time
        self.report_gen_completed_time = report_gen_completed_time
        self.additional_props = additional_props

        def get_token():
            headers = {"Content-Type" : "application/x-www-form-urlencoded" , "Accept" : "application/json"};
            post_data = {"grant_type": "client_credentials", "client_id" : API_KEY, "client_secret" : API_SECRET};
            token_url = BASE_API_URL + "/tenancy/auth/oauth/token";
            response = requests.post(token_url,data=post_data,headers=headers,verify=False);
            json = response.json();
            auth = str(json["access_token"])
            return auth


        def get_headers():
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'Authorization': f'Bearer {get_token()}'
            }

            return headers

        def call_requests(method, url, params=None, data=None, json=None, verify=True):
            headers = get_headers()
            retry = 1
            resp = None
            while retry <= API_RETRY:
                try:
                    resp = requests.request(method, url, params=params, data=data, json=json, headers=headers, verify=verify)
                    if not resp.ok:
                        time.sleep(retry * 2)
                        retry+=1
                        continue
                except requests.exceptions.ConnectionError:
                    time.sleep(retry * 2)
                    retry+=1
                    continue

                return resp

            return resp


        def diff_sec(st, et):
            difference = int(et - st)
            return difference

        url = BASE_API_URL + f'/reporting/api/v3/tenants/{tenantId}/runs/{analysis_run}/summary'
        api_proce_before_time = time.time()
        res = call_requests('GET', url)
        if res == None or not res.ok:
            logger.info('Get excel summary is failed')
            raise Exception("Get excel summary API FAILED", res)
        else:
            logger.info('Get excel summary API response is %s', res)
        api_proce_after_time = time.time()
        api_proce_diff = diff_sec(api_proce_before_time, api_proce_after_time)
        if api_proce_diff > API_TIME_STACKS:
            logging.info('Get excel summary API response took %d (greater than %d) seconds', api_proce_diff, API_TIME_STACKS)

        self.res = res

    def get_value(self, value):
        if value:
            try:
                value = float(value)
                return value
            except ValueError:
                return value
        return value
    

    def _load_data(self, ws, data, row_start, col_start, adjust_column=True, x_axis_date_format=None):
        column_widths = {}
        if isinstance(data, list):
            if len(data) >= 1000005:
                data = data[0:1000005]

        for row_delta, row in enumerate(data):
            for col_delta, val in enumerate(row):
                col = col_start + col_delta
                cell = ws.cell(row=row_start+row_delta, column=col)
                if x_axis_date_format and row_delta > 0 and col_delta == 0:
                    cell.value = datetime.strptime(val, x_axis_date_format).date()
                else:
                    try:
                        if APP_ID == 'AVAILABILITY-DETAILS':
                            if isinstance(val, str) and 'COLOR-CODE_' in val:
                                vals = val.split('_') #COLOR-CODE_{color_code}_{value}
                                if len(vals) == 3:
                                    val = vals[2]
                                    if vals[1] is not None and len(vals[1]):
                                        cell.fill = PatternFill("solid", fgColor=vals[1])
                        cell.value = self.get_value(val)
                    except Exception as e:
                        if isinstance(val, str):
                            cell.value = ILLEGAL_CHARACTERS_RE.sub(r'', val)
                        else:
                            cell.value = self.get_value(val)
                if col in column_widths:
                    column_widths[col] = max(len(str(val)), column_widths[col])
                else:
                    column_widths[col] = max(len(str(val)), ws.column_dimensions[get_column_letter(col)].width)
        # handle cell width
        if adjust_column:
            for col, column_width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = column_width + 2

    def add_table(self, ws, table_data):
        """
        add table component
        """
        row_start = ExcelRenderer.ROW_START + table_data['start-row'] + 1
        col_start = table_data['start-col']
        self._load_data(ws, table_data['data'], row_start, col_start)
        row_span = len(table_data['data'])
        col_span = len(table_data['data'][0])

        ref = f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start+col_span-1)}{row_start+row_span-1}'
        ts = time.time()
        table_name = f'Table{random.randint(0, 100)}_{ts}'

        table = Table(displayName=table_name, ref=ref, tableStyleInfo=TableStyleInfo(name="TableStyleMedium9"))
        ws.add_table(table)


    def add_metric_table(self, ws, table_data):
        """
        add metric table component
        """
        row_start = ExcelRenderer.ROW_START + table_data['start-row'] + 1
        col_start = table_data['start-col']
        self._load_data(ws, table_data['data'], row_start, col_start)
        row_span = len(table_data['data'])
        col_span = len(table_data['data'][0])
        if 'merge_cells' in table_data:
            for merge_cell in table_data['merge_cells']:
                ws.merge_cells(merge_cell)
        fill_color = 'eeeeee'
        if 'cell_color' in table_data:
            fill_color = table_data['cell_color']
        if 'color_cells' in table_data:
            min_row = table_data['color_cells'][0]
            max_row = table_data['color_cells'][1]
            min_col = table_data['color_cells'][2]
            max_col = table_data['color_cells'][3]

            for rows in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in rows:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.alignment = Alignment(horizontal="center")

        div_color = 'dee6ef'
        col_min_col = 13
        if 'divider-color' in table_data:
            div_color = table_data['divider-color']
        if 'col_min_col' in table_data:
            col_min_col = table_data['col_min_col']
        if 'col_divider' in table_data:
            if table_data['col_divider'] and table_data['col_divider'] is not None:
                col_min_row = table_data['col_divider'][0]
                col_max_row = table_data['col_divider'][1]
                col_length = table_data['col_divider'][2]
                count = 0
                for i in range(col_length):
                    if count > 0:
                        col_min_col = col_min_col+4
                    for rows in ws.iter_rows(min_row=col_min_row, max_row=col_max_row, min_col=col_min_col, max_col=col_min_col):
                        for cell in rows:
                            cell.fill = PatternFill(start_color=div_color, end_color=div_color, fill_type='solid')
                    count+=1

        ref = f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start+col_span-1)}{row_start+row_span-1}'
        ts = time.time()
        table_name = f'Table{random.randint(0, 100)}_{ts}'

        table = Table(displayName=table_name, ref=ref, tableStyleInfo=TableStyleInfo(name="TableStyleMedium9"))
        ws.add_table(table)


    def add_doughnut_chart(self, ws, chart_data):
        """
        add doughnut chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        hole_size = 50
        if 'hole-size' in chart_data:
            hole_size = chart_data.get('hole-size')

        chart = DoughnutChart(holeSize=hole_size)
        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = chart_data['chart-title']
        chart.style = 12
        chart.width = chart_data.get('width', 11)
        chart.height = chart_data.get('height', 5)

        ##Set the DataLabelList to show percentage of pie chart labels.
        #chart.dataLabels = DataLabelList()
        #chart.dataLabels.showVal = False
        #chart.dataLabels.showCatName = False
        #chart.dataLabels.showLegendKey = False

        slices = [DataPoint(idx=i) for i in range(row_span-1)]
        chart.series[0].data_points = slices

        #colors = ["0077C8", "32a3df", "673ab7", "9c27b0"]
        chart_colors = self.generate_pie_chart_colors(len(slices))
        colors = chart_colors

        if 'colors' in chart_data and chart_data['colors']:
            colors = chart_data['colors']+colors

        for idx, slice in enumerate(slices):
            #slice.graphicalProperties.solidFill = colors[idx % 4]
            slice.graphicalProperties.solidFill = colors[idx % len(colors)]

        ws.add_chart(chart, chart_data['chart-position'])


    def add_pie_chart(self, ws, chart_data):
        """
        add doughnut chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        chart = PieChart()
        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = chart_data['chart-title']
        chart.style = 12
        chart.width = chart_data.get('width', 11)
        chart.height = chart_data.get('height', 5)

        ##Set the DataLabelList to show percentage of pie chart labels.
        #chart.dataLabels = DataLabelList()
        #chart.dataLabels.showVal = False
        #chart.dataLabels.showCatName = False
        #chart.dataLabels.showLegendKey = False

        slices = [DataPoint(idx=i) for i in range(row_span-1)]
        chart.series[0].data_points = slices

        #colors = ["0077C8", "32a3df", "673ab7", "9c27b0"]
        chart_colors = self.generate_pie_chart_colors(len(slices))
        colors = chart_colors

        if 'colors' in chart_data and chart_data['colors']:
            colors = chart_data['colors']+colors

        for idx, slice in enumerate(slices):
            #slice.graphicalProperties.solidFill = colors[idx % 4]
            slice.graphicalProperties.solidFill = colors[idx % len(colors)]

        ws.add_chart(chart, chart_data['chart-position'])


    def add_bar_chart(self, ws, chart_data):
        """
        add bar chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.style = 4
        chart.legend = None
        chart.width = chart_data.get('width', 17)
        chart.height = chart_data.get('height', 8.5)

        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "0077C8"
        s1.graphicalProperties.solidFill = "0077C8"

        multicolor_check = False
        multi_colors = chart_colors
        if 'is-multicolored-chart' in chart_data:
            multicolor_check = chart_data['is-multicolored-chart']
        if 'colors' in chart_data and chart_data['colors']:
            multi_colors = chart_data['colors']

        for idx, s in enumerate(chart.series):
            if multicolor_check:
                for i in range(len(multi_colors)):
                    pt = DataPoint(idx=i)
                    pt.graphicalProperties.solidFill = multi_colors[i]
                    s.dPt.append(pt)
            else:
                index = len(chart_colors)%(idx+1)
                if index >= len(chart_colors):
                    index = 0
                s.graphicalProperties.line.solidFill = chart_colors[index]
                s.graphicalProperties.solidFill = chart_colors[index]

        ws.add_chart(chart, chart_data['chart-position'])

    def add_stack_bar_chart(self, ws, chart_data):
        """
        add stack bar chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])
        no_of_labels = chart_data['no-of-labels']
        if not no_of_labels:
            no_of_labels = 1

        chart = BarChart()
        chart.type = "col"
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1, max_col=col_start+no_of_labels)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.style = 4
        if no_of_labels == 1:
            chart.legend = None
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.width = chart_data.get('width', 20)
        chart.height = chart_data.get('height', 10)

        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "0077C8"
        s1.graphicalProperties.solidFill = "0077C8"

        add_chart_colors = chart_colors
        colors_check = False
        if 'colors' in chart_data and chart_data['colors']:
            add_chart_colors = chart_data['colors']+add_chart_colors
            colors_check = True

        for idx, s in enumerate(chart.series):
            index = idx
            if not colors_check:
                index = len(add_chart_colors)%(idx+1)
            if index >= len(add_chart_colors):
                index = 0
            s.graphicalProperties.line.solidFill = add_chart_colors[index]
            s.graphicalProperties.solidFill = add_chart_colors[index]

        ws.add_chart(chart, chart_data['chart-position'])

    def add_scatter_chart(self, ws, chart_data):
        """
        add scatter chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        x_axis_date_format = chart_data.get('x-axis-date-format')
        self._load_data(ws, chart_data['data'], row_start, col_start, False, x_axis_date_format)
        row_span = len(chart_data['data'])
        no_of_lines = chart_data.get('no-of-lines')
        if not no_of_lines:
            no_of_lines = 1
        colors = ["0077C8", "e35420", "32a3df", "673ab7", "9c27b0"]

        chart = ScatterChart()
        chart.style = 5
        chart.legend = None
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')
        chart.width = chart_data.get('width', 15)
        chart.height = chart_data.get('height', 7.5)

        for i in range(no_of_lines):
            labels = Reference(ws, min_col=col_start, min_row=row_start+i+1, max_row=row_start+row_span-1)
            data = Reference(ws, min_col=col_start+i+1, min_row=row_start, max_row=row_start+row_span-1)
            series = Series(data, labels, title_from_data=True)
            series.marker.symbol = "circle"
            series.marker.graphicalProperties.solidFill = colors[i % 5]  # Marker filling
            series.marker.graphicalProperties.line.solidFill = colors[i % 5]  # Marker outline
            series.graphicalProperties.line.solidFill = colors[i % 5]
            series.graphicalProperties.line.width = 24050  # width in EMUs
            if x_axis_date_format:
                chart.x_axis.number_format = 'm/d'
                chart.x_axis.majorTimeUnit = "days"
            chart.series.append(series)

        ws.add_chart(chart, chart_data['chart-position'])

    def add_line_chart(self, ws, chart_data):
        """
        add line chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        x_axis_date_format = chart_data.get('x-axis-date-format')
        self._load_data(ws, chart_data['data'], row_start, col_start, False, x_axis_date_format)
        row_span = len(chart_data['data'])
        no_of_lines = chart_data.get('no-of-lines')
        if not no_of_lines:
            no_of_lines = 1
        colors = ["0077C8", "e35420", "32a3df", "673ab7", "9c27b0"]
        if 'colors' in chart_data and chart_data['colors']:
            colors = chart_data['colors']+colors

        chart = LineChart()
        chart.style = 5
        chart.legend.position = 'b'     # Location of legend
        if no_of_lines == 1:
            chart.legend = None
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')
        chart.width = chart_data.get('width', 20)
        chart.height = chart_data.get('height', 10)

        for i in range(no_of_lines):
            labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
            data = Reference(ws, min_col=col_start+i+1, min_row=row_start, max_row=row_start+row_span-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            series = chart.series[i]
            series.marker.graphicalProperties.solidFill = colors[i % no_of_lines]  # Marker filling
            series.marker.graphicalProperties.line.solidFill = colors[i % no_of_lines]  # Marker outline
            series.graphicalProperties.line.solidFill = colors[i % no_of_lines]
            series.graphicalProperties.line.width = 24050  # width in EMUs
            series.smooth=True        # Draw a smooth line
            if x_axis_date_format:
                chart.x_axis.number_format = 'm/d'
                chart.x_axis.majorTimeUnit = "days"

        ws.add_chart(chart, chart_data['chart-position'])


    def add_bar_line_trend_chart(self, ws, chart_data):
        """
        add bar line trend chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.style = 4
        chart.legend = None
        chart.width = chart_data.get('width', 17)
        chart.height = chart_data.get('height', 8.5)

        chart2 = LineChart()
        #chart2.type = "col"
        chart2.style = 10
        chart2.title = chart_data['chart-title']
        chart2.x_axis.title = chart_data.get('x-axis-title')
        chart2.y_axis.title = chart_data.get('y-axis-title')

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        chart2.style = 4
        chart2.legend = None
        chart2.width = chart_data.get('width', 17)
        chart2.height = chart_data.get('height', 8.5)

        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "0077C8"
        s1.graphicalProperties.solidFill = "0077C8"

        for idx, s in enumerate(chart.series):
            index = len(chart_colors)%(idx+1)
            if index >= len(chart_colors):
                index = 0
            s.graphicalProperties.line.solidFill = chart_colors[index]
            s.graphicalProperties.solidFill = chart_colors[index]

        chart += chart2
        ws.add_chart(chart, chart_data['chart-position'])


    def client_names_ids(self, customer_name):
        names = ''
        client_id = ''
        if customer_name and customer_name is not None:
            for i in customer_name:
                for j in i.items():
                    names+=j[1]['name']
                    client_id+=j[1]['uniqueId']
                    names+=', '
                    client_id+=', '
            names = names.rstrip(', ')
            client_id = client_id.rstrip(', ')
            return names, client_id
        else:
            return names, client_id


    def generate_glossary(self, ws, sheet_resp, title):
        """
        add glossary sheet
        """
        data_len = 35
        if sheet_resp is not None and 'data' in sheet_resp and sheet_resp['data'] is not None and len(sheet_resp['data']) >= data_len:
            data_len = len(sheet_resp['data'])+5
        for row_idx in range(1, data_len):
            row = ws.row_dimensions[row_idx]
            row.fill = PatternFill("solid", fgColor="eeeeee")

        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 24

        header_idx = 2
        value_idx = 2
        if sheet_resp is not None and 'data' in sheet_resp and sheet_resp['data'] is not None:
            data = sheet_resp['data']
            for key in data.keys():
                header_idx += 1
                value_idx += 1
                display_name = key
                display_value = data[key]

                cell = ws[f'B{header_idx}']
                cell.value = display_name
                cell.font = Font(color="00598B")
                cell.fill = PatternFill("solid", fgColor="eeeeee")

                cell = ws[f'C{value_idx}']
                cell.value = display_value
                cell.font = Font(color="000000")
                cell.fill = PatternFill("solid", fgColor="eeeeee")


    def generate_summary(self, res, tenantId, additional_props, report_gen_start_time, report_gen_completed_time):
        """
        add summary sheet
        """
        ws = self.wb.active
        ws.title = 'SUMMARY'

        for row_idx in range(1, 30):
            row = ws.row_dimensions[row_idx]
            row.fill = PatternFill("solid", fgColor="eeeeee")

        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24

        c_name = ''
        c_id = ''
        opsql_params = None
        if additional_props and additional_props is not None:
            if 'client_names' in additional_props:
                client_res = additional_props['client_names']
                result = self.client_names_ids(client_res)
                c_name = result[0]
                c_id = result[1]

            if 'opsql_query' in additional_props:
                opsql_params = additional_props['opsql_query']

        cell = ws['B4']
        cell.value = 'App'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C4']
        cell.value = APP_DISPLAY_NAME.replace('-', ' ').title()
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B5']
        cell.value = 'Tenant Name'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C5']
        # cell.value = res.json()['tenantInfo']['tenantName']
        cell.value = c_name
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        #cell = ws['B6']
        #cell.value = 'Tenant Id'
        #cell.font = Font(color="00598B")
        #cell.fill = PatternFill("solid", fgColor="eeeeee")

        #cell = ws['C6']
        ## cell.value = tenantId
        #cell.value = c_id
        #cell.font = Font(color="000000")
        #cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B6']
        cell.value = 'Run date'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C6']
        cell.value = report_gen_start_time
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B7']
        cell.value = 'Completion date'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C7']
        cell.value = report_gen_completed_time
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B8']
        cell.value = 'User'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C8']
        first_name = res.json()['analysisRun']['createdBy']['firstName']
        last_name = res.json()['analysisRun']['createdBy']['lastName']
        login_name = res.json()['analysisRun']['createdBy']['loginName']
        cell.value = f'{first_name} {last_name} ({login_name})'
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        header_idx = 8
        value_idx = 8
        if opsql_params is not None:

            for key in opsql_params.keys():
                header_idx += 1
                value_idx += 1
                display_name = self.get_display_name(key)
                display_value = self.get_display_value(opsql_params[key])

                cell = ws[f'B{header_idx}']
                cell.value = display_name
                cell.font = Font(color="00598B")
                cell.fill = PatternFill("solid", fgColor="eeeeee")

                cell = ws[f'C{value_idx}']
                cell.value = display_value
                cell.font = Font(color="000000")
                cell.fill = PatternFill("solid", fgColor="eeeeee")

        # cell = ws['B10']
        # cell.value = 'Analysis Parameters'
        # cell.font = Font(color="00598B")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")

        # cell = ws['B11']
        # cell.value = 'Analysis Period'
        # cell.alignment = Alignment(horizontal="right")
        # cell.font = Font(color="00598B")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")

        # cell = ws['C11']
        # start_date = res.json()['analysisPeriodInUserTZ']['displayStartTime']
        # end_date = res.json()['analysisPeriodInUserTZ']['displayEndTime']
        # cell.value = f"{start_date} - {end_date}"
        # cell.font = Font(color="000000")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")


        # cell = ws['B13']
        # cell.value = 'App version'
        # cell.alignment = Alignment(horizontal="right")
        # cell.font = Font(color="00598B")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")
        #
        # cell = ws['C13']
        # cell.value = res['analysisRun']['version']
        # cell.font = Font(color="000000")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")

    def get_display_name(self, key):
        if key == 'filterCriteria':
            return 'Query'
        if key == 'fields':
            return 'Attributes'
        if key == 'groupBy':
            return 'Group By'
        if key == 'soryBy':
            return 'Sort By'
        if key == 'sortByOrder':
            return 'Sort By Order'
        return key

    def get_display_value(self, value):
        if value is None and len(value) <= 0:
            return '-'
        else:
            if isinstance(value, list):
                all_values = ', '.join([val for val in value])
                return all_values
        return value


    def generate_pie_chart_colors(self, num_colors):
        colors=["0077C8", "00A3E0", "673AB7", "9C27B0", "E91E63", "F47925"]
        random.seed(123)  # 123 Fixed seed value
        #colors = []
        for _ in range(num_colors):
            hex_color = '{:06x}'.format(random.randint(0, 0xFFFFFF))
            colors.append(hex_color)
        return colors


    def add_header(self, ws, sheet_data):
        """
        add header to the sheet
        """
        for row_idx in range(1, ExcelRenderer.ROW_START):
            row = ws.row_dimensions[row_idx]
            row.fill = PatternFill("solid", fgColor="eeeeee")

        ws.merge_cells('B1:Z2')
        cell = ws['B1']
        cell.value = APP_DISPLAY_NAME.replace('-', ' ').title()
        cell.font = Font(color="00598B", bold=True)
        cell.fill = PatternFill("solid", fgColor="eeeeee")
        cell.alignment = Alignment(vertical="center")

    def add_component_title(self, ws, component_data):
        """
        add component title
        """
        cell = ws.cell(row=ExcelRenderer.ROW_START+component_data['start-row'], column=component_data['start-col'])
        cell.value = component_data['title']
        color = component_data.get('color', '00598B')
        cell.font = Font(color=color, bold=True)

    def add_component(self, ws, component_data):

        if component_data and 'data' in component_data:
            data = component_data['data']
            if isinstance(data, list):
                if len(data) >= 1000005:
                    logger.error('Data exceeding 1Million records.. Excluding excessive records..')
                    data = data[0:1000005]
                    component_data['data'] = data

        if component_data['type'] == 'table':
            if component_data.get('title'):
                self.add_component_title(ws, component_data)
            if 'metric_sheet' in component_data:
                self.add_metric_table(ws, component_data)
            else:
                self.add_table(ws, component_data)
        elif component_data['type'] == 'name':
            if component_data.get('title'):
                self.add_component_title(ws, component_data)
        elif component_data['type'] == 'doughnut-chart':
            self.add_doughnut_chart(ws, component_data)
        elif component_data['type'] == 'pie-chart':
            self.add_pie_chart(ws, component_data)
        elif component_data['type'] == 'bar-chart':
            self.add_bar_chart(ws, component_data)
        elif component_data['type'] == 'scatter-chart':
            self.add_scatter_chart(ws, component_data)
        elif component_data['type'] == 'line-chart':
            self.add_line_chart(ws, component_data)
        elif component_data['type'] == 'bar-line-trend-chart':
            self.add_bar_line_trend_chart(ws, component_data)
        elif component_data['type'] == 'stack-bar-chart':
            self.add_stack_bar_chart(ws, component_data)

    def render(self):
        """
        :return: Workbook
        """
        res = self.res
        tenantId = self.tenantId
        report_gen_start_time = self.report_gen_start_time
        report_gen_completed_time = self.report_gen_completed_time
        additional_props = self.additional_props
        self.generate_summary(res, tenantId, additional_props, report_gen_start_time, report_gen_completed_time)
        # run_data = self.analysis_run
        run_data = self.excel_data
        for sheet in run_data['sheets']:
            if 'documentation' in sheet and sheet['documentation'] == 'true':
                ws = self.wb.create_sheet(title=sheet['title'])
                self.generate_glossary(ws, sheet, title=sheet['title'])
            else:
                ws = self.wb.create_sheet(title=sheet['title'])
                self.add_header(ws, sheet)

                for component_data in sheet['components']:
                    self.add_component(ws, component_data)

        return self.wb
