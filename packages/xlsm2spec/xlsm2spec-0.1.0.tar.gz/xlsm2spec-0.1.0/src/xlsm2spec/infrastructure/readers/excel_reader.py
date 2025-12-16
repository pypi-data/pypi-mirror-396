"""Excel読み込みアダプター - xlrd, openpyxlを使用"""

import os
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from xlsm2spec.domain.exceptions import FileNotFoundError, UnsupportedFormatError
from xlsm2spec.domain.models import Workbook


class ExcelReader:
    """Excelファイルを読み込み、Workbookオブジェクトを生成する"""

    SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}

    def read(self, filepath: str) -> Workbook:
        """
        Excelファイルを読み込み、Workbookオブジェクトを返す

        Args:
            filepath: 読み込むExcelファイルのパス

        Returns:
            Workbook: 読み込んだワークブック情報

        Raises:
            FileNotFoundError: ファイルが存在しない場合
            UnsupportedFormatError: サポートされていない形式の場合
        """
        path = Path(filepath)

        # ファイル存在チェック
        if not path.exists():
            raise FileNotFoundError(filepath)

        # 拡張子チェック
        ext = path.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise UnsupportedFormatError(filepath, ext)

        # 形式に応じて読み込み
        if ext == ".xls":
            return self._read_xls(filepath)
        else:
            return self._read_xlsx(filepath)

    def _read_xls(self, filepath: str) -> Workbook:
        """xls形式（BIFF8）を読み込む"""
        try:
            wb = xlrd.open_workbook(filepath)
            sheets = [sheet.name for sheet in wb.sheets()]
            return Workbook(
                filename=os.path.basename(filepath),
                sheets=sheets,
                vba_project=None,
            )
        except Exception as e:
            raise UnsupportedFormatError(filepath, ".xls") from e

    def _read_xlsx(self, filepath: str) -> Workbook:
        """xlsx/xlsm形式（OOXML）を読み込む"""
        try:
            wb = load_workbook(filepath, read_only=True, keep_vba=True)
            sheets = wb.sheetnames
            wb.close()
            return Workbook(
                filename=os.path.basename(filepath),
                sheets=sheets,
                vba_project=None,
            )
        except Exception as e:
            ext = Path(filepath).suffix.lower()
            raise UnsupportedFormatError(filepath, ext) from e
