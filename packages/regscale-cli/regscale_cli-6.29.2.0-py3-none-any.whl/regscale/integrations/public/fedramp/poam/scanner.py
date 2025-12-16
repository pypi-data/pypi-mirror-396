#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""FedRAMP Scanner Integration"""

import logging
import re
from typing import Iterator, List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    import numpy

from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import column_index_from_string  # type: ignore
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
from openpyxl.workbook import Workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet

from regscale.core.app.utils.app_utils import error_and_exit, get_current_datetime
from regscale.core.utils.date import date_str
from regscale.integrations.scanner_integration import (
    IntegrationAsset,
    IntegrationFinding,
    ScannerIntegration,
    issue_due_date,
)
from regscale.models import ImportValidater, IssueSeverity, Mapping, regscale_models
from regscale.validation.address import validate_ip_address, validate_mac_address

logger = logging.getLogger("regscale")

WEAKNESS_DETECTOR_SOURCE = "Weakness Detector Source"
ASSET_IDENTIFIER = "Asset Identifier"
SCHEDULED_COMPLETION_DATE = "Scheduled Completion Date"
MILESTONE_CHANGES = "Milestone Changes"
ORIGINAL_RISK_RATING = "Original Risk Rating"
ADJUSTED_RISK_RATING = "Adjusted Risk Rating"
FILE_PATH_ERROR = "File path is required."


class FedrampPoamIntegration(ScannerIntegration):
    """Integration class for FedRAMP POAM scanning."""

    # Keys set in the `set_keys` method of `ScannerIntegration`
    title = "FedRAMP"
    file_path: str = ""
    poam_sheets: List[str] = []
    validators: dict = {}
    workbook: Optional[Workbook] = None

    asset_identifier_field = "otherTrackingNumber"
    finding_severity_map = {
        "Low": regscale_models.IssueSeverity.Low,
        "Moderate": regscale_models.IssueSeverity.Moderate,
        "High": regscale_models.IssueSeverity.High,
        "Critical": regscale_models.IssueSeverity.Critical,
    }
    poam_id_header = "POAM ID"
    blank_records: int = 0
    blank_threshold: int = 3
    error_records: int = 0
    skipped_records: int = 0
    processed_assets: set[str] = set()  # Track processed assets across all methods

    fedramp_poam_columns = [
        "POAM ID",
        "Weakness Name",
        "Weakness Description",
        WEAKNESS_DETECTOR_SOURCE,
        "Weakness Source Identifier",
        ASSET_IDENTIFIER,
        "Point of Contact",
        "Resources Required",
        "Overall Remediation Plan",
        "Original Detection Date",
        SCHEDULED_COMPLETION_DATE,
        "Planned Milestones",
        MILESTONE_CHANGES,
        "Status Date",
        ORIGINAL_RISK_RATING,
        ADJUSTED_RISK_RATING,
        "Risk Adjustment",
        "False Positive",
        "Operational Requirement",
        "Deviation Rationale",
        "Comments",
    ]

    """
    Unused columns:
    # "Vendor Dependency",
    # "Last Vendor Check-in Date",
    # "Vendor Dependent Product Name",
    # "Supporting Documents",
    # "Auto-Approve",
    # "Binding Operational Directive 22-01 tracking",
    # "Binding Operational Directive 22-01 Due Date",
    # "CVE",
    # "Service Name",
    """

    def __init__(self, plan_id: int, **kwargs: dict):
        super().__init__(plan_id=plan_id)
        try:
            # Use read_only mode for memory efficiency, purposefully use kwarg index to force KeyError
            if "file_path" in kwargs:
                self.file_path = kwargs["file_path"]
            if not self.file_path:
                error_and_exit(FILE_PATH_ERROR)
            self.workbook = self.workbook or load_workbook(filename=self.file_path, data_only=True, read_only=True)
            self.poam_sheets = kwargs.get("poam_sheets") or [
                sheet for sheet in self.workbook.sheetnames if re.search("POA&M Items|Configuration Findings", sheet)
            ]
        except (FileNotFoundError, InvalidFileException, KeyError) as e:
            logger.error(f"Failed to load workbook: {e}")
            return
        # Validate Here
        if not self.validators and isinstance(self.poam_sheets, list):
            for sheet in self.poam_sheets:
                ws = self.workbook[sheet]
                mapping_path = "./mappings/fedramp_poam/" + sheet
                validator = ImportValidater(
                    file_path=self.file_path,
                    disable_mapping=True,
                    required_headers=self.fedramp_poam_columns,
                    worksheet_name=sheet,
                    mapping_file_path=mapping_path,
                    prompt=True,
                    skip_rows=self.find_header_row(ws),
                    ignore_unnamed=True,
                )
                self.validators[sheet] = validator
        self.processed_assets = set()  # Reset processed assets on init

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup resources."""
        if self.workbook:
            self.workbook.close()

    def fetch_findings(self, *args, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Fetches findings from FedRAMP POAM files.

        :yield: Iterator of validated integration findings
        """
        if not self.file_path:
            error_and_exit(FILE_PATH_ERROR)

        findings = []
        try:
            for sheet in self.poam_sheets:
                validator = self.validators.get(sheet)
                if not validator:
                    logger.warning(f"No validator found for sheet: {sheet}")
                    continue

                sheet_kwargs = {**kwargs, "sheet": sheet}
                sheet_findings = self._process_sheet(**sheet_kwargs)
                findings.extend(sheet_findings)

            self.num_findings_to_process = len(findings)
            return iter(findings)

        except Exception as e:
            logger.error(f"Error fetching findings from POAM file: {str(e)}")
            return iter(findings)

    def _process_sheet(self, **kwargs: dict) -> List[IntegrationFinding]:
        """
        Process a single sheet from the POAM workbook.

        :param str sheet: The sheet name
        :param **kwargs: Arbitrary keyword arguments
        :return: List of IntegrationFinding objects
        :rtype: List[IntegrationFinding]
        """

        finding_lst = []
        if not self.workbook:
            return finding_lst
        sheet = kwargs.get("sheet")
        previous_status_date: str = None
        resolve_status = kwargs.get("resolve_empty_status_date", "CURRENT_DATE")
        ws = self.workbook[sheet]
        validator = self.validators.get(sheet)
        category = ws["C3"].value or "Low"
        if not ws["C3"].value:
            logger.warning(f"Category is required in cell C3. Defaulting to Low for sheet {sheet}.")

        status = self.determine_status(sheet)
        if status is None:
            logger.warning(f"Unable to determine POA&M status for sheet {sheet}. Skipping import.")
            return finding_lst

        try:
            start_row = self.find_start_row(validator.data.values)
        except IndexError:
            return finding_lst

        if start_row is None:
            logger.warning(f"No POAM entries found in sheet {sheet}. Skipping.")
            return finding_lst

        logger.info("Processing sheet: %s for findings, rows: %i", sheet, len(validator.data))
        for index, row in enumerate(validator.data.values):
            try:
                if index < start_row:
                    continue
                if not validator and validator.mapping:
                    logger.error("Validator mapping or validator mapping is None")
                    break
                val_mapping = validator.mapping  # convert tuple to dict
                data = dict(zip(val_mapping.mapping, row))

                if not isinstance(data, dict):
                    logger.error("data must be a dictionary")
                    continue

                if parsed_category := self.determine_category(data, validator):
                    category = parsed_category
                    # Category must be in IssueSeverity
                    if category not in [IssueSeverity.Low.name, IssueSeverity.Moderate.name, IssueSeverity.High.name]:
                        logger.warning(f"Invalid Original Risk Rating: {category} in sheet {sheet}. Skipping.")
                        continue
                logger.debug(f"Processing row {index} in sheet {sheet.strip()} for findings")
                logger.debug(f"Status: {status}, Category: {category}")
                if not status:
                    logger.warning(f"Status is required in sheet {sheet}. Skipping.")
                    continue
                if not category:
                    logger.warning(f"Category is required in sheet {sheet}. Skipping.")
                    continue
                findings = self.parse_finding(
                    data=data,
                    previous_status_date=previous_status_date,
                    status=status,
                    category=category,
                    index=index,
                    sheet=sheet,
                    validator=validator,
                    resolve_status=resolve_status,
                )
                for finding in findings:
                    previous_status_date = finding.date_last_updated
                    if isinstance(finding, IntegrationFinding):
                        finding_lst.append(finding)

            except Exception as e:
                logger.error(f"Error processing row {index} in sheet {sheet}: {str(e)}")
                self.error_records += 1
                continue

        return finding_lst

    def determine_category(self, data: dict, validator: ImportValidater) -> str:
        """
        Determine the category of the finding by direct string or from a mapping.

        :param dict data: The row data
        :param ImportValidater validator: The ImportValidater object
        :return: The category of the finding
        """
        dat_map = {
            "medium": IssueSeverity.Moderate.name,
            "high": IssueSeverity.High.name,
            "critical": IssueSeverity.High.name,
            "low": IssueSeverity.Low.name,
        }
        res = validator.mapping.get_value(data, ORIGINAL_RISK_RATING)
        if res.lower() not in [mem.lower() for mem in IssueSeverity.__members__]:
            res = dat_map.get(res.lower(), IssueSeverity.Low.name)
        return res

    @staticmethod
    def is_poam(finding: IntegrationFinding) -> bool:
        """
        Determine if this finding is a POAM.

        :param IntegrationFinding finding: The finding to check
        :return: True if this is a POAM finding
        :rtype: bool
        """
        return True  # All FedRAMP findings are POAMs

    @staticmethod
    def get_issue_title(finding: IntegrationFinding) -> str:
        """
        Get the title for an issue.

        :param IntegrationFinding finding: The finding
        :return: The issue title
        :rtype: str
        """
        return finding.title[:255]  # Enforce title length limit

    def parse_finding(self, data: dict, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationFinding objects.
        Creates a separate finding for each asset and CVE combination.

        :param dict data: The row data
        :param kwargs: Arbitrary keyword arguments
        :rtype: Iterator[IntegrationFinding]
        :yields: IntegrationFinding
        """
        findings = []
        status = kwargs.get("status")
        if not isinstance(status, str):
            raise TypeError("status must be a string")

        category = kwargs.get("category")
        if not isinstance(category, str):
            raise TypeError("category must be a string")

        index = kwargs.get("index")
        if not isinstance(index, int):
            raise TypeError("index must be an integer")

        sheet = kwargs.get("sheet")
        if not isinstance(sheet, str):
            raise TypeError("sheet must be a string")

        resolve_status = kwargs.get("resolve_status")
        if not isinstance(resolve_status, str):
            raise TypeError("resolve_status must be a string")

        val_mapping = kwargs.get("validator").mapping

        try:
            poam_id = val_mapping.get_value(data, self.poam_id_header)
            weakness_name = str(val_mapping.get_value(data, "Weakness Name"))

            if not poam_id and weakness_name in [None, "None", ""]:
                self.blank_records += 1
                yield from findings

            if not poam_id or not poam_id.upper():
                logger.debug(
                    f"Invalid POAM ID on row {index}, sheet {sheet}: weakness_name={weakness_name}, poam_id={poam_id}"
                )
                logger.warning(f"Invalid POAM ID on row {index}, sheet {sheet}. Skipping.")
                yield from findings

            if not weakness_name:
                logger.warning(f"Title is required on row {index}, sheet {sheet}. Unable to import")
                yield from findings

            # Get and validate plugin ID
            raw_plugin_id = val_mapping.get_value(data, "Weakness Source Identifier")
            try:
                plugin_id_int = (
                    int(raw_plugin_id)
                    if raw_plugin_id and str(raw_plugin_id).isdigit()
                    else abs(hash(str(raw_plugin_id or ""))) % (10**9)
                )
            except (ValueError, TypeError):
                plugin_id_int = abs(hash(poam_id)) % (10**9)

            # Get asset identifiers
            asset_ids = val_mapping.get_value(data, ASSET_IDENTIFIER)
            if not asset_ids:
                logger.warning(f"No asset identifier found on row {index}, sheet {sheet}. Skipping.")
                yield from findings

            # Clean asset identifiers
            asset_id_list = self.gen_asset_list(asset_ids)

            if not asset_id_list:
                logger.warning(f"No valid asset identifiers found on row {index}, sheet {sheet}. Skipping.")
                yield from findings

            # Get and validate CVEs
            cves = self.process_cve(val_mapping.get_value(data, "CVE"), index, sheet)
            cve_list = cves.split("\n") if cves else [""]  # Use empty string if no CVEs

            # Create a finding for each asset and CVE combination
            for asset_id in asset_id_list:
                for cve in cve_list:
                    # Create unique plugin ID for each CVE
                    if cve:
                        unique_plugin_id = abs(hash(f"{plugin_id_int}:{cve}")) % (10**9)
                    else:
                        unique_plugin_id = plugin_id_int

                    date_created = (
                        date_str(val_mapping.get_value(data, "Original Detection Date")) or get_current_datetime()
                    )
                    due_date = date_str(
                        val_mapping.get_value(data, SCHEDULED_COMPLETION_DATE)
                        if val_mapping.get_value(data, SCHEDULED_COMPLETION_DATE) != "#REF!"
                        else ""
                    )
                    severity: IssueSeverity = getattr(IssueSeverity, category.title(), IssueSeverity.NotAssigned)
                    if date_created and not due_date:
                        due_date = issue_due_date(severity, date_created)

                    # Status Date
                    status_date = date_str(val_mapping.get_value(data, "Status Date"))
                    if not status_date or status_date == "NaT":
                        status_date = self.determine_status_date(**kwargs)
                        # if status date is still None, skip this finding
                        if not status_date:
                            continue

                    # Extract Controls field (Column B) for Configuration Findings
                    controls = val_mapping.get_value(data, "Controls")
                    affected_controls = str(controls) if controls else None

                    # Validate pluginText
                    finding = IntegrationFinding(
                        control_labels=[],
                        title=f"{weakness_name[:240]} - {cve}" if cve else weakness_name[:255],
                        category=f"FedRAMP POAM: {category}",
                        description=val_mapping.get_value(data, "Weakness Description") or "",
                        severity=severity,
                        status=(
                            regscale_models.IssueStatus.Closed
                            if status.lower() == "closed"
                            else regscale_models.IssueStatus.Open
                        ),
                        asset_identifier=asset_id,
                        external_id=f"{poam_id}:{cve}" if cve else poam_id,
                        date_created=date_created,
                        date_last_updated=status_date,
                        due_date=due_date,
                        cve=cve,  # Single CVE per finding
                        plugin_name=val_mapping.get_value(data, WEAKNESS_DETECTOR_SOURCE) or "",
                        plugin_id=str(unique_plugin_id),
                        observations=str(val_mapping.get_value(data, MILESTONE_CHANGES)) or "",
                        poam_comments=self.empty(val_mapping.get_value(data, "Comments")),
                        remediation=self.empty(val_mapping.get_value(data, "Overall Remediation Plan")),
                        basis_for_adjustment=str(self.get_basis_for_adjustment(val_mapping=val_mapping, data=data)),
                        vulnerability_type="FedRAMP",
                        source_report=str(val_mapping.get_value(data, WEAKNESS_DETECTOR_SOURCE)),
                        point_of_contact=str(val_mapping.get_value(data, "Point of Contact")),
                        milestone_changes=str(val_mapping.get_value(data, MILESTONE_CHANGES)),
                        planned_milestone_changes=str(val_mapping.get_value(data, "Planned Milestones")),
                        adjusted_risk_rating=val_mapping.get_value(data, ADJUSTED_RISK_RATING),
                        risk_adjustment=self.determine_risk_adjustment(val_mapping.get_value(data, "Risk Adjustment")),
                        operational_requirements=str(val_mapping.get_value(data, "Operational Requirement")),
                        deviation_rationale=str(val_mapping.get_value(data, "Deviation Rationale")),
                        affected_controls=affected_controls,
                        poam_id=poam_id,
                    )
                    if finding.is_valid():
                        findings.append(finding)

        except Exception as e:
            logger.error(f"Error processing row {index} in sheet {sheet}: {str(e)}")
            self.error_records += 1

        yield from findings

    def determine_status_date(self, **kwargs):
        """
        Determine the status date.

        :param kwargs: Arbitrary keyword arguments
        :return: The status date
        :rtype: str
        """
        index = kwargs.get("index")
        sheet = kwargs.get("sheet")
        resolve_status = kwargs.get("resolve_status")
        status_map = {
            "CURRENT_DATE": date_str(get_current_datetime()),
            "USE_NEIGHBOR": date_str(kwargs.get("previous_status_date")),
        }
        res = date_str(status_map.get(resolve_status), "%m-%d-%Y")
        if res:
            logger.warning(
                "Status Date missing on row %i, sheet %s, defaulting to %s: %s",
                index,
                sheet,
                resolve_status.lower().replace("_", " "),
                res,
            )
            return res
        logger.warning(
            f"Status Date missing on row {index}, sheet {sheet}. Unable to find valid neighbor, falling back to current date."
        )
        return date_str(status_map.get("CURRENT_DATE"), "%Y-%m-%d")

    # flake8: noqa: C901
    def parse_asset(self, row: List, validator: ImportValidater) -> List[IntegrationAsset]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationAsset objects.
        Handles multiple comma-separated asset identifiers.

        :param List row: The row data from the spreadsheet
        :param ImportValidater validator: The ImportValidater object
        :rtype: List[IntegrationAsset]
        """
        row_assets = []
        try:
            if validator and validator.mapping:
                val_mapping = validator.mapping  # convert tuple to dict
                data = dict(zip(val_mapping.mapping, row))
            else:
                logger.error("Validator mapping is None")
                return row_assets
            asset_ids = val_mapping.get_value(data, ASSET_IDENTIFIER)
            if not asset_ids:
                return row_assets

            # Skip rows where asset identifier contains date/description text (header rows)
            asset_ids_str = str(asset_ids).lower()
            if any(
                keyword in asset_ids_str
                for keyword in [
                    "date the weakness",
                    "aka discovery",
                    "permanent column",
                    "date of intended",
                    "last changed or closed",
                    "port/protocol",
                    "specified in the inventory",
                ]
            ):
                logger.debug(f"Skipping row with header/description text in asset identifier: {str(asset_ids)[:100]}")
                return row_assets

            asset_id_list = self.gen_asset_list(asset_ids)

            if not asset_id_list:
                return row_assets

            def clean_str(val: Optional[str], default: str = "") -> str:
                """Clean and validate string values."""
                if not val:
                    return default
                if not isinstance(val, str):
                    return default

                # Remove problematic patterns
                val = str(val).strip()
                if any(
                    pattern in val.lower()
                    for pattern in [
                        "n/a",
                        "none",
                        "null",
                        "undefined",
                        "planned",
                        "pending",
                        "tbd",
                        "remediation",
                        "deviation",
                        "request",
                        "vulnerability",
                    ]
                ):
                    return default

                # Remove date-like strings
                if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", val):
                    return default

                # Remove long descriptions
                if len(val) > 100 or "\n" in val:
                    return default

                return val

            def determine_asset_type(asset_id: str, raw_type: str) -> str:
                """Determine asset type based on asset ID and raw type."""
                if not raw_type or raw_type == "Other":
                    # Check for common patterns in asset ID
                    if any(pattern in asset_id.lower() for pattern in ["docker", "container", "image", "registry"]):
                        return "Container"
                    elif any(pattern in asset_id.lower() for pattern in ["lambda", "function", "azure-function"]):
                        return "Function"
                    elif any(pattern in asset_id.lower() for pattern in ["s3", "bucket", "blob", "storage"]):
                        return "Storage"
                    elif any(pattern in asset_id.lower() for pattern in ["db", "database", "rds", "sql"]):
                        return "Database"
                    elif any(pattern in asset_id.lower() for pattern in ["ec2", "vm", "instance"]):
                        return "Virtual Machine"
                    else:
                        return "Other"
                return raw_type

            for asset_id in asset_id_list:
                # Handle long asset names
                asset_name, asset_notes = self._handle_long_asset_name(asset_id)

                # Get raw values and clean them
                raw_values = {
                    "ip": asset_id if validate_ip_address(asset_id) else "",
                    "type": clean_str(val_mapping.get_value(data, "Resources Required")),
                    "fqdn": asset_id if self.is_valid_fqdn(asset_id) else "",
                    "mac": asset_id if validate_mac_address(asset_id) else "",
                }

                # Determine proper asset type
                asset_type = determine_asset_type(asset_id, raw_values["type"])

                res = IntegrationAsset(
                    name=asset_name,  # Use shortened name if needed
                    identifier=asset_name,  # Use shortened name as identifier
                    asset_type=asset_type,  # Use determined asset type
                    asset_category=regscale_models.AssetCategory.Hardware,
                    parent_id=self.plan_id,
                    parent_module=regscale_models.SecurityPlan.get_module_string(),
                    status="Active (On Network)",
                    ip_address=raw_values["ip"],
                    fqdn=raw_values["fqdn"],
                    mac_address=raw_values["mac"],
                    notes=asset_notes,  # Store full name if truncated
                    date_last_updated=get_current_datetime(),
                )
                row_assets.append(res)
        except (KeyError, ValueError, TypeError) as kex:
            logger.error(f"Error parsing asset from row: {str(kex)} (Exception type: {type(kex).__name__})")
        except Exception as ex:
            logger.error(f"Unknown Error parsing asset from row: {str(ex)}")

        return row_assets

    def _handle_long_asset_name(self, asset_id: str, max_length: int = 450) -> tuple[str, str]:
        """
        Handle asset names that exceed database field limits.
        Generates a hash-based identifier for long names and preserves full name in notes.

        :param str asset_id: The asset identifier
        :param int max_length: Maximum allowed length (default: 450)
        :return: Tuple of (shortened_name, notes)
        :rtype: tuple[str, str]
        """
        if len(asset_id) <= max_length:
            return asset_id, ""

        # Generate hash-based identifier
        import hashlib

        hash_suffix = hashlib.sha256(asset_id.encode()).hexdigest()[:8]
        truncated = asset_id[: max_length - 9]  # Leave room for underscore and hash
        short_name = f"{truncated}_{hash_suffix}"
        notes = f"Full identifier: {asset_id}"

        logger.warning(f"Asset identifier exceeds {max_length} chars, truncated to: {short_name[:100]}...")
        return short_name, notes

    def gen_asset_list(self, asset_ids: str):
        """
        Generate a list of asset identifiers from a string.
        Handles multiple separator types: commas, semicolons, pipes, tabs, newlines.
        Preserves spaces within asset identifiers (e.g., "10.10.160.200 ( 2049 / TCP )").
        Also removes surrounding brackets that might wrap the list.

        :param str asset_ids: The asset identifier string
        :return: The list of asset identifiers
        :rtype: List[str]
        """
        # Remove surrounding brackets if present (handles cases like "[10.10.1.1, 10.10.1.2]")
        asset_ids = asset_ids.strip()
        if asset_ids.startswith("[") and asset_ids.endswith("]"):
            asset_ids = asset_ids[1:-1].strip()

        # Split only on actual delimiters: commas, semicolons, pipes, tabs, newlines, carriage returns
        # Do NOT split on spaces to preserve port/protocol info like "10.10.160.200 ( 2049 / TCP )"
        return [
            aid.strip()
            for aid in re.split(r"[,;\|\t\n\r]+", asset_ids)
            if isinstance(aid, str) and aid.strip() and len(aid.strip()) > 0
        ]

    @staticmethod
    def empty(string: Optional[str]) -> Optional[str]:
        """
        Convert empty strings and "None" to None.

        :param Optional[str] string: The input string
        :return: The processed string or None
        :rtype: Optional[str]
        """
        if not isinstance(string, str):
            return None
        if string.lower() in ["none", "n/a"]:
            return None
        return string

    @staticmethod
    def determine_status(sheet: str) -> Optional[str]:
        """
        Determine the status based on sheet name.

        :param str sheet: The sheet name
        :return: The status (Open/Closed) or None
        :rtype: Optional[str]
        """
        sheet_lower = sheet.lower()
        if "closed" in sheet_lower:
            return "Closed"
        elif "open" in sheet_lower or "configuration findings" in sheet_lower:
            return "Open"
        return None

    def find_start_row(self, array: "numpy.ndarray") -> Optional[int]:
        """
        Find the first row containing POAM data.

        :param array: NumPy array containing the data
        :return: The row number where POAM entries start
        :rtype: Optional[int]
        """
        if array[0][0] == "Unique identifier for each POAM Item" and array[1][0] == "Unique Identifier":
            if array[2][0] == "V-1Example":
                return 3
            return 2

        return 0

    def get_basis_for_adjustment(self, val_mapping: Mapping, data: dict) -> Optional[str]:
        """
        Get the basis for risk adjustment.

        :param Mapping val_mapping: The mapping object
        :param dict data: The row data
        :return: The basis for adjustment
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(val_mapping.get_value(data, "Comments"))  # e.g. row 23
        risk_rating = val_mapping.get_value(data, ORIGINAL_RISK_RATING)
        adjusted_risk_rating = val_mapping.get_value(data, ADJUSTED_RISK_RATING)

        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    def process_cve(self, cve: Optional[str], index: int, sheet: str) -> Optional[str]:
        """
        Process and validate CVE string. Handles multiple comma-separated CVEs.

        :param Optional[str] cve: The CVE string
        :param int index: The row index
        :param str sheet: The sheet name
        :return: The processed CVE string, multiple CVEs joined by newlines
        :rtype: Optional[str]
        """
        cve = self.empty(cve)
        if not cve:
            return None

        # Split by comma and clean
        cve_list = [c.strip() for c in cve.split(",") if c.strip()]
        if not cve_list:
            return None

        valid_cves = []
        cve_pattern = r"(?:CVE-\d{4}-\d{4,7}|RHSA-\d{4}:\d+|GHSA-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4})"

        for single_cve in cve_list:
            # Search for CVE pattern in the string
            cve_match = re.search(cve_pattern, single_cve, re.IGNORECASE)
            if cve_match:
                valid_cves.append(cve_match.group(0).upper())
            else:
                logger.warning(f"Invalid CVE format: {single_cve} on row {index}, sheet {sheet}. Skipping this CVE.")

        # Return newline-separated CVEs or None if no valid CVEs found
        return "\n".join(valid_cves) if valid_cves else None

    def is_valid_fqdn(self, hostname: str) -> bool:
        """
        Check if the hostname is valid.

        :param str hostname: The hostname string
        :return: True if the hostname is valid
        :rtype: bool
        """
        if validate_ip_address(hostname):
            return False

        if not hostname or len(hostname) > 255:
            return False

        allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-.")
        if not all(char in allowed for char in hostname):
            return False

        parts = hostname.split(".")
        if len(parts) < 2:
            return False

        if hostname[-1] == ".":
            hostname = hostname[:-1]

        return all(
            1 <= len(part) <= 63 and not part.startswith("-") and not part.endswith("-") for part in hostname.split(".")
        )

    def find_header_row(self, ws: Worksheet) -> int:
        """
        Find the header row in the POAM sheet.

        :param ws: Worksheet
        :return: The header row number
        :rtype: int
        """
        # Loop every row
        header_row = None
        for ix, row in enumerate(ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, values_only=True)):
            for cell in row:
                if cell and self.poam_id_header in str(cell):
                    header_row = ix + 1
                    break
            if header_row:
                break
        if not header_row:
            error_and_exit("Unable to find the header row in the POAM sheet.")
        return header_row

    def progress_bar(self, progress, total, width=50):
        filled = int(width * progress // total)
        bar = "=" * filled + "-" * (width - filled)
        percent = progress / total * 100
        return f"[{bar}] {percent:.1f}%"

    def fetch_assets(self, *args, **kwargs) -> Iterator[IntegrationAsset]:
        """
        Fetch assets from FedRAMP POAM files.

        :raises POAMProcessingError: If there's an error processing the POAM file
        :return: Iterator of parsed integration assets
        :rtype: Iterator[IntegrationAsset]
        """
        if not self.file_path:
            error_and_exit(FILE_PATH_ERROR)

        assets = []
        total_processed = 0

        try:
            logger.info(f"Starting POAM sheets processing from {self.file_path}")

            with self._get_lock("processed_assets"):
                for sheet_name in self.poam_sheets:
                    try:
                        validator = self.validators.get(sheet_name)
                        if not validator:
                            logger.warning(f"No validator found for sheet: {sheet_name}")
                            continue

                        data = validator.data
                        if data.empty:
                            logger.warning(f"Empty sheet found: {sheet_name}")
                            continue

                        start_row = self.find_start_row(data.values)
                        rows_count = len(data.values)

                        logger.info(
                            f"Processing sheet '{sheet_name}' with {rows_count} rows starting from row {start_row}"
                        )

                        for ix, row in enumerate(data.values[start_row:], start=start_row):
                            try:
                                new_assets = self.parse_asset(row, validator)
                                assets.extend(new_assets)
                                total_processed += len(new_assets)
                            except Exception as row_error:
                                logger.error(
                                    f"Failed to process row {ix} in sheet '{sheet_name}': {str(row_error)}",
                                    exc_info=True,
                                )
                                self.error_records += 1

                    except Exception as sheet_error:
                        logger.error(f"Failed to process sheet '{sheet_name}': {str(sheet_error)}", exc_info=True)
                        continue

        except Exception as e:
            error_msg = f"Critical error while processing POAM file: {str(e)}"
            logger.error(error_msg, exc_info=True)

        finally:
            logger.info(f"Completed processing with {total_processed} assets and {self.error_records} errors")

        self.num_assets_to_process = len(assets)
        return iter(assets)

    def find_max_row(self, start_row: int, ws: Worksheet) -> int:
        """
        A Method to find the max row in the worksheet.

        :param start_row: int
        :param ws: Worksheet
        :return: The max row number
        :rtype: int
        """
        last_row = ws.max_row
        for row in range(start_row, last_row):
            if ws.cell(row=row, column=1).value:
                continue
            else:
                return row
        return last_row

    def determine_risk_adjustment(self, param):
        """
        Determine the risk adjustment.

        Yes, No or Pending

        :param param: The parameter to check
        :return: The risk adjustment
        """
        adjustment_map = {
            "false": "No",
            "no": "No",
            "": "No",
            None: "No",
            "true": "Yes",
            "yes": "Yes",
            "pending": "Pending",
            "closed": "No",
            "n/a": "No",
        }
        # BMC Prefers this
        return adjustment_map.get(param.lower(), "No")
