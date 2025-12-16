#  Copyright (c) 2024 Higher Bar AI, PBC
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

"""Utilities for reading and processing documents for AI workflows."""

from ai_workflows.llm_utilities import LLMInterface
import fitz  # (PyMuPDF)
from fitz.utils import get_pixmap
import pymupdf4llm
from PIL import Image
import json
import subprocess
import os
from typing import Tuple, List, Union, Dict, Optional, Any, Callable
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.merge import MergedCellRange
from unstructured.partition.auto import partition
from unstructured.documents.elements import (
    Element, Text, Title, NarrativeText, ListItem, Table, Image as ImageElement, PageBreak,
    Header, Footer, Address
)
from dataclasses import dataclass
from pathlib import Path
import re
import tempfile
import logging
import markdown as mdpackage
from bs4 import BeautifulSoup
from docling.document_converter import DocumentConverter
from docling.datamodel.base_models import FormatToExtensions
from functools import reduce
import concurrent.futures


class DocumentInterface:
    """Utility class for reading and processing documents for AI workflows."""

    # member variables
    via_pdf_file_extensions = [".docx", ".doc", ".pptx"]
    # use Docling, except not for .html files since it doesn't preserve hyperlinks and not .txt because it's broken
    docling_file_extensions = [f".{ext}" for exts in FormatToExtensions.values() for ext in exts
                               if ext not in ["html", "txt"]]
    pymupdf_file_extensions = [".pdf", ".xps", ".epub", ".mobi", ".fb2", ".cbz", ".svg", ".txt"]
    libreoffice_file_extensions = [
        ".odt", ".csv", ".db", ".doc", ".docx", ".dotx", ".fodp", ".fods", ".fodt", ".mml", ".odb", ".odf", ".odg",
        ".odm", ".odp", ".ods", ".otg", ".otp", ".ots", ".ott", ".oxt", ".pdf", ".pptx", ".psw", ".sda", ".sdc", ".sdd",
        ".sdp", ".sdw", ".slk", ".smf", ".stc", ".std", ".sti", ".stw", ".sxc", ".sxg", ".sxi", ".sxm", ".sxw", ".uof",
        ".uop", ".uos", ".uot", ".vsd", ".vsdx", ".wdb", ".wps", ".wri", ".xls", ".xlsx"
    ]
    max_xlsx_via_pdf_pages: int = 10
    max_json_via_markdown_chunk_tokens = None
    max_parallel_requests = 5

    llm_interface: LLMInterface = None
    pdf_image_dpi: int = 150
    pdf_image_max_bytes = 1024 * 1024 * 4  # 4MB

    def __init__(self, llm_interface: LLMInterface = None, pdf_image_dpi: int = 150,
                 pdf_image_max_bytes: int = 1024 * 1024 * 4, max_parallel_requests: int = 5):
        """
        Initialize the document interface for reading and processing documents.

        :param llm_interface: LLM interface for interacting with LLMs in AI workflows (defaults to None, which won't
            use an LLM to convert supported document types to markdown).
        :type llm_interface: LLMInterface
        :param pdf_image_dpi: DPI to use for rendering PDF images. Default is 150, which is generally plenty for LLM
            applications.
        :type pdf_image_dpi: int
        :param pdf_image_max_bytes: Maximum size in bytes for an image to be processed. Default is 4MB.
        :type pdf_image_max_bytes: int
        :param max_parallel_requests: Maximum number of parallel requests to make when accessing LLM. Default is 5.
        :type max_parallel_requests: int
        """

        # if specified, remember our LLM interface
        if llm_interface:
            self.llm_interface = llm_interface

            # also set max chunk size for Markdown processing to 75% of the LLM's max output tokens (leaving overhead
            # for JSON and other output formatting)
            self.max_json_via_markdown_chunk_tokens = int(self.llm_interface.max_tokens * 0.75)

        # remember other settings
        self.pdf_image_dpi = pdf_image_dpi
        self.pdf_image_max_bytes = pdf_image_max_bytes
        self.max_parallel_requests = max_parallel_requests

    def convert_to_markdown(self, filepath: str, use_text: bool = False) -> str:
        """
        Convert a document to markdown.

        :param filepath: Path to the file.
        :type filepath: str
        :param use_text: Whether to use extracted text to help the LLM with extracting text from the images. Default
            is False.
        :type use_text: bool
        :return: Markdown output.
        :rtype: str
        """

        # use internal conversion function
        return self._convert(filepath, to_format="md", use_text=use_text)

    def convert_to_json(self, filepath: str, json_context: str, json_job: str, json_output_spec: str,
                        markdown_first: Optional[bool] = None, json_output_schema: str | None = "") -> list[dict]:
        """
        Convert a document to JSON.

        :param filepath: Path to the file.
        :type filepath: str
        :param json_context: Context for the LLM prompt used in JSON conversion (e.g., "The file contains a survey
          instrument administered by trained enumerators to households in Zimbabwe."). (Required for JSON output.)
        :type json_context: str
        :param json_job: Description of the job to do for the LLM prompt used in JSON conversion (e.g., "Your job is to
          extract each question or form field included in the text or page given."). (Required for JSON output.)
        :type json_job: str
        :param json_output_spec: JSON output specification for the LLM prompt (e.g., "Respond in correctly-formatted
          JSON with a single key named `questions` that is a list of dicts, one for each question or form field, each
          with the keys listed below..."). (Required for JSON output.)
        :type json_output_spec: str
        :param markdown_first: Whether to convert to Markdown first and then to JSON using an LLM. Set this to true if
          page-by-page conversion is not working well for elements that span pages; the Markdown-first approach will
          convert page-by-page to Markdown and then convert to JSON as the next step. The default is None, which will
          use the Markdown path for small PDF files and the page-by-page path for larger ones.
        :type markdown_first: Optional[bool]
        :param json_output_schema: JSON schema for output validation. Defaults to "", which auto-generates a validation
          schema based on the json_output_spec. If explicitly set to None, will skip JSON validation.
        :type json_output_schema: str
        :return: List of dicts, one for each batch (e.g., for each page). Use the merge_dicts() function to combine
          these into a single dict.
        :rtype: list[dict]
        """

        if markdown_first is None:
            # figure out whether we should convert to Markdown page-by-page and then JSON all in one go or convert
            # to JSON page-by-page directly

            # extract file extension
            ext = os.path.splitext(filepath)[1].lower()

            # if we're going to convert from PDF using an LLM, we need to figure out the right choice
            if self.llm_interface and (ext == '.pdf' or ext in DocumentInterface.via_pdf_file_extensions):
                # first convert to Markdown without LLM assistance
                doc_interface_no_llm = DocumentInterface()
                markdown = doc_interface_no_llm.convert_to_markdown(filepath)
                markdown_tokens = self.llm_interface.count_tokens(markdown)

                if markdown_tokens <= self.max_json_via_markdown_chunk_tokens:
                    # if we can do JSON conversion all in one shot, use doc->Markdown->JSON path
                    markdown_first = True
                else:
                    # otherwise, use doc->JSON path (the page-by-page approach)
                    markdown_first = False
            else:
                # for other file types or without an LLM, they always use Markdown first anyway
                markdown_first = True

        # use internal conversion function
        return self._convert(filepath, to_format="json" if not markdown_first else "mdjson",
                             json_context=json_context, json_job=json_job, json_output_spec=json_output_spec,
                             json_output_schema=json_output_schema)

    def _convert(self, filepath: str, to_format: str = "md", json_context: str = "", json_job: str = "",
                 json_output_spec: str = "", json_output_schema: str | None = "",
                 use_text: bool = False) -> str | list[dict]:
        """
        Convert a document to Markdown or JSON.

        :param filepath: Path to the file.
        :type filepath: str
        :param to_format: Format to convert to ("md" for Markdown, "json" for JSON, or "mdjson" for JSON from Markdown).
          Default is "md" for Markdown. The "mdjson" option is a special case that converts to Markdown first and then
          to JSON using an LLM. In contrast, the "json" option converts directly to JSON using an LLM when it can,
          bypassing the Markdown step (but when it does, it processes the document page-by-page, which can lead to
          worse results if elements span pages). Note that all JSON conversion requires an LLM interface be passed to
          the DocumentInterface constructor.
        :type to_format: str
        :param json_context: Context for the LLM prompt used in JSON conversion (e.g., "The file contains a survey
          instrument administered by trained enumerators to households in Zimbabwe."). (Required for JSON output.)
        :type json_context: str
        :param json_job: Description of the job to do for the LLM prompt used in JSON conversion (e.g., "Your job is to
          extract each question or form field included in the text or page given."). (Required for JSON output.)
        :type json_job: str
        :param json_output_spec: JSON output specification for the LLM prompt (e.g., "Respond in correctly-formatted
          JSON with a single key named `questions` that is a list of dicts, one for each question or form field, each
          with the keys listed below..."). (Required for JSON output.)
        :type json_output_spec: str
        :param json_output_schema: JSON schema for output validation. Defaults to "", which auto-generates a validation
          schema based on the json_output_spec. If explicitly set to None, will skip JSON validation.
        :type json_output_schema: str
        :param use_text: Whether to use extracted text to help the LLM with extracting text from the images. Default
            is False.
        :type use_text: bool
        :return: Markdown output or list of dicts containing JSON results.
        :rtype: str | list[dict]
        """

        # validate parameters
        if to_format not in ["md", "json", "mdjson"]:
            raise ValueError("Invalid 'to_format' parameter; must be 'md' for Markdown, 'json' for JSON, or 'mdjson' "
                             "for JSON via Markdown.")
        if to_format in ["json", "mdjson"] and (not json_context or not json_job or not json_output_spec):
            raise ValueError("For JSON output, 'context', 'job_to_do', and 'output_format' parameters are required.")
        if to_format in ["json", "mdjson"] and self.llm_interface is None:
            raise ValueError("LLM interface required for JSON output.")

        # extract file extension
        ext = os.path.splitext(filepath)[1].lower()

        if ext == '.md':
            # if the input file is already Markdown, read it as-is
            with open(filepath, 'r', encoding='utf-8') as file:
                markdown = file.read()
        else:
            # otherwise, if we have an LLM interface, use it when we can
            if self.llm_interface is not None:
                # always convert PDFs with the LLM
                if ext == '.pdf':
                    # convert PDF using LLM
                    pdf_converter = PDFDocumentConverter(llm_interface=self.llm_interface,
                                                         pdf_image_dpi=self.pdf_image_dpi,
                                                         pdf_image_max_bytes=self.pdf_image_max_bytes,
                                                         max_parallel_requests=self.max_parallel_requests)
                    if to_format == "md":
                        # convert to Markdown
                        return pdf_converter.pdf_to_markdown(filepath, use_text=use_text)
                    elif to_format == "json":
                        # convert directly to JSON
                        return pdf_converter.pdf_to_json(filepath, json_context, json_job, json_output_spec,
                                                         json_output_schema, use_text=use_text)
                    else:
                        # convert to Markdown and then to JSON
                        markdown = pdf_converter.pdf_to_markdown(filepath, use_text=use_text)
                        return self.markdown_to_json(markdown, json_context, json_job, json_output_spec,
                                                     json_output_schema)

                # convert certain other file types to PDF to then convert with the LLM
                if ext in DocumentInterface.via_pdf_file_extensions:
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # convert to PDF in temporary directory
                        pdf_path = self.convert_to_pdf(filepath, temp_dir)
                        # convert PDF using LLM
                        pdf_converter = PDFDocumentConverter(llm_interface=self.llm_interface,
                                                             pdf_image_dpi=self.pdf_image_dpi,
                                                             pdf_image_max_bytes=self.pdf_image_max_bytes,
                                                             max_parallel_requests=self.max_parallel_requests)
                        if to_format == "md":
                            # convert to Markdown
                            return pdf_converter.pdf_to_markdown(pdf_path, use_text=use_text)
                        elif to_format == "json":
                            # convert directly to JSON
                            return pdf_converter.pdf_to_json(pdf_path, json_context, json_job, json_output_spec,
                                                             json_output_schema, use_text=use_text)
                        else:
                            # convert to Markdown and then to JSON
                            markdown = pdf_converter.pdf_to_markdown(pdf_path, use_text=use_text)
                            return self.markdown_to_json(markdown, json_context, json_job, json_output_spec,
                                                         json_output_schema)

            # if Excel, see if we can convert to Markdown using our custom converter
            if ext == '.xlsx':
                # convert Excel to Markdown using custom converter
                # (try to keep images and charts if we have an LLM available and we're after Markdown output)
                result, markdown = (ExcelDocumentConverter.convert_excel_to_markdown
                                    (filepath,
                                     lose_unsupported_content=(not (self.llm_interface and to_format == "md"))))
                if result:
                    if to_format == "json":
                        # if we're after JSON, convert the Markdown to JSON using the LLM
                        return self.markdown_to_json(markdown, json_context, json_job, json_output_spec,
                                                     json_output_schema)
                    else:
                        # otherwise, just return the Markdown
                        return markdown
                else:
                    # log reason from returned Markdown
                    logging.info(f"Failed to convert {filepath} to Markdown: {markdown}")

                    # if we have an LLM and we're after Markdown, PDF it and then convert with LLM if we can
                    # (we don't want to use an LLM on Excel files headed for JSON)
                    if self.llm_interface is not None and to_format == "md":
                        with (tempfile.TemporaryDirectory() as temp_dir):
                            # convert to PDF in temporary directory
                            pdf_path = self.convert_to_pdf(filepath, temp_dir)

                            # check number of PDF pages and only move forward with LLM conversion if within the limit
                            doc = fitz.open(pdf_path)
                            if len(doc) <= DocumentInterface.max_xlsx_via_pdf_pages:
                                # convert PDF to Markdown using LLM
                                pdf_converter = PDFDocumentConverter(llm_interface=self.llm_interface,
                                                                     pdf_image_dpi=self.pdf_image_dpi,
                                                                     pdf_image_max_bytes=self.pdf_image_max_bytes,
                                                                     max_parallel_requests=self.max_parallel_requests)
                                return pdf_converter.pdf_to_markdown(pdf_path, use_text=use_text)
                            else:
                                logging.info(f"{filepath} converted to {len(doc)} pages, which is over the limit "
                                             f"({DocumentInterface.max_xlsx_via_pdf_pages}); converting without images "
                                             f"or charts...")
                                result, markdown = (ExcelDocumentConverter.convert_excel_to_markdown
                                                    (filepath, lose_unsupported_content=True))
                                if result:
                                    return markdown
                                else:
                                    # log reason from returned Markdown
                                    # then fall through to let Unstructured have a try at it
                                    logging.info(f"Failed to convert {filepath} to Markdown: {markdown}")

            # otherwise, fall back to converting using Docling, PyMuPDF4LLM, or Unstructured (in that preference order)
            markdown = ""
            if ext in DocumentInterface.docling_file_extensions:
                # try to convert using Docling
                try:
                    doc_converter = DocumentConverter()
                    markdown = doc_converter.convert(filepath).document.export_to_markdown()
                except Exception as e:
                    logging.warning(f"Error converting {filepath} using Docling: {e}")
            if not markdown and ext in DocumentInterface.pymupdf_file_extensions:
                # try to convert using PyMuPDF4LLM
                try:
                    markdown = pymupdf4llm.to_markdown(filepath)
                except Exception as e:
                    logging.warning(f"Error converting {filepath} using PyMuPDF4LLM: {e}")
            if not markdown:
                # otherwise, use Unstructured to convert
                doc_converter = UnstructuredDocumentConverter()
                markdown = doc_converter.convert_to_markdown(filepath)

        if to_format in ["json", "mdjson"]:
            # if we're after JSON, convert the Markdown to JSON using the LLM
            return self.markdown_to_json(markdown, json_context, json_job, json_output_spec, json_output_schema)
        else:
            # otherwise, just return the Markdown
            return markdown

    @staticmethod
    def convert_to_pdf(filepath: str, output_dir: str) -> str:
        """
        Convert a document to PDF using LibreOffice.

        :param filepath: Path to the document file.
        :type filepath: str
        :param output_dir: Path to the output directory.
        :type output_dir: str
        :return: Path to the converted PDF file. Throws exception on failure.
        :rtype: str
        """

        # confirm that it's a format we can convert (i.e., a supported LibreOffice format)
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in DocumentInterface.libreoffice_file_extensions:
            raise ValueError(f"Unsupported file format: {ext}. convert_to_pdf() only support formats that LibreOffice "
                             f"supports.")

        # call to LibreOffice to convert the document to PDF format
        subprocess.run([
            'soffice', '--headless', '--convert-to', 'pdf',
            filepath, '--outdir', output_dir
        ], check=True)

        # return path to the converted PDF file
        return os.path.join(output_dir, os.path.splitext(os.path.basename(filepath))[0] + '.pdf')

    def markdown_to_json(self, markdown: str, json_context: str, json_job: str, json_output_spec: str,
                         json_output_schema: str | None = "", max_chunk_size: int = 0,
                         min_chunk_size: int = 2000) -> list[dict]:
        """
        Convert Markdown text to JSON using an LLM. If needed, will automatically split text into chunks and process
        each separately. Returns a list of dicts with JSON results, one for each chunk.

        :param markdown: Markdown text to convert to JSON.
        :type markdown: str
        :param json_context: Context for the LLM prompt (e.g., "The file contains a survey instrument administered by
          trained enumerators to households in Zimbabwe.").
        :type json_context: str
        :param json_job: Job to do for the LLM prompt (e.g., "Your job is to extract each question or form field
          included in the text.").
        :type json_job: str
        :param json_output_spec: Output format for the LLM prompt (e.g., "Respond in correctly-formatted JSON with a
          single key named `questions` that is a list of dicts, one for each question or form field, each with the keys
          listed below...").
        :type json_output_spec: str
        :param json_output_schema: JSON schema for output validation. Defaults to "", which auto-generates a validation
          schema based on the json_output_spec. If explicitly set to None, will skip JSON validation.
        :type json_output_schema: str
        :param max_chunk_size: Maximum number of tokens allowed per chunk of Markdown processed. Default is 0, which
            will use a default based on the LLM's maximum output tokens.
        :type max_chunk_size: int
        :param min_chunk_size: Minimum number of desired tokens in a chunk of Markdown processed. Default is 2000.
            Set to -1 to disable.
        :type min_chunk_size: int
        :return: List of dicts with JSON results, one for each chunk. Use the merge_dicts() function to combine
            these into a single dict.
        :rtype: list[dict]
        """

        # require LLM interface to continue
        if self.llm_interface is None:
            raise ValueError("LLM interface required for JSON conversion")

        # default max_chunk_size if not specified
        if not max_chunk_size:
            max_chunk_size = self.max_json_via_markdown_chunk_tokens

        # handle automatic schema generation
        json_validation_desc = ""
        if json_output_schema is None:
            # explicitly skip schema validation
            json_output_schema = ""
        elif not json_output_schema:
            # auto-generate validation schema based on description
            json_validation_desc = json_output_spec

        # handle processing of Markdown chunks in batches
        def process_chunks_in_batches(chunks, max_parallel_requests):
            results = []

            def process_chunk(chunk):
                json_prompt = f"""Consider the Markdown text below, which has been extracted from a file.

{json_context}

{json_job}

{json_output_spec}

Markdown text enclosed by |@| delimiters:

|@|{chunk}|@|

Your JSON response precisely following the instructions given above the Markdown text:"""

                response_dict, response_text, error = self.llm_interface.get_json_response(
                    prompt=json_prompt,
                    json_validation_schema=json_output_schema,
                    json_validation_desc=json_validation_desc
                )

                if error:
                    logging.error(f"Error extracting JSON from Markdown: {error}")
                    raise RuntimeError(f"Error extracting JSON from Markdown: {error}")

                logging.info(f"Extracted JSON from Markdown: {json.dumps(response_dict, indent=2)}")
                return response_dict

            with concurrent.futures.ThreadPoolExecutor(max_workers=max_parallel_requests) as executor:
                future_to_chunk = {executor.submit(process_chunk, chunk): i for i, chunk in enumerate(chunks)}
                for future in concurrent.futures.as_completed(future_to_chunk):
                    index = future_to_chunk[future]
                    result = future.result()
                    results.append((index, result))

            results.sort(key=lambda x: x[0])
            return [result for _, result in results]

        # split into chunks as needed, then process the chunks in batches
        markdown_chunks = self.split_markdown(markdown, max_tokens=max_chunk_size, min_tokens=min_chunk_size)
        result_list = process_chunks_in_batches(markdown_chunks, self.max_parallel_requests)

        # return results
        return result_list

    @staticmethod
    def markdown_to_text(markdown: str) -> str:
        """
        Convert Markdown text to plain text by removing formatting.

        :param markdown: Input Markdown text to be converted.
        :type markdown: str
        :return: Plain text with Markdown formatting removed.
        :rtype: str
        """

        # remove lines that have only "BOXOUT:", "FOOTNOTE:", or "OTHER:"
        markdown = re.sub(r'^BOXOUT:.*$', '', markdown, flags=re.MULTILINE)
        markdown = re.sub(r'^FOOTNOTE:.*$', '', markdown, flags=re.MULTILINE)
        markdown = re.sub(r'^OTHER:.*$', '', markdown, flags=re.MULTILINE)

        # convert Markdown to HTML, then to text
        html = mdpackage.markdown(markdown, extensions=['tables'])
        soup = BeautifulSoup(html, 'html.parser')
        text = soup.get_text()

        return text

    def split_markdown(self, text: str, max_tokens: int, min_tokens: int = 2000) -> List[str]:
        """
        Split Markdown text into chunks based on token count and document structure.

        This function provides a convenient interface to the MarkdownSplitter class, creating chunks that respect both
        markdown structure and token limits.

        :param text: The Markdown text to split
        :type text: str
        :param max_tokens: Maximum number of tokens allowed per chunk
        :type max_tokens: int
        :param min_tokens: Minimum number of desired tokens in a chunk. Default is 2000. Set to -1 to disable.
        :type min_tokens: int
        :return: List of text chunks, each within the token limit
        :rtype: List[str]
        """

        splitter = MarkdownSplitter(self.llm_interface.count_tokens, max_tokens=max_tokens, min_tokens=min_tokens)
        return splitter.split_text(text)

    @staticmethod
    def merge_dicts(dict_list: list[dict], strategy: str = 'retain') -> dict:
        """
        Merge a list of dictionaries into a single dictionary.

        :param dict_list: List of dictionaries to merge.
        :type dict_list: list[dict]
        :param strategy: Strategy for handling non-list duplicate items.
                         'retain' (default): retain the original value.
                         'overwrite': overwrite with the last value.
                         'collect': collect values into a list.
        :type strategy: str
        :return: A single merged dictionary.
        :rtype: dict
        """

        def merge(a, b):
            result = {}
            for key in set(a.keys()).union(b.keys()):
                val_a = a.get(key)
                val_b = b.get(key)

                if key in a and key in b:
                    # both dictionaries have the key
                    if isinstance(val_a, dict) and isinstance(val_b, dict):
                        # recursively merge dictionaries
                        result[key] = merge(val_a, val_b)
                    elif isinstance(val_a, list) and isinstance(val_b, list):
                        # extend lists
                        result[key] = val_a + val_b
                    elif isinstance(val_a, list):
                        # append non-list to list
                        result[key] = val_a + [val_b]
                    elif isinstance(val_b, list):
                        # prepend non-list to list
                        result[key] = [val_a] + val_b
                    else:
                        if strategy == 'overwrite':
                            # overwrite with the last value
                            result[key] = val_b
                        elif strategy == 'collect':
                            # collect both values into a list
                            result[key] = [val_a, val_b]
                        elif strategy == 'retain':
                            # retain the original value
                            result[key] = val_a
                        else:
                            raise ValueError(f"Unknown strategy: {strategy}")
                elif key in a:
                    result[key] = val_a
                else:
                    result[key] = val_b
            return result

        return reduce(merge, dict_list)


class PDFDocumentConverter:
    """Utility class for converting PDF files to Markdown."""

    # member variables
    llm_interface: LLMInterface = None
    pdf_image_dpi: int = 150
    pdf_image_max_bytes: int = 1024 * 1024 * 4  # 4MB
    max_parallel_requests: int = 5

    def __init__(self, llm_interface: LLMInterface = None, pdf_image_dpi: int = 150,
                 pdf_image_max_bytes: int = 1024 * 1024 * 4, max_parallel_requests: int = 5):
        """
        Initialize for converting PDF files.

        :param llm_interface: LLM interface for interacting with LLMs in AI workflows (defaults to None, which won't
            use an LLM to convert PDF files to Markdown).
        :type llm_interface: LLMInterface
        :param pdf_image_dpi: DPI to use for rendering PDF images. Default is 150, which is generally plenty for LLM
            applications.
        :type pdf_image_dpi: int
        :param pdf_image_max_bytes: Maximum size in bytes for an image to be processed. Default is 4MB.
        :type pdf_image_max_bytes: int
        :param max_parallel_requests: Maximum number of parallel requests to make when accessing LLM. Default is 5.
        :type max_parallel_requests: int
        """

        # if specified, remember our LLM interface
        if llm_interface:
            self.llm_interface = llm_interface

        # remember other settings
        self.pdf_image_dpi = pdf_image_dpi
        self.pdf_image_max_bytes = pdf_image_max_bytes
        self.max_parallel_requests = max_parallel_requests

    @staticmethod
    def pdf_to_images(pdf_path: str, dpi: int = 150) -> list[Image.Image]:
        """
        Convert a PDF to a list of PIL Images.

        This function opens a PDF file, renders each page as an image at the specified DPI, and returns a list of these
        images.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param dpi: DPI to use for rendering the PDF. Default is 150, which is generally plenty for LLM applications.
        :type dpi: int
        :return: List of PIL Images representing the pages within the PDF.
        :rtype: list[Image.Image]
        """

        images_with_text = PDFDocumentConverter.pdf_to_images_and_text(pdf_path, dpi=dpi)
        return [image for image, _ in images_with_text]

    @staticmethod
    def pdf_to_images_and_text(pdf_path: str, dpi: int = 150) -> list[tuple[Image.Image, str]]:
        """
        Convert a PDF to a list of PIL Images, each with extracted text.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param dpi: DPI to use for rendering the PDF. Default is 150, which is generally plenty for LLM applications.
        :type dpi: int
        :return: List of tuples, one for each page, each with an image and the text extracted from that page.
        :rtype: list[tuple[Image.Image, str]]
        """

        # open PDF
        doc = fitz.open(pdf_path)

        # calculate scaling factor based on DPI
        zoom = dpi / 72  # 72 is the default PDF DPI
        matrix = fitz.Matrix(zoom, zoom)

        # first pass to determine minimum page dimensions
        min_page_width = None
        min_page_height = None
        for page in doc:
            page_width = page.rect.width
            page_height = page.rect.height

            if min_page_width is None or page_width < min_page_width:
                min_page_width = page_width
            if min_page_height is None or page_height < min_page_height:
                min_page_height = page_height

        # tolerance for dimension comparison
        tolerance = 1e-2  # Adjust if necessary for your PDFs

        # function to check if two dimensions are approximately equal
        def approx_equal(a, b, tol=tolerance):
            return abs(a - b) / max(abs(a), abs(b)) < tol

        # second pass to process pages
        images_with_text = []
        for page in doc:
            # get page dimensions
            page_width = page.rect.width
            page_height = page.rect.height

            # render page to pixmap
            pix = get_pixmap(page, matrix=matrix, alpha=False)

            # convert to PIL Image
            img = Image.frombytes("RGB", (pix.w, pix.h), pix.samples)

            # get page content for text extraction
            page_dict = page.get_text("dict")

            # check if page is a double-page spread
            if approx_equal(page_height, min_page_height) and approx_equal(page_width, 2 * min_page_width):
                # if so, split the image into two images
                left_img = img.crop((0, 0, img.width // 2, img.height))
                right_img = img.crop((img.width // 2, 0, img.width, img.height))

                # also split the page_dict into two halves
                left_dict, right_dict = PDFDocumentConverter._split_page_dict(page_dict)

                # add to list
                images_with_text.extend([(left_img, PDFDocumentConverter._plain_text_from_page_dict(left_dict)),
                                         (right_img, PDFDocumentConverter._plain_text_from_page_dict(right_dict))])
            else:
                # add to list
                images_with_text.append((img, PDFDocumentConverter._plain_text_from_page_dict(page_dict)))

        doc.close()
        return images_with_text

    @staticmethod
    def _split_page_dict(page_dict: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """
        Split a page dictionary into left and right halves.

        :param page_dict: Page dictionary from PyMuPDF.
        :type page_dict: Dict[str, Any]
        :return: Tuple of dictionaries for the left and right halves of the page.
        :rtype: Tuple[Dict[str, Any], Dict[str, Any]]
        """

        # get page dimensions
        page_width = page_dict.get("width", 0)
        mid_x = page_width / 2

        # create new dictionaries for left and right pages
        left_dict = {
            "width": mid_x,
            "height": page_dict.get("height", 0),
            "blocks": []
        }
        right_dict = {
            "width": mid_x,
            "height": page_dict.get("height", 0),
            "blocks": []
        }

        # process each block in the original page
        for block in page_dict.get("blocks", []):
            # get block position
            block_bbox = block.get("bbox", [0, 0, 0, 0])
            block_x = block_bbox[0]  # Left edge of block
            block_width = block_bbox[2] - block_bbox[0]
            block_center = block_x + (block_width / 2)

            # create a copy of the block
            new_block = block.copy()

            # adjust block position relative to new page
            if block_center <= mid_x:
                # block belongs to left page
                left_dict["blocks"].append(new_block)
            else:
                # block belongs to right page
                new_bbox = list(block_bbox)
                new_bbox[0] -= mid_x  # Adjust x coordinates
                new_bbox[2] -= mid_x
                new_block["bbox"] = new_bbox
                right_dict["blocks"].append(new_block)

        return left_dict, right_dict

    @staticmethod
    def _plain_text_from_page_dict(page_dict: Dict[str, Any]) -> str:
        """
        Extract plain text from a page dictionary.

        :param page_dict: Page dictionary from PyMuPDF.
        :type page_dict: Dict[str, Any]
        :return: Plain text extracted from the page.
        :rtype: str
        """

        plain_text = ""
        for block in page_dict.get("blocks", []):
            if block.get("type") == 0:  # text block
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        plain_text += span.get("text", "")
                    plain_text += "\n"  # newline at the end of each line
                plain_text += "\n"  # additional newline at the end of each block
        return plain_text

    @staticmethod
    def _starts_with_heading(content: str) -> bool:
        """
        Check if the content appears to start with a heading.

        This function checks if the given content starts with a heading, either in ATX or Setext style.

        :param content: Content to check.
        :type content: str
        :return: True if the content starts with a heading, False otherwise.
        :rtype: bool
        """

        # split content into lines
        lines = content.splitlines()
        i = 0
        # skip blank lines
        while i < len(lines) and lines[i].strip() == '':
            i += 1
        if i >= len(lines):
            # if no non-blank lines, return False
            return False
        # get first non-blank line, stripping whitespace
        line = lines[i].strip()
        # check for ATX heading (starts with '#')
        if line.startswith('#'):
            # we found a heading, so return True
            return True
        # check for setext-style heading (underlines with '=' or '-')
        if i + 1 < len(lines):
            underline = lines[i + 1].strip()
            if len(underline) >= len(line) and (all(c == '-' for c in underline) or all(c == '=' for c in underline)):
                # we found a heading, so return True
                return True

        # if no heading found, return False
        return False

    @staticmethod
    def _clean_and_reorder_elements(elements: list[dict]) -> list[dict]:
        """
        Clean and reorder elements, dropping page headers and footers and reordering body text sections as needed to
        ensure uninterrupted cross-page flow within sections.

        :param elements: List of elements to clean and reorder.
        :type elements: list[dict]
        :return: Cleaned and reordered list of elements.
        :rtype: list[dict]
        """

        output_elements = []
        last_body_text_idx = None
        for element in elements:
            element_type = element.get('type')
            # drop elements with type 'page_header' or 'page_footer'
            if element_type in ['page_header', 'page_footer']:
                continue
            elif element_type == 'body_text_section':
                content = element.get('content', '')
                if PDFDocumentConverter._starts_with_heading(content):
                    # sections with headings can appear in their existing order
                    output_elements.append(element)
                    last_body_text_idx = len(output_elements) - 1
                else:
                    if last_body_text_idx is not None:
                        # body sections without headings should move to just after the last body section, if any
                        last_body_text_idx += 1
                        output_elements.insert(last_body_text_idx, element)
                    else:
                        # otherwise, if no prior body section, leave it where it is
                        output_elements.append(element)
                        last_body_text_idx = len(output_elements) - 1
            else:
                # other sections can stay in their existing order
                output_elements.append(element)

        # return trimmed and reordered list of elements
        return output_elements

    @staticmethod
    def _assemble_markdown(elements: list[dict]) -> str:
        """
        Assemble a list of elements into a single markdown output.

        :param elements: List of elements to assemble.
        :type elements: list[dict]
        :return: Markdown output.
        :rtype: str
        """

        markdown_output = ''
        for element in elements:
            element_type = element.get('type')
            content = element.get('content', '')
            if element_type == 'body_text_section':
                # just add body content with paragraph separation
                # (could be smarter about this, only adding paragraph breaks where it looks appropriate — but it's
                # not clear that over-breaking will overly affect downstream uses)
                markdown_output += content.strip() + '\n\n'
            elif element_type:
                # add other types of elements with labels and more visual separation
                markdown_output += f'***\n\n{element_type.upper()}:\n\n{content.strip()}\n\n***\n\n'
            else:
                # add other types of elements with labels and more visual separation
                markdown_output += f'***\n\n{content.strip()}\n\n***\n\n'

        # return assembled markdown output with extra newlines at the end stripped out
        return markdown_output.strip()

    def pdf_to_json(self, pdf_path: str, json_context: str, json_job: str, json_output_spec: str,
                    json_output_schema: str | None = "", use_text: bool = False) -> list[dict]:
        """
        Process a PDF file page-by-page to extract elements and output JSON text.

        This function reads a PDF file, converts pages to images, processes each image with an LLM, and assembles the
        returned elements into a single JSON output.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param json_context: Context for the LLM prompt (e.g., "The file contains a survey instrument.").
        :type json_context: str
        :param json_job: Job to do for the LLM prompt (e.g., "Your job is to extract each question or form field
          included on the page."). In this case, the job will be to process each page, one at a time.
        :type json_job: str
        :param json_output_spec: Output format for the LLM prompt (e.g., "Respond in correctly-formatted JSON with a
          single key named `questions` that is a list of dicts, one for each question or form field, each with the keys
          listed below...").
        :type json_output_spec: str
        :param json_output_schema: JSON schema for output validation. Defaults to "", which auto-generates a validation
          schema based on the json_output_spec. If explicitly set to None, will skip JSON validation.
        :type json_output_schema: str
        :param use_text: Whether to use extracted text to help the LLM with extracting text from the images. Default
          is False.
        :type use_text: bool
        :return: List of parsed results from all pages, one per page, in order.
        :rtype: list[dict]
        """

        # require LLM interface to continue
        if self.llm_interface is None:
            raise ValueError("LLM interface required for PDF to JSON conversion")

        # convert PDF to images
        images_and_text = PDFDocumentConverter.pdf_to_images_and_text(pdf_path=pdf_path, dpi=self.pdf_image_dpi)

        # handle automatic schema generation
        json_validation_desc = ""
        if json_output_schema is None:
            # explicitly skip schema validation
            json_output_schema = ""
        elif not json_output_schema:
            # auto-generate validation schema based on description
            json_validation_desc = json_output_spec

        # function to process a single page
        def process_page(i, img, txt):
            logging.log(logging.INFO, f"Processing PDF page {i + 1}: Size={img.size}, Mode={img.mode}")

            # assemble prompt
            prompt_with_image = self._image_prompt(img, json_context, json_job, json_output_spec,
                                                   txt if txt and use_text else "")

            # call out to the LLM and process the returned JSON
            response_dict, response_text, error = self.llm_interface.get_json_response(
                prompt=prompt_with_image,
                json_validation_schema=json_output_schema,
                json_validation_desc=json_validation_desc
            )

            # raise exception on error
            if error:
                logging.error(f"ERROR: Error extracting JSON from page {i + 1}: {error}")
                raise RuntimeError(f"Error extracting JSON from page {i + 1} of {pdf_path}: {error}")

            # log all returned elements
            logging.info(f"Extracted JSON for page {i + 1}: {json.dumps(response_dict, indent=2)}")

            return response_dict

        # process pages in batches
        def process_batches(images_n_text, max_parallel_requests):
            results = []

            with concurrent.futures.ThreadPoolExecutor(max_workers=max_parallel_requests) as executor:
                future_to_page = {executor.submit(process_page, i, img, txt): i for i, (img, txt) in
                                  enumerate(images_n_text)}
                for future in concurrent.futures.as_completed(future_to_page):
                    index = future_to_page[future]
                    result = future.result()
                    results.append((index, result))

            results.sort(key=lambda x: x[0])
            return [result for _, result in results]

        # process all pages in batches
        logging.log(logging.INFO, f"Processing PDF {pdf_path} from {len(images_and_text)} images")
        all_dicts = process_batches(images_and_text, self.max_parallel_requests)

        # return all results
        return all_dicts

    def _image_prompt(self, image: Image.Image, json_context: str, json_job: str, json_output_spec: str,
                      txt: str = "") -> list[dict]:
        """
        Construct an LLM prompt for processing an image.

        :param image: Image to process.
        :type image: Image.Image
        :param json_context: Context for the LLM prompt.
        :type json_context: str
        :param json_job: Job to do for the LLM prompt.
        :type json_job: str
        :param json_output_spec: Output format for the LLM prompt.
        :type json_output_spec: str
        :param txt: Text to include (optional, as extracted from the image).
        :type txt: str
        :return: LLM prompt for processing the image.
        :rtype: list[dict]
        """

        if txt:
            image_prompt = f"""Consider the attached image, which shows a single page from a PDF file.

Here is plain text we have extracted from the image, in case it's helpful (delimited by #|# delimiters), but always use the image to guide your response:

#|#{txt}#|#

{json_context}

{json_job}

{json_output_spec}

Your JSON response precisely following the instructions above:"""
        else:
            image_prompt = f"""Consider the attached image, which shows a single page from a PDF file.

{json_context}

{json_job}

{json_output_spec}

Your JSON response precisely following the instructions above:"""

        return [self.llm_interface.user_message_with_image(
                user_message=image_prompt, image=image, max_bytes=self.pdf_image_max_bytes,
                current_dpi=self.pdf_image_dpi)]

    def pdf_to_markdown(self, pdf_path: str, use_text: bool = False) -> str:
        """
        Process a PDF file to extract elements and output Markdown text.

        This function reads a PDF file, converts it to images, processes each image with an LLM, and assembles the
        returned elements into a single markdown output. If no LLM is available, the function falls back to PyMuPDFLLM
        for Markdown conversion.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param use_text: Whether to use extracted text to help the LLM with extracting text from the images. Default
            is False.
        :type use_text: bool
        :return: Markdown text.
        :rtype: str
        """

        # check for LLM interface
        if self.llm_interface is None:
            # since no LLM interface, try using Docling to convert PDF to Markdown
            try:
                converter = DocumentConverter()
                result = converter.convert(pdf_path)
                return result.document.export_to_markdown()
            except Exception as e:
                logging.warning(f"Error converting {pdf_path} using Docling: {e}")

            # if that failed, fall back to PyMuPDF4LLM
            return pymupdf4llm.to_markdown(pdf_path)

        # otherwise, we'll use the LLM to process the PDF
        json_context = "The page might include a mix of text, tables, figures, charts, images, or other elements."

        json_job = f"""Your job is to:

First, scan the image to identify each distinct element in the image, where each element is a part of the page that can be handled separately from the other parts of the page. Elements include, for example:

   1. The main body text (if any), possibly separated into sections. This is the primary text of the page, which can begin on prior pages and/or continue on to future pages.

   2. Boxout text (if any). These might be sidebars, callout boxes, or other separated sub-sections that are self-contained within the page.

   3. Tables (if any). These might include a title, the table itself, and, possibly, end notes or captions.

   4. Charts or graphs (if any). These might include a title, the chart or graph itself, and, possibly, notes or captions just beneath the chart or graph.

   5. Image or figure (if any). These might include a title, the image or figure itself, and, possibly, notes or captions just beneath the image or figure.

   6. Footnotes (if any). These should include a note number or letter as well as text.

   7. Page headers and footers (if any). These are the thin, one-or-two-line headers or footers that tend to appear at the top or bottom of pages, often with a title and page number. (Do not consider larger, more-substantive headers or footers with real page content to be headers and footers, but rather part of the main body text.)

   8. Watermarks or other background images or design elements that are purely decorative and are not needed to understand the meaning of the page content (if any). These you want to ignore completely.

   9. Other (if any). Any other content that doesn't fit into one of the categories above.

Then, respond in correctly-formatted JSON according to the format described below."""

        json_output_spec = f"""Your JSON response should include a single key named `elements` that is a list of dicts, one for each element. Each of these element-specific dicts should include the keys listed below. These elements should be ordered as a human reader is meant to read them (generally from left to right and top to bottom, but might vary depending on the visual layout of the page or pages).

   1. `type` (string): This must be one of the following values, according to the element type descriptions above.

      a. "body_text_section": One section of main body text (or all of the main body text on the page when there are no clear section breaks). Don't forget to capture the page or section heading, if any, which might be stylized in some way.
      b. "boxout": One section of boxout text
      c. "table": One table
      d. "chart": One chart or graph
      e. "image": One image or figure (but remember that watermarks, backgrounds, and purely-decorative images should be ignored and not considered page elements in your JSON output)
      f. "footnote": One footnote
      g. "page_header": One page header
      h. "page_footer": One page footer
      i. "other": One other element that doesn't fit into one of the above categories

   2. `content` (string): This is the content of the element, in markdown format. The content should depend on the `type` as follows:

      a. "body_text_section": A markdown version of the text in the section, beginning with the section header (if any). The text should be verbatim, with no omissions, additions, or revisions other than to format the text in appropriate markdown syntax and remove soft hyphens that were added at the ends of lines when possible (and when the words are not naturally hyphenated). Do not add hyperlink formatting to the markdown.
      b. "boxout": A markdown version of the text in the section, beginning with the section header (if any). The text should be verbatim, with no omissions, additions, or revisions other than to format the text in appropriate markdown syntax and remove soft hyphens that were added at the ends of lines when possible (and when the words are not naturally hyphenated). Do not add hyperlink formatting to the markdown.
      c. "table": A markdown version of the complete table, including any title, column and row labels, cell text or data, and any end notes.
      d. "chart": Describe the chart or graph as if to a blind person, using markdown text and exact details and numbers whenever possible. Be sure to include any title, labels, notes, captions, or other information presented with the chart or graph (generally just above, below, or to the side of the chart or graph). Approximate numeric values from the visual elements, using the axes, in order to report the approximate numeric scale of features in the graph or chart.
      e. "image": Describe the image or figure as if to a blind person, using markdown text and exact details and numbers whenever possible. Be sure to include any title, labels, notes, captions, or other information presented with the image or figure (generally just above, below, or to the side of the image or figure). Do not add hyperlink formatting to the markdown.
      f. "footnote": Markdown text with the exact footnote, including the number or label identifying the footnote.
      g. "page_header": Markdown text with the exact header.
      h. "page_footer": Markdown text with the exact footer.
      i. "other": Markdown text with the content of the element.

Be sure to follow these JSON instructions faithfully, returning a single `elements` object list (each with `type` and `content` keys)."""

        json_output_schema = """{
  "$schema": "http://json-schema.org/draft-07/schema#",
  "type": "object",
  "properties": {
    "elements": {
      "type": "array",
      "items": {
        "type": "object",
        "properties": {
          "type": {
            "type": "string",
            "enum": [
              "body_text_section",
              "boxout",
              "table",
              "chart",
              "image",
              "footnote",
              "page_header",
              "page_footer",
              "other"
            ]
          },
          "content": {
            "type": "string"
          }
        },
        "required": ["type", "content"],
        "additionalProperties": false
      }
    }
  },
  "required": ["elements"],
  "additionalProperties": false
}"""

        # process PDF to JSON
        all_dicts = self.pdf_to_json(pdf_path, json_context, json_job, json_output_spec, json_output_schema,
                                     use_text=use_text)

        # aggregate all elements into a single list
        all_elements = []
        for d in all_dicts:
            all_elements.extend(d['elements'])

        # drop all headers and footers, re-order body text to flow continuously within sections
        all_elements = self._clean_and_reorder_elements(all_elements)

        # assemble and return markdown output
        all_markdown = self._assemble_markdown(all_elements)
        return all_markdown


class ExcelDocumentConverter:
    """Utility class to convert Excel files Markdown tables (if they don't have any images or charts)."""

    class ExcelContent:
        """Class for representing Excel file content."""

        @dataclass
        class TableRange:
            """Represents a contiguous table range in a worksheet."""

            start_row: int
            end_row: int
            start_col: int
            end_col: int
            has_header: bool = True
            is_pivot_table: bool = False

        # initialize class-level member variables
        wb: Workbook = None
        filepath: str = None

        def __init__(self, filepath: str):
            """
            Initialize the Excel content object.

            :param filepath: Path to the Excel file.
            :type filepath: str
            """

            self.wb = load_workbook(filepath, data_only=True)
            self.filepath = filepath

        def has_unsupported_content(self) -> Tuple[bool, str]:
            """
            Check if workbook contains content that we don't support.
            Only checks for images and charts, allowing all other formatting to be quietly lost
            in conversion.

            :return: Tuple indicating if PDF conversion is needed and the reason why.
            :rtype: Tuple[bool, str]
            """

            for sheet in self.wb.worksheets:
                # check for images
                # noinspection PyProtectedMember
                if hasattr(sheet, '_images') and sheet._images:
                    return True, f"Sheet '{sheet.title}' contains images"

                # check for charts
                # noinspection PyProtectedMember
                if hasattr(sheet, '_charts') and sheet._charts:
                    return True, f"Sheet '{sheet.title}' contains charts"

            return False, "Content is suitable for direct markdown conversion"

        @staticmethod
        def _is_pivot_table(sheet: Worksheet, table_range: TableRange) -> bool:
            """
            Detect if a table range is likely a pivot table based on characteristics.
            This function uses a combination of heuristics to detect pivot tables in Excel.

            :param sheet: Worksheet object containing the table range.
            :type sheet: Worksheet
            :param table_range: TableRange object representing the table range.
            :type table_range: TableRange
            :return: True if the table range is likely a pivot table, False otherwise.
            :rtype: bool
            """

            # check if sheet has pivot tables defined
            if hasattr(sheet, 'pivotTables') and sheet.pivotTables:
                # check if our range intersects with any pivot table range
                for pivot in sheet.pivotTables:
                    if (pivot.location.min_row <= table_range.end_row and
                            pivot.location.max_row >= table_range.start_row and
                            pivot.location.min_col <= table_range.end_col and
                            pivot.location.max_col >= table_range.start_col):
                        return True

            # use fallback detection based on pivot table characteristics
            # noinspection PyBroadException
            try:
                first_row = list(sheet.iter_rows(
                    min_row=table_range.start_row,
                    max_row=table_range.start_row,
                    min_col=table_range.start_col,
                    max_col=table_range.end_col
                ))[0]

                # check for common pivot table indicators
                pivot_indicators = ['Total', 'Grand Total', 'Sum of', 'Count of', 'Average of']
                cell_values = [str(cell.value).strip() if cell.value else '' for cell in first_row]
                if any(any(indicator in value for indicator in pivot_indicators)
                       for value in cell_values):
                    return True
            except:
                pass

            return False

        @staticmethod
        def find_tables(sheet: Worksheet) -> List[TableRange]:
            """
            Identify contiguous table ranges in a worksheet.

            :param sheet: Worksheet object to analyze.
            :type sheet: Worksheet
            :return: List of TableRange objects representing the identified table ranges.
            :rtype: List[TableRange]
            """

            tables = []
            current_table = None

            # first check for explicitly defined tables
            if hasattr(sheet, 'tables'):
                for table in sheet.tables.values():
                    table_range = ExcelDocumentConverter.ExcelContent.TableRange(
                        start_row=table.ref.min_row,
                        end_row=table.ref.max_row,
                        start_col=table.ref.min_col,
                        end_col=table.ref.max_col,
                        has_header=True  # (Excel tables always have headers)
                    )
                    tables.append(table_range)

            # Then look for implicit tables in the data
            data_rows = list(sheet.rows)
            if not data_rows:
                return tables

            for row_idx, row in enumerate(data_rows, 1):
                row_empty = all(cell.value is None for cell in row)
                row_merged = any(isinstance(cell, MergedCell) for cell in row)

                if not row_empty and not row_merged:
                    # find the start and end columns for this row
                    start_col = None
                    end_col = None
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value is not None:
                            if start_col is None:
                                start_col = col_idx
                            end_col = col_idx

                    if start_col is not None:
                        if current_table is None:
                            # start new table
                            current_table = ExcelDocumentConverter.ExcelContent.TableRange(
                                start_row=row_idx,
                                end_row=row_idx,
                                start_col=start_col,
                                end_col=end_col
                            )
                        else:
                            # extend current table
                            current_table.end_row = row_idx
                            current_table.start_col = min(current_table.start_col, start_col)
                            current_table.end_col = max(current_table.end_col, end_col)
                else:
                    # empty row - close current table if it exists
                    if current_table is not None:
                        # check if it overlaps with any existing tables
                        if not any(ExcelDocumentConverter.ExcelContent._ranges_overlap(current_table, t)
                                   for t in tables):
                            tables.append(current_table)
                        current_table = None

            # add final table if exists and doesn't overlap
            if (current_table is not None and
                    not any(ExcelDocumentConverter.ExcelContent._ranges_overlap(current_table, t) for t in tables)):
                tables.append(current_table)

            # analyze tables for headers and pivot tables
            for table in tables:
                table.has_header = ExcelDocumentConverter.ExcelContent._detect_header_row(sheet, table)
                table.is_pivot_table = ExcelDocumentConverter.ExcelContent._is_pivot_table(sheet, table)

            # return identified tables
            return tables

        @staticmethod
        def _ranges_overlap(range1: TableRange, range2: TableRange) -> bool:
            """
            Check if two table ranges overlap.

            :param range1: First TableRange object.
            :type range1: TableRange
            :param range2: Second TableRange object.
            :type range2: TableRange
            :return: True if the two ranges overlap, False otherwise.
            :rtype: bool
            """

            return not (range1.end_row < range2.start_row or
                        range1.start_row > range2.end_row or
                        range1.end_col < range2.start_col or
                        range1.start_col > range2.end_col)

        @staticmethod
        def _detect_header_row(sheet: Worksheet, table: TableRange) -> bool:
            """
            Detect if the first row of a table range is likely a header row.

            :param sheet: Worksheet object containing the table range.
            :type sheet: Worksheet
            :param table: TableRange object representing the table range.
            :type table: TableRange
            :return: True if the first row is likely a header row, False otherwise.
            :rtype: bool
            """

            if table.start_row == table.end_row:
                return False

            first_row = list(sheet.iter_rows(
                min_row=table.start_row,
                max_row=table.start_row,
                min_col=table.start_col,
                max_col=table.end_col
            ))[0]

            second_row = list(sheet.iter_rows(
                min_row=table.start_row + 1,
                max_row=table.start_row + 1,
                min_col=table.start_col,
                max_col=table.end_col
            ))[0]

            # check for header formatting
            header_indicators = 0

            # check for bold font in first row
            if any(cell.font and cell.font.bold for cell in first_row):
                header_indicators += 1

            # check for different font properties between first and second row
            if any(
                    first_row[i].font != second_row[i].font
                    for i in range(len(first_row))
                    if first_row[i].font and second_row[i].font
            ):
                header_indicators += 1

            # check if first row values look like headers
            first_row_values = [str(cell.value).strip() if cell.value else '' for cell in first_row]

            # headers often have different data types than the data
            if any(isinstance(first_row[i].value, str) and
                   isinstance(second_row[i].value, (int, float))
                   for i in range(len(first_row))):
                header_indicators += 1

            # headers typically don't contain empty cells
            if all(val for val in first_row_values):
                header_indicators += 1

            # consider it a header if we have at least 2 indicators
            return header_indicators >= 2

    @staticmethod
    def _excel_to_strftime_format(excel_format: str) -> str:
        """
        Convert an Excel date format to a strftime format.

        :param excel_format: Excel date format.
        :type excel_format: str
        :return: strftime format.
        :rtype: str
        """

        # define a mapping from Excel format codes to strftime format codes
        format_mapping = {
            'yyyy': '%Y',
            'yy': '%y',
            'mmmm': '%B',
            'mmm': '%b',
            'mm': '%m',  # month as a zero-padded decimal number
            'm': '%m',  # month as a zero-padded decimal number
            'dddd': '%A',
            'ddd': '%a',
            'dd': '%d',
            'd': '%d',  # day of the month as a zero-padded decimal number
            'hh': '%H',
            'h': '%H',  # hour (24-hour clock) as a zero-padded decimal number
            'ss': '%S',
            's': '%S',  # second as a zero-padded decimal number
            'AM/PM': '%p',
            'am/pm': '%p'
        }

        # create a regex pattern that matches any of the Excel format codes
        pattern = re.compile('|'.join(re.escape(key) for key in sorted(format_mapping.keys(), key=len, reverse=True)))

        # replace Excel format codes with strftime format codes
        strftime_format = pattern.sub(lambda x: format_mapping[x.group()], excel_format)

        return strftime_format

    @staticmethod
    def _format_cell_value(cell: Cell, value: Any) -> str:
        """
        Format cell value with appropriate markdown styling based on cell format.

        :param cell: Cell object to format.
        :type cell: Cell
        :param value: Value to format.
        :type value: Any
        :return: Formatted value as a string.
        :rtype: str
        """

        # if no value, just return an empty string
        if value is None:
            return ''

        # handle dates and times
        if cell.is_date and value:
            # noinspection PyBroadException
            try:
                from datetime import datetime
                if isinstance(value, datetime):
                    # do our best to format dates and times as specified in the document, otherwise use a default format
                    strftime_format = ExcelDocumentConverter._excel_to_strftime_format(cell.number_format
                                                                                       or '%Y-%m-%d %H:%M:%S')
                    return value.strftime(strftime_format)
                return str(value)
            except:
                return str(value)

        # handle numbers
        if isinstance(value, (int, float)):
            number_format = cell.number_format or 'General'
            if number_format == 'General':
                # for whole numbers, don't show decimal places
                if isinstance(value, int) or value.is_integer():
                    return str(int(value))
                return str(value)
            elif number_format.endswith('%'):
                # handle percentages
                match = re.search(r'0+(\.0+)?%', number_format)
                if match:
                    # if we find requested number of decimal places, respect it
                    decimal_places = match.group(1).count('0') if match.group(1) else 0
                    return f"{value * 100:.{decimal_places}f}%"
                else:
                    # default to 2 decimal places
                    return f"{value * 100:.2f}%"
            elif '#' in number_format or '0' in number_format:
                # try to respect decimal places specified in format (otherwise default to 2 for non-integers)
                decimal_places = 2
                match = re.search(r'0+(\.0+)?', number_format)
                if match:
                    decimal_places = match.group(1).count('0') if match.group(1) else 0
                if decimal_places == 0 or isinstance(value, int) or value.is_integer():
                    return str(int(value))
                return f"{value:.{decimal_places}f}"

            return str(value)

        # handle text by stripping extra spacing and escaping pipe characters
        value = str(value).strip()
        formatted = value.replace('|', '\\|')

        # add hyperlinks if needed
        if cell.hyperlink:
            if cell.hyperlink.target:
                formatted = f'[{formatted}]({cell.hyperlink.target})'

        # add text formatting if needed (note: order matters here!)
        if cell.font:
            if cell.font.strike:
                formatted = f'~~{formatted}~~'          # strikethrough
            if cell.font.bold:
                formatted = f'**{formatted}**'          # bold
            if cell.font.italic:
                formatted = f'*{formatted}*'            # italic
            if cell.font.vertAlign == 'superscript':
                formatted = f'<sup>{formatted}</sup>'   # superscript
            elif cell.font.vertAlign == 'subscript':
                formatted = f'<sub>{formatted}</sub>'   # subscript

        # handle multi-line text if needed (since it's going to go in a Markdown table, can't leave newlines as-is)
        if '\n' in formatted:
            formatted = formatted.replace('\n', '<br>')

        return formatted

    @staticmethod
    def _get_merge_range_for_cell(cell: Cell, merge_ranges: List[MergedCellRange]) -> MergedCellRange | None:
        """
        Get the merge range containing this cell, if any.

        :param cell: Cell object to check.
        :type cell: Cell
        :param merge_ranges: List of merged cell ranges in the worksheet.
        :type merge_ranges: List[MergedCellRange]
        :return: Merged cell range containing the cell, or None if not merged.
        :rtype: MergedCellRange | None
        """

        # can for cell in supplied merge ranges
        for merge_range in merge_ranges:
            if cell.coordinate in merge_range:
                return merge_range
        return None

    @staticmethod
    def _is_first_cell_in_merge_range(cell: Cell, merge_range: MergedCellRange) -> bool:
        """
        Check if cell is the top-left cell in its merge range.

        :param cell: Cell object to check.
        :type cell: Cell
        :param merge_range: MergedCellRange containing the cell.
        :type merge_range: MergedCellRange
        :return: True if cell is the top-left cell in the merge range, False otherwise.
        :rtype: bool
        """

        return cell.row == merge_range.min_row and cell.column == merge_range.min_col

    @staticmethod
    def _create_markdown_table(sheet: Worksheet, table_range: ExcelContent.TableRange) -> str:
        """
        Convert a table range to Markdown format.

        :param sheet: Worksheet object containing the table range.
        :type sheet: Worksheet
        :param table_range: TableRange object representing the table range.
        :type table_range: TableRange
        :return: Markdown-formatted table.
        :rtype: str
        """

        # track merge ranges that affect our table
        relevant_merges = [
            merge_range for merge_range in sheet.merged_cells.ranges
            if (merge_range.min_row <= table_range.end_row and
                merge_range.max_row >= table_range.start_row and
                merge_range.min_col <= table_range.end_col and
                merge_range.max_col >= table_range.start_col)
        ]

        # initialize
        rows = []
        merge_notes = []

        # handle pivot table headers specially
        if table_range.is_pivot_table:
            # process headers differently for pivot tables
            header_rows = []
            data_start_row = table_range.start_row

            # collect all header rows (those with merged cells or different formatting)
            curr_row = table_range.start_row
            while curr_row < table_range.end_row:
                row = list(sheet.iter_rows(
                    min_row=curr_row,
                    max_row=curr_row,
                    min_col=table_range.start_col,
                    max_col=table_range.end_col
                ))[0]

                if any(isinstance(cell, MergedCell) or
                       (cell.font and cell.font.bold) for cell in row):
                    header_rows.append(row)
                    data_start_row = curr_row + 1
                else:
                    break
                curr_row += 1

            # process header rows
            if header_rows:
                for row in header_rows:
                    row_values = []
                    for cell in row:
                        merge_range = ExcelDocumentConverter._get_merge_range_for_cell(cell, relevant_merges)
                        if merge_range:
                            if ExcelDocumentConverter._is_first_cell_in_merge_range(cell, merge_range):
                                value = sheet.cell(merge_range.min_row, merge_range.min_col).value
                                formatted_value = ExcelDocumentConverter._format_cell_value(cell, value)
                                row_values.append(formatted_value)
                            else:
                                row_values.append('')
                        else:
                            value = cell.value
                            formatted_value = ExcelDocumentConverter._format_cell_value(cell, value) \
                                if value is not None else ''
                            row_values.append(formatted_value)
                    rows.append(row_values)

            # update table range to exclude processed headers
            table_range.start_row = data_start_row
            table_range.has_header = False  # Headers already processed

        # get all cells in range
        for row in sheet.iter_rows(
                min_row=table_range.start_row,
                max_row=table_range.end_row,
                min_col=table_range.start_col,
                max_col=table_range.end_col
        ):
            row_values = []

            for cell in row:
                merge_range = ExcelDocumentConverter._get_merge_range_for_cell(cell, relevant_merges)
                if merge_range:
                    if ExcelDocumentConverter._is_first_cell_in_merge_range(cell, merge_range):
                        value = sheet.cell(merge_range.min_row, merge_range.min_col).value
                        formatted_value = ExcelDocumentConverter._format_cell_value(cell, value)
                        row_values.append(formatted_value)

                        # consider adding comment about the merge
                        span_cols = merge_range.max_col - merge_range.min_col + 1
                        span_rows = merge_range.max_row - merge_range.min_row + 1
                        if span_cols > 1 or span_rows > 1:
                            # add comment about merged cell
                            curr_row = len(rows) + 1
                            curr_col = len(row_values)
                            merge_notes.append(
                                f"* Cell at row {curr_row}, column {curr_col} " +
                                f"spans {span_rows} row{'' if span_rows == 1 else 's'} and " +
                                f"{span_cols} column{'' if span_cols == 1 else 's'}"
                            )
                    else:
                        row_values.append('')
                else:
                    value = cell.value
                    formatted_value = ExcelDocumentConverter._format_cell_value(cell, value) if value is not None \
                        else ''
                    row_values.append(formatted_value)

            # only add rows that aren't completely empty
            if any(val.strip() for val in row_values):
                rows.append(row_values)

        if not rows:
            return ''

        # create Markdown table
        md_lines = []

        # add headers (either pivot table headers or regular headers)
        header_row = rows[0]
        md_lines.append('| ' + ' | '.join(header_row) + ' |')

        # add separator
        separator = '|'
        for _ in range(len(header_row)):
            separator += ' --- |'
        md_lines.append(separator)

        # add data rows (skip first row only if it's a regular header)
        if table_range.has_header and not table_range.is_pivot_table:
            data_rows = rows[1:]
        else:
            data_rows = rows[1:] if rows else []  # (for pivot tables, first row is already a header)

        for row in data_rows:
            md_lines.append('| ' + ' | '.join(row) + ' |')

        # add merge notes (if any)
        if merge_notes:
            md_lines.extend(['', '<!-- Merged cells:', *merge_notes, '-->'])

        # combine all lines together for output
        return '\n'.join(md_lines)

    @staticmethod
    def convert_excel_to_markdown(excel_path: str, include_hidden_sheets: bool = False,
                                  lose_unsupported_content: bool = False) -> Tuple[bool, str]:
        """
        Convert Excel file to Markdown if possible, otherwise indicate PDF conversion needed.

        :param excel_path: Path to the Excel file.
        :type excel_path: str
        :param include_hidden_sheets: Whether to include hidden sheets in the conversion. Default is False.
        :type include_hidden_sheets: bool
        :param lose_unsupported_content: Whether to quietly lose unsupported content in the conversion (if False, will
          return failure when file contains images and/or charts). Default is False.
        :type lose_unsupported_content: bool
        :return: Tuple indicating if conversion was successful and the Markdown text.
        """

        try:
            excel_content = ExcelDocumentConverter.ExcelContent(excel_path)
            # drop out if we have images or charts and we're not losing them
            if not lose_unsupported_content:
                has_unsupported_content, reason = excel_content.has_unsupported_content()
                if has_unsupported_content:
                    return False, reason

            # process each sheet
            markdown_content = []

            for sheet in excel_content.wb.worksheets:
                # skip hidden sheets if desired
                if sheet.sheet_state == 'hidden' and not include_hidden_sheets:
                    continue

                markdown_content.append(f'# {sheet.title}\n')

                # find and convert tables
                tables = excel_content.find_tables(sheet)
                if not tables:
                    markdown_content.append('*No data found in this sheet*\n')
                    continue

                # group tables by row to maintain original layout
                row_grouped_tables = {}
                for table in tables:
                    row_grouped_tables.setdefault(table.start_row, []).append(table)

                # process tables in order by row
                last_end_row = 0
                for start_row in sorted(row_grouped_tables.keys()):
                    # add spacing if there's a gap between tables
                    if last_end_row > 0 and start_row > last_end_row + 1:
                        markdown_content.append('\n---\n')

                    # process tables that start on this row
                    row_tables = row_grouped_tables[start_row]

                    # sort tables by column if multiple tables on same row
                    row_tables.sort(key=lambda t: t.start_col)

                    for i, table in enumerate(row_tables):
                        if i > 0:
                            # add horizontal separator between side-by-side tables
                            markdown_content.append('\n___\n')

                        # add table title if it's a pivot table
                        if table.is_pivot_table:
                            # try to find a title from cells above the table
                            title = ExcelDocumentConverter._find_table_title(sheet, table)
                            if title:
                                markdown_content.append(f'### {title}\n')

                        # convert table to markdown
                        markdown_table = ExcelDocumentConverter._create_markdown_table(sheet, table)
                        if markdown_table:
                            markdown_content.append(f"{markdown_table}\n")

                        last_end_row = max(last_end_row, table.end_row)

            # return Markdown text
            return True, '\n'.join(markdown_content)

        except Exception as e:
            return False, f"Error processing file: {str(e)}"

    @staticmethod
    def _find_table_title(sheet: Worksheet, table: ExcelContent.TableRange) -> str | None:
        """
        Look for a title above the table by checking for merged cells or
        cells with larger/bold font.

        :param sheet: Worksheet object containing the table.
        :type sheet: Worksheet
        :param table: TableRange object representing the table.
        :type table: TableRange
        :return: Title text if found, None otherwise.
        :rtype: str | None
        """

        if table.start_row <= 1:
            return None

        # look up to 3 rows above the table
        for row in range(table.start_row - 1, max(0, table.start_row - 4), -1):
            # check cells in and around the table's column range
            start_col = max(1, table.start_col - 1)
            end_col = min(sheet.max_column, table.end_col + 1)

            for cell in sheet[row][start_col - 1:end_col]:
                # skip empty cells
                if not cell.value:
                    continue

                # check if cell is merged
                is_merged = isinstance(cell, MergedCell)
                if is_merged:
                    merge_range = next(
                        merge_range for merge_range in sheet.merged_cells.ranges
                        if cell.coordinate in merge_range
                    )
                    if merge_range.min_row == row and merge_range.min_col == cell.column:
                        master_cell = sheet.cell(merge_range.min_row, merge_range.min_col)
                        return str(master_cell.value).strip()

                # check for title-like formatting
                if cell.font:
                    font_size = cell.font.size if hasattr(cell.font, 'size') else None
                    if ((font_size and font_size > 11) or  # larger font
                            cell.font.bold or  # bold
                            (  # centered over table
                                    cell.alignment and
                                    cell.alignment.horizontal == 'center' and
                                    abs((end_col - start_col) / 2 + start_col - cell.column) <= 1
                            )):
                        return str(cell.value).strip()

        return None


class UnstructuredDocumentConverter:
    """Convert various document types to markdown using Unstructured."""

    @dataclass
    class DocumentElement:
        """Represents a processed document element."""

        type: str
        content: str
        metadata: Optional[Dict] = None
        level: Optional[int] = None

    # initialize class-level member variables
    heading_style: str = "atx"
    element_handlers: Dict[type, Callable] = {}

    def __init__(self, heading_style: str = "atx"):
        """
        Initialize the Unstructured document converter.

        :param heading_style: 'atx' for # style or 'setext' for underline style.
        :type heading_style: str
        """

        # set heading style
        self.heading_style = heading_style

        # map Unstructured element types to markdown handlers
        self.element_handlers = {
            Title: UnstructuredDocumentConverter._handle_title,
            Header: UnstructuredDocumentConverter._handle_header,
            Text: UnstructuredDocumentConverter._handle_text,
            NarrativeText: UnstructuredDocumentConverter._handle_narrative,
            ListItem: UnstructuredDocumentConverter._handle_list_item,
            Table: UnstructuredDocumentConverter._handle_table,
            ImageElement: UnstructuredDocumentConverter._handle_image,
            PageBreak: UnstructuredDocumentConverter._handle_page_break,
            Footer: UnstructuredDocumentConverter._handle_footer,
            Address: UnstructuredDocumentConverter._handle_address,
        }

    def convert_to_markdown(self, file_path: Union[str, Path]) -> str:
        """
        Convert document to Markdown format.

        :param file_path: Path to input file.
        :type file_path: Union[str, Path]
        :return: Markdown formatted string.
        :rtype: str
        """

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # pre-process .html files (to replace description list elements since Unstructured doesn't support them)
        if file_path.suffix.lower() == '.html':
            # create a temporary file to use instead of the original
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html') as temp_file:
                # load the HTML content
                with open(file_path, 'r') as file:
                    soup = BeautifulSoup(file, 'html.parser')

                # replace all description list elements
                while True:
                    dls = soup.find_all('dl')
                    if not dls:
                        break

                    for dl in dls:
                        ul = soup.new_tag('ul')
                        if dl.has_attr('class'):
                            ul['class'] = dl['class']

                        for tag in dl.find_all(['dt', 'dd'], recursive=False):
                            li = soup.new_tag('li')
                            li.extend(tag.contents)
                            ul.append(li)

                        dl.replace_with(ul)

                # save the modified HTML content to temp_file
                temp_file.write(str(soup))

                # partition document using Unstructured
                elements: List[Element] = partition(temp_file.name)
        else:
            # partition document using Unstructured
            elements: List[Element] = partition(str(file_path))

        # process elements to DocumentElement objects
        doc_elements = []
        for element in elements:
            handler = self.element_handlers.get(type(element))
            if handler:
                # noinspection PyArgumentList
                processed = handler(element)
                if processed:
                    doc_elements.extend(processed if isinstance(processed, list) else [processed])
            else:
                # default to handling as text
                processed = UnstructuredDocumentConverter.DocumentElement(
                    type="text",
                    content=element.text.strip(),
                    metadata={
                        "coordinates": getattr(element, "coordinates", None),
                        "link_urls": element.metadata.link_urls,
                        "link_texts": element.metadata.link_texts,
                        "link_start_indexes": element.metadata.link_start_indexes
                    }
                )
                doc_elements.extend([processed])

        # convert to markdown
        return self._elements_to_markdown(doc_elements)

    @staticmethod
    def _handle_title(element: Title) -> DocumentElement:
        """
        Process title elements.

        :param element: Title element to process.
        :type element: Title
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="heading",
            content=element.text.strip(),
            level=1,
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_header(element: Header) -> DocumentElement:
        """
        Process header elements.

        :param element: Header element to process.
        :type element: Header
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="heading",
            content=element.text.strip(),
            level=2,
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_text(element: Text) -> DocumentElement:
        """
        Process basic text elements.

        :param element: Text element to process.
        :type element: Text
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="text",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_narrative(element: NarrativeText) -> DocumentElement:
        """
        Process narrative text elements.

        :param element: NarrativeText element to process.
        :type element: NarrativeText
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="paragraph",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_list_item(element: ListItem) -> DocumentElement:
        """
        Process list items.

        :param element: ListItem element to process.
        :type element: ListItem
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="list_item",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "indent_level": getattr(element, "indent_level", 0),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_table(element: Table) -> DocumentElement | None:
        """
        Process table elements.

        :param element: Table element to process.
        :type element: Table
        :return: DocumentElement object or None if table is not supported.
        :rtype: DocumentElement | None
        """

        # extract table text and format as Markdown table
        table_text = element.text.strip()

        # lines are rows
        rows = table_text.split('\n')
        if not rows:
            return None

        # create Markdown table, assuming whitespace separates columns and the first row is the header
        md_table = ['| ' + ' | '.join(rows[0].split()) + ' |', '|' + '---|' * (len(rows[0].split()) - 1) + '---|']
        for row in rows[1:]:
            md_table.append('| ' + ' | '.join(row.split()) + ' |')

        return UnstructuredDocumentConverter.DocumentElement(
            type="table",
            content='\n'.join(md_table),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_image(element: ImageElement) -> DocumentElement:
        """
        Process image elements.

        :param element: ImageElement to process.
        :type element: ImageElement
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        # for now, we just add a placeholder for the image
        return UnstructuredDocumentConverter.DocumentElement(
            type="image",
            content="![Image]",
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "image_data": getattr(element, "image_data", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_page_break(element: PageBreak) -> DocumentElement:
        """
        Process page breaks.

        :param element: PageBreak element to process.
        :type element: PageBreak
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="page_break",
            content="",
            metadata={"page_number": getattr(element, "page_number", None)}
        )

    @staticmethod
    def _handle_footer(element: Footer) -> DocumentElement:
        """
        Process footer elements.

        :param element: Footer element to process.
        :type element: Footer
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="footer",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    @staticmethod
    def _handle_address(element: Address) -> DocumentElement:
        """
        Process address elements.

        :param element: Address element to process.
        :type element: Address
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="address",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "link_urls": element.metadata.link_urls,
                "link_texts": element.metadata.link_texts,
                "link_start_indexes": element.metadata.link_start_indexes
            }
        )

    def _elements_to_markdown(self, elements: List[DocumentElement]) -> str:
        """
        Convert processed elements to markdown string.

        :param elements: List of DocumentElement objects.
        :type elements: List[DocumentElement]
        :return: Markdown formatted string.
        :rtype: str
        """

        # keep track as we process each element
        markdown_parts = []
        list_stack = []
        for element in elements:
            if element.type == "heading":
                # clear any active lists
                list_stack = []

                # add heading
                if self.heading_style == "atx":
                    markdown_parts.append(f"{'#' * element.level} {self.content_with_links(element)}")
                else:
                    markdown_parts.append(self.content_with_links(element))
                    markdown_parts.append('=' if element.level == 1 else '-' * len(self.content_with_links(element)))
            elif element.type == "paragraph":
                # clear any active lists
                list_stack = []

                # add paragraph content
                markdown_parts.append(self.content_with_links(element))
            elif element.type == "list_item":
                indent_level = element.metadata.get("indent_level", 0)
                # adjust list stack
                while len(list_stack) > indent_level:
                    list_stack.pop()
                while len(list_stack) < indent_level:
                    list_stack.append(0)

                # add list item
                prefix = "  " * indent_level + "* "
                markdown_parts.append(f"{prefix}{self.content_with_links(element)}")
            elif element.type == "table":
                # clear any active lists
                list_stack = []

                # add table content
                markdown_parts.append(self.content_with_links(element))
            elif element.type == "image":
                # clear any active lists
                list_stack = []

                # add image content
                markdown_parts.append(self.content_with_links(element))
            elif element.type == "page_break":
                # add page break
                markdown_parts.append("\n---\n")
            elif element.type == "footer":
                # add footer content
                markdown_parts.append(f"\n---\n{self.content_with_links(element)}\n")
            elif element.type == "address":
                # add address content
                markdown_parts.append(f"> {self.content_with_links(element)}")
            else:
                # add any other content as-is
                markdown_parts.append(self.content_with_links(element))

            # add spacing between elements
            markdown_parts.append("")

        # return combined markdown content
        return "\n".join(markdown_parts)

    @staticmethod
    def content_with_links(element: DocumentElement) -> str:
        """
        Convert content to Markdown with links as needed.

        :param element: DocumentElement object.
        :type element: DocumentElement
        :return: String with content, including hyperlinks.
        :rtype: str
        """

        # provisional return value is just the content
        retval = element.content

        # add hyperlinks, if present
        if ("link_urls" in element.metadata and element.metadata["link_urls"] and "link_texts" in element.metadata
                and element.metadata["link_texts"]):
            # see if we have start indexes for the hyperlinks
            if "link_start_indexes" in element.metadata and element.metadata["link_start_indexes"]:
                # combine link data and sort by start index in descending order
                links = sorted(
                    zip(element.metadata["link_start_indexes"],
                        element.metadata["link_texts"],
                        element.metadata["link_urls"]),
                    reverse=True
                )
                # replace each link text with Markdown link syntax
                for start_index, link_text, link_url in links:
                    # only replace absolute http and mailto links for now
                    if link_url.startswith("https:") or link_url.startswith("http:") or link_url.startswith("mailto:"):
                        end_index = start_index + len(link_text)
                        markdown_link = f"[{link_text}]({link_url})"
                        retval = retval[:start_index] + markdown_link + retval[end_index:]
            else:
                # if no start indexes, just use blind text replacement (which is error-prone)
                for link_text, link_url in zip(element.metadata["link_texts"], element.metadata["link_urls"]):
                    if link_url.startswith("https:") or link_url.startswith("http:") or link_url.startswith("mailto:"):
                        # replace the link text with Markdown hyperlink syntax only if it's not already hyperlinked
                        markdown_link = f"[{link_text}]({link_url})"
                        if not re.search(rf'\[{link_text}]\([^)]+\)', retval):
                            retval = retval.replace(link_text, markdown_link)

        return retval


class MarkdownSplitter:
    """Split Markdown text into chunks while preserving document structure."""

    def __init__(self, count_tokens: Callable[[str], int], max_tokens: int, min_tokens: int = 2000):
        """
        Initialize the Markdown splitter with token counting function and maximum tokens.

        This class splits Markdown text into chunks while preserving the document structure and ensuring each chunk
        stays within a specified token limit.

        :param count_tokens: Function that takes a string and returns its token count
        :type count_tokens: Callable[[str], int]
        :param max_tokens: Maximum number of tokens allowed per chunk
        :type max_tokens: int
        :param min_tokens: Minimum number of tokens desired per chunk. Defaults to 2000. Set to -1 to disable.
        :type min_tokens: int
        """

        self.count_tokens = count_tokens
        self.max_tokens = max_tokens
        self.min_tokens = min_tokens

    def split_text(self, text: str) -> List[str]:
        """
        Split Markdown text recursively according to heading hierarchy and structure.

        This function splits text using a hierarchical approach, starting with highest level headers and progressively
        moving to finer-grained splits until all chunks are within the token limit.

        :param text: The Markdown text to split
        :type text: str
        :return: List of text chunks, each within the token limit
        :rtype: List[str]
        """

        # start with one big chunk and return it right away if it's within the limit
        chunks = [text]
        if self.count_tokens(text) <= self.max_tokens:
            return [text]

        # rank order our splitting strategies
        splitting_strategies = [
            self._split_by_pattern(r'^# '),  # h1
            self._split_by_pattern(r'^## '),  # h2
            self._split_by_pattern(r'^### '),  # h3
            self._split_by_pattern(r'^#### '),  # h4
            self._split_by_pattern(r'^##### '),  # h5
            self._split_by_pattern(r'^\*\*[^*]+\*\*$'),  # bold headers
            self._split_by_pattern(r'^(?:\*\*\*|---)$'),  # horizontal rules
            self._split_by_paragraphs,  # paragraphs
            self._split_by_lines,  # lines
            self._split_by_tokens  # token-level split
        ]

        # apply each strategy one at a time, only to chunks that are over the limit
        for strategy in splitting_strategies:
            new_chunks = []
            for chunk in chunks:
                if self.count_tokens(chunk) <= self.max_tokens:
                    new_chunks.append(chunk)
                else:
                    new_chunks.extend(strategy(chunk))
            chunks = new_chunks

            # break out of the loop once everybody is within the limit
            if all(self.count_tokens(chunk) <= self.max_tokens for chunk in chunks):
                break

        # now run back through and merge adjacent chunks when they're too small and merging won't put a chunk over the
        # limit
        merged_chunks = []
        prior_chunk_tokens = 0
        prior_chunk_prefers_merge = False
        for chunk in chunks:
            chunk_tokens = self.count_tokens(chunk)
            # if prior chunk was hoping to merge, merge if possible
            if prior_chunk_prefers_merge and prior_chunk_tokens + chunk_tokens <= self.max_tokens:
                merged_chunks[-1] += chunk
                prior_chunk_tokens += chunk_tokens
                # if we're still under the desired chunk size, keep trying to merge
                prior_chunk_prefers_merge = (prior_chunk_tokens < self.min_tokens)
            else:
                # otherwise, see if we'd like to merge with somebody
                if chunk_tokens < self.min_tokens:
                    # merge with previous chunk if it won't put it over the limit
                    if merged_chunks and prior_chunk_tokens + chunk_tokens <= self.max_tokens:
                        merged_chunks[-1] += chunk
                        prior_chunk_tokens += chunk_tokens
                        prior_chunk_prefers_merge = False
                    else:
                        # add chunk, but flag that we'd like to merge with the next chunk if possible
                        merged_chunks.append(chunk)
                        prior_chunk_tokens = chunk_tokens
                        prior_chunk_prefers_merge = True
                else:
                    # add chunk, with no explicit desire to merge with the next one
                    merged_chunks.append(chunk)
                    prior_chunk_tokens = chunk_tokens
                    prior_chunk_prefers_merge = False
        chunks = merged_chunks

        return chunks

    @staticmethod
    def _split_by_pattern(pattern: str) -> Callable[[str], List[str]]:
        """
        Create a function that splits text by a given regex pattern.

        This function returns a splitting function that can be used to split text at specified Markdown patterns
        like headers.

        :param pattern: Regular expression pattern to split on
        :type pattern: str
        :return: Function that splits text using the provided pattern
        :rtype: Callable[[str], List[str]]
        """

        def splitter(text: str) -> List[str]:
            if not text.strip():
                return []

            return re.split(f'(?m)(?={pattern})', text)

        return splitter

    @staticmethod
    def _split_by_paragraphs(text: str) -> List[str]:
        """
        Split text by paragraphs (double newlines).

        This function splits text at paragraph boundaries, identified by double newlines in the Markdown text.

        :param text: Text to split
        :type text: str
        :return: List of paragraphs
        :rtype: List[str]
        """

        if not text.strip():
            return []

        return re.split(r'\n\s*\n', text)

    @staticmethod
    def _split_by_lines(text: str) -> List[str]:
        """
        Split text by single newlines.

        This function splits text into individual lines, used when paragraph-level splitting isn't sufficient.

        :param text: Text to split
        :type text: str
        :return: List of lines
        :rtype: List[str]
        """

        if not text.strip():
            return []

        return text.splitlines(keepends=True)

    def _split_by_tokens(self, text: str) -> List[str]:
        """
        Split text at token boundaries as a final fallback.

        This function performs the finest-grained splitting, breaking text into chunks of approximately max_tokens,
        attempting to split at word boundaries when possible.

        :param text: Text to split
        :type text: str
        :return: List of chunks, each within token limit
        :rtype: List[str]
        """

        if self.count_tokens(text) <= self.max_tokens:
            return [text]

        chunks = []
        current_chunk = ""
        current_tokens = 0

        words = text.split()

        for word in words:
            word_tokens = self.count_tokens(word)

            if word_tokens > self.max_tokens:
                if current_chunk:
                    chunks.append(current_chunk)
                    current_chunk = ""
                    current_tokens = 0

                remaining_word = word
                while remaining_word:
                    for i in range(len(remaining_word), 0, -1):
                        substr = remaining_word[:i]
                        if self.count_tokens(substr) <= self.max_tokens:
                            chunks.append(substr)
                            remaining_word = remaining_word[i:]
                            break
                continue

            space_tokens = self.count_tokens(" ") if current_chunk else 0
            if current_tokens + word_tokens + space_tokens <= self.max_tokens:
                current_chunk += (" " if current_chunk else "") + word
                current_tokens = self.count_tokens(current_chunk)
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = word
                current_tokens = word_tokens

        if current_chunk:
            chunks.append(current_chunk)

        return chunks
