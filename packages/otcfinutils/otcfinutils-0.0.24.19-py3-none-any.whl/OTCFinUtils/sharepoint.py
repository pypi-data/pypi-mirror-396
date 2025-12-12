from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from OTCFinUtils.security import create_flow_headers, get_graph_token
from io import BytesIO
from http import HTTPStatus as status
import pandas as pd
import requests
import base64


# TODO - review and refactor code


def load_document(dataverse_url: str, token: str, document_path: str, drive: str, use_flow_headers: bool = False) -> BytesIO:
    """
    Load a document from SharePoint as a `BytesIO` object.
    """
    site_id, drive_id = get_connection_details(dataverse_url, token, drive, use_flow_headers)
    access_token = get_graph_token()

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    # TODO - review if provided logic corresponds to SharePoint URL format
    if document_path.startswith(drive):
        document_path = document_path[len(drive):]

    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}"
    url += f"/root:/{document_path}?select=id,@microsoft.graph.downloadUrl"

    response = requests.get(url, headers=headers)

    if response.status_code != status.OK.value:
        raise Exception(f"Failed to fetch JSON data. Status: {response.status_code}")

    json_data = response.json()
    download_url = json_data.get("@microsoft.graph.downloadUrl")

    if not download_url:
        raise Exception("Download URL not found in JSON data")

    download_response = requests.get(download_url)

    if download_response.status_code != status.OK.value:
        raise Exception(f"Failed to download file. Status: {download_response.status_code}")

    content = download_response.content

    return BytesIO(content)


# TODO - Remove dependence on Dataverse system variable, use key vault instead
# TODO - Use DVHandler object, instead of passing the url and token

def get_connection_details(dataverse_url: str, token: str, drive: str = "account", use_flow_headers: bool = False) -> tuple[str, str]:
    """
    Retrieves the SharePoint site id and drive id.
    Needed for reading and writing to SharePoint.
    """
    dataverse_url = f"{dataverse_url}/api/data/v9.2/new_systemvariableses?"
    dataverse_url += f"$select=new_value&"
    dataverse_url += f"$filter=%20new_name%20eq%20%27HTTP_SHAREPOINT_CONNECTION%27"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    response = requests.get(dataverse_url, headers=headers)
    
    if response.status_code != status.OK.value:
        raise Exception(f"Failed to get sharepoint connection details. Response: {response.text}")
    
    data = response.json()

    if "value" not in data or len(data["value"]) == 0:
        raise Exception("No data found in the response.")

    connection_details_url = data["value"][0]["new_value"]

    if use_flow_headers:
        flow_headers = create_flow_headers()
    else:
        flow_headers = None

    response = requests.post(connection_details_url, json={"drive": drive}, headers=flow_headers)

    if response.status_code != status.OK.value:
        raise Exception(f"Failed to get sharepoint connection details. Response: {response.text}")

    site_id = response.json().get("siteid")
    drive_id = response.json().get("driveid")

    return site_id, drive_id


def create_excel_document(
    df: pd.DataFrame,
    dataverse: str,
    token: str,
    drive: str,
    document_path: str,
    file_name: str,
    sheet: str,
) -> requests.Response:
    """
    Creates an Excel document on SharePoint.
    """
    if document_path.startswith(drive):
        document_path = document_path[len(drive):]
    
    access_token = get_graph_token()
    site_id, drive_id = get_connection_details(dataverse, token, drive,use_flow_headers=True)

    url = f"https://graph.microsoft.com/v1.0/sites"
    url += f"/{site_id}/drives/{drive_id}/root:"
    url += f"/{document_path}/{file_name}:/content"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/octet-stream",
    }

    with BytesIO() as output:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
        output.seek(0)
        response = requests.put(url, output, headers=headers)

    if response.status_code not in (status.CREATED.value, status.OK.value):
        raise Exception(f"Failed to upload file: {response.status_code}")

    return response


def update_excel(excel_content: bytes, new_json_rows: list) -> str:
    """
    Work in progress...
    
    The goal was to have a function that can 
    append rows to an Excel file.
    """
    df = pd.read_excel(pd.ExcelFile(BytesIO(excel_content)))

    # If DataFrame is empty, create columns using keys from the first row of new_json_rows
    if df.empty:
        df = pd.DataFrame(columns=new_json_rows[0].keys())

    # Convert each JSON row into a DataFrame and concatenate them with the existing DataFrame
    for row_data in new_json_rows:
        row_df = pd.DataFrame([row_data])
        df = pd.concat([df, row_df], ignore_index=True)

    # Convert DataFrame to Excel format and return as bytes
    output = BytesIO()
    df.to_excel(output, index=False)

    # Load the workbook
    output.seek(0)
    workbook = load_workbook(output)
    sheet = workbook.active

    if sheet is None:
        raise RuntimeError("Value of 'sheet' cannot be None.")

    # Get the range of new rows in the sheet
    start_row = len(df) - len(new_json_rows) + 2
    end_row = len(df) + 2

    # Apply yellow fill only to the range of new rows
    for row in range(start_row, end_row):
        for column in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # Adjust column widths to fit the text
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value is not None]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO
    output.seek(0)
    workbook.save(output)

    updated_excel_content = output.getvalue()

    # Encode Excel content to base64
    updated_excel_base64 = base64.b64encode(updated_excel_content).decode("utf-8")

    return updated_excel_base64
