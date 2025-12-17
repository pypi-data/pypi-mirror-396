"""测试 common.py - 通用工具函数

测试覆盖:
- random_string() - 生成随机字符串
- random_email() - 生成随机邮箱
- random_phone() - 生成随机手机号
- load_json() - 加载JSON文件
- load_excel() - 加载Excel文件
"""

import json
import string
from pathlib import Path

import pytest
from openpyxl import Workbook

from df_test_framework.utils.common import (
    load_excel,
    load_json,
    random_email,
    random_phone,
    random_string,
)


class TestRandomGenerators:
    """测试随机数据生成函数"""

    def test_random_string_default_length(self):
        """测试生成默认长度的随机字符串"""
        result = random_string()
        assert len(result) == 10
        assert all(c in string.ascii_letters + string.digits for c in result)

    def test_random_string_custom_length(self):
        """测试生成自定义长度的随机字符串"""
        result = random_string(length=20)
        assert len(result) == 20

    def test_random_string_custom_chars(self):
        """测试使用自定义字符集生成随机字符串"""
        chars = "ABC123"
        result = random_string(length=10, chars=chars)
        assert len(result) == 10
        assert all(c in chars for c in result)

    def test_random_string_only_digits(self):
        """测试只使用数字生成随机字符串"""
        result = random_string(length=8, chars=string.digits)
        assert len(result) == 8
        assert result.isdigit()

    def test_random_string_empty_length(self):
        """测试生成空字符串"""
        result = random_string(length=0)
        assert result == ""

    def test_random_email_default_domain(self):
        """测试生成默认域名的随机邮箱"""
        result = random_email()
        assert "@test.com" in result
        assert len(result.split("@")[0]) == 8

    def test_random_email_custom_domain(self):
        """测试生成自定义域名的随机邮箱"""
        result = random_email(domain="example.org")
        assert "@example.org" in result

    def test_random_email_format(self):
        """测试随机邮箱格式正确"""
        result = random_email()
        parts = result.split("@")
        assert len(parts) == 2
        assert len(parts[0]) > 0
        assert len(parts[1]) > 0

    def test_random_phone_default_prefix(self):
        """测试生成默认前缀的随机手机号"""
        result = random_phone()
        assert result.startswith("13")
        assert len(result) == 11
        assert result.isdigit()

    def test_random_phone_custom_prefix(self):
        """测试生成自定义前缀的随机手机号"""
        result = random_phone(prefix="18")
        assert result.startswith("18")
        assert len(result) == 11
        assert result.isdigit()

    def test_random_phone_format(self):
        """测试随机手机号格式正确"""
        result = random_phone()
        assert len(result) == 11
        assert all(c in string.digits for c in result)


class TestLoadJson:
    """测试JSON加载函数"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """临时目录"""
        yield tmp_path

    def test_load_json_simple_object(self, temp_dir):
        """测试加载简单JSON对象"""
        json_file = temp_dir / "test.json"
        data = {"name": "Alice", "age": 25}
        json_file.write_text(json.dumps(data), encoding="utf-8")

        result = load_json(str(json_file))
        assert result == data

    def test_load_json_array(self, temp_dir):
        """测试加载JSON数组"""
        json_file = temp_dir / "array.json"
        data = [{"id": 1}, {"id": 2}, {"id": 3}]
        json_file.write_text(json.dumps(data), encoding="utf-8")

        result = load_json(str(json_file))
        assert result == data

    def test_load_json_nested_object(self, temp_dir):
        """测试加载嵌套JSON对象"""
        json_file = temp_dir / "nested.json"
        data = {"user": {"name": "Bob", "address": {"city": "Beijing", "country": "China"}}}
        json_file.write_text(json.dumps(data), encoding="utf-8")

        result = load_json(str(json_file))
        assert result == data

    def test_load_json_with_chinese(self, temp_dir):
        """测试加载包含中文的JSON"""
        json_file = temp_dir / "chinese.json"
        data = {"姓名": "张三", "年龄": 30}
        json_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        result = load_json(str(json_file))
        assert result == data

    def test_load_json_file_not_found(self):
        """测试加载不存在的JSON文件"""
        with pytest.raises(FileNotFoundError, match="文件不存在"):
            load_json("nonexistent.json")

    def test_load_json_invalid_json(self, temp_dir):
        """测试加载无效的JSON文件"""
        json_file = temp_dir / "invalid.json"
        json_file.write_text("{ invalid json }", encoding="utf-8")

        with pytest.raises(json.JSONDecodeError):
            load_json(str(json_file))


class TestLoadExcel:
    """测试Excel加载函数"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """临时目录"""
        yield tmp_path

    def create_excel_file(self, file_path: Path, data: list, headers: list = None):
        """创建Excel测试文件"""
        wb = Workbook()
        ws = wb.active

        if headers:
            ws.append(headers)

        for row in data:
            ws.append(row)

        wb.save(file_path)
        wb.close()

    def test_load_excel_with_header(self, temp_dir):
        """测试加载带表头的Excel文件"""
        excel_file = temp_dir / "test.xlsx"
        headers = ["name", "age", "email"]
        data = [["Alice", 25, "alice@test.com"], ["Bob", 30, "bob@test.com"]]
        self.create_excel_file(excel_file, data, headers)

        result = load_excel(str(excel_file), skip_header=True)

        assert len(result) == 2
        assert result[0] == {"name": "Alice", "age": 25, "email": "alice@test.com"}
        assert result[1] == {"name": "Bob", "age": 30, "email": "bob@test.com"}

    def test_load_excel_without_header(self, temp_dir):
        """测试加载不带表头的Excel文件"""
        excel_file = temp_dir / "test_no_header.xlsx"
        data = [["Alice", 25, "alice@test.com"], ["Bob", 30, "bob@test.com"]]
        self.create_excel_file(excel_file, data)

        result = load_excel(str(excel_file), skip_header=False)

        assert len(result) == 2
        assert result[0] == {"col_0": "Alice", "col_1": 25, "col_2": "alice@test.com"}

    def test_load_excel_with_specific_sheet(self, temp_dir):
        """测试加载指定工作表"""
        excel_file = temp_dir / "multi_sheet.xlsx"
        wb = Workbook()

        # 创建第一个工作表
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["name", "age"])
        ws1.append(["Alice", 25])

        # 创建第二个工作表
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["product", "price"])
        ws2.append(["Book", 50])

        wb.save(excel_file)
        wb.close()

        result = load_excel(str(excel_file), sheet_name="Sheet2", skip_header=True)

        assert len(result) == 1
        assert result[0] == {"product": "Book", "price": 50}

    def test_load_excel_empty_file(self, temp_dir):
        """测试加载空Excel文件"""
        excel_file = temp_dir / "empty.xlsx"
        wb = Workbook()
        wb.save(excel_file)
        wb.close()

        result = load_excel(str(excel_file))
        assert result == []

    def test_load_excel_with_empty_rows(self, temp_dir):
        """测试加载包含空行的Excel文件"""
        excel_file = temp_dir / "with_empty_rows.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 25])
        ws.append([None, None])  # 空行
        ws.append(["Bob", 30])
        wb.save(excel_file)
        wb.close()

        result = load_excel(str(excel_file), skip_header=True)

        assert len(result) == 2
        assert result[0]["name"] == "Alice"
        assert result[1]["name"] == "Bob"

    def test_load_excel_file_not_found(self):
        """测试加载不存在的Excel文件"""
        with pytest.raises(FileNotFoundError, match="文件不存在"):
            load_excel("nonexistent.xlsx")

    def test_load_excel_with_chinese(self, temp_dir):
        """测试加载包含中文的Excel文件"""
        excel_file = temp_dir / "chinese.xlsx"
        headers = ["姓名", "年龄", "城市"]
        data = [["张三", 25, "北京"], ["李四", 30, "上海"]]
        self.create_excel_file(excel_file, data, headers)

        result = load_excel(str(excel_file), skip_header=True)

        assert len(result) == 2
        assert result[0] == {"姓名": "张三", "年龄": 25, "城市": "北京"}


__all__ = [
    "TestRandomGenerators",
    "TestLoadJson",
    "TestLoadExcel",
]
