"""通用工具函数"""

import json
import random
import string
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def random_string(length: int = 10, chars: str = string.ascii_letters + string.digits) -> str:
    """
    生成随机字符串

    Args:
        length: 字符串长度
        chars: 字符集

    Returns:
        随机字符串
    """
    return "".join(random.choice(chars) for _ in range(length))


def random_email(domain: str = "test.com") -> str:
    """
    生成随机邮箱

    Args:
        domain: 邮箱域名

    Returns:
        随机邮箱地址
    """
    username = random_string(8)
    return f"{username}@{domain}"


def random_phone(prefix: str = "13") -> str:
    """
    生成随机手机号

    Args:
        prefix: 前缀

    Returns:
        随机手机号
    """
    suffix = "".join(random.choice(string.digits) for _ in range(9))
    return f"{prefix}{suffix}"


def load_json(file_path: str) -> Any:
    """
    加载JSON文件

    Args:
        file_path: 文件路径

    Returns:
        JSON数据
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_excel(
    file_path: str,
    sheet_name: str = None,
    skip_header: bool = True,
) -> list[dict[str, Any]]:
    """
    加载Excel文件为字典列表

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称,如果为None则使用第一个工作表
        skip_header: 是否跳过第一行作为表头

    Returns:
        字典列表,每个字典代表一行数据

    示例:
        Excel内容:
        | name  | age | email          |
        | Alice | 25  | alice@test.com |
        | Bob   | 30  | bob@test.com   |

        返回:
        [
            {"name": "Alice", "age": 25, "email": "alice@test.com"},
            {"name": "Bob", "age": 30, "email": "bob@test.com"}
        ]
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 加载工作簿
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # 获取工作表
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 读取所有行
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # 提取表头
    if skip_header:
        headers = rows[0]
        data_rows = rows[1:]
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]
        data_rows = rows

    # 转换为字典列表
    result = []
    for row in data_rows:
        # 跳过空行
        if all(cell is None for cell in row):
            continue

        row_dict = {}
        for header, value in zip(headers, row):
            row_dict[header] = value
        result.append(row_dict)

    wb.close()
    return result


__all__ = [
    "random_string",
    "random_email",
    "random_phone",
    "load_json",
    "load_excel",
]
