"""Professional Excel report generator."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .statistics import calculate_statistics, filter_suspicious
from .styles import (
    BORDER,
    COLUMN_WIDTHS,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    SCORE_COLORS,
    get_score_style,
)

logger = logging.getLogger(__name__)


class ExcelReporter:
    """Professional Excel report generation."""

    def __init__(self):
        """Initializes the report generator."""
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore
        if self.ws is None:
            self.ws = self.wb.create_sheet()
        self.ws.title = "Suspicious Files"

    def generate_report(self, results: List[Dict], output_file: Path):
        """Generates the Excel report with results.

        Args:
            results: List of analysis results.
            output_file: Output Excel file path.
        """
        # Filter only suspicious files (score < 90)
        suspicious = filter_suspicious(results)

        # Generate data sheet
        self._write_data_sheet(suspicious)

        # Add summary sheet
        self._add_summary_sheet(results, suspicious)

        # Save
        self.wb.save(output_file)
        logger.info(f"✅ Excel report generated: {output_file}")

    def _write_data_sheet(self, suspicious: List[Dict]):
        """Writes the suspicious files data sheet.

        Args:
            suspicious: List of suspicious files.
        """
        # Headers
        headers = [
            "Full Path",
            "Filename",
            "FLAC Score (%)",
            "Reason for Doubt",
            "Cutoff Frequency (Hz)",
            "Sample Rate",
            "Bit Depth",
            "Encoder",
            "Duration Issue",
            "Metadata Duration",
            "Real Duration",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Write data
        for row_idx, result in enumerate(suspicious, start=2):
            self._write_result_row(row_idx, result)

        # Auto-adjust column widths
        for col_letter, width in COLUMN_WIDTHS.items():
            self.ws.column_dimensions[col_letter].width = width

        # Freeze first row
        self.ws.freeze_panes = "A2"

    def _write_result_row(self, row_idx: int, result: Dict):
        """Writes a result row.

        Args:
            row_idx: Row index.
            result: Analysis result.
        """
        # Full path
        cell = self.ws.cell(row=row_idx, column=1)
        cell.value = result["filepath"]
        cell.border = BORDER

        # Filename
        cell = self.ws.cell(row=row_idx, column=2)
        cell.value = result["filename"]
        cell.border = BORDER

        # Score with color code
        cell = self.ws.cell(row=row_idx, column=3)
        cell.value = result["score"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        fill, font = get_score_style(result["score"])
        cell.fill = fill
        if font:
            cell.font = font

        # Reason
        cell = self.ws.cell(row=row_idx, column=4)
        cell.value = result["reason"]
        cell.alignment = Alignment(wrap_text=True)
        cell.border = BORDER

        # Cutoff frequency
        cell = self.ws.cell(row=row_idx, column=5)
        cell.value = result["cutoff_freq"]
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Sample Rate
        cell = self.ws.cell(row=row_idx, column=6)
        cell.value = result["sample_rate"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Bit Depth
        cell = self.ws.cell(row=row_idx, column=7)
        cell.value = result["bit_depth"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Encodeur
        cell = self.ws.cell(row=row_idx, column=8)
        cell.value = result["encoder"]
        cell.border = BORDER

        # Duration Issue
        cell = self.ws.cell(row=row_idx, column=9)
        duration_mismatch = result.get("duration_mismatch")
        if duration_mismatch:
            cell.value = duration_mismatch
            cell.fill = SCORE_COLORS["suspect"]
            cell.font = Font(bold=True)
        else:
            cell.value = "✓ OK"
            cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Metadata Duration
        cell = self.ws.cell(row=row_idx, column=10)
        cell.value = result.get("duration_metadata", "N/A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Real Duration
        cell = self.ws.cell(row=row_idx, column=11)
        cell.value = result.get("duration_real", "N/A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    def _add_summary_sheet(self, all_results: List[Dict], suspicious: List[Dict]):
        """Adds a statistical summary sheet.

        Args:
            all_results: Complete list of results.
            suspicious: List of suspicious files.
        """
        ws_summary = self.wb.create_sheet("Summary", 0)

        # Calculate statistics
        stats = calculate_statistics(all_results)

        # Title
        ws_summary["A1"] = "FLAC ANALYSIS REPORT"
        ws_summary["A1"].font = Font(size=16, bold=True, color="366092")

        # Date
        ws_summary["A2"] = f'Generated on: {datetime.now().strftime("%d/%m/%Y at %H:%M:%S")}'
        ws_summary["A2"].font = Font(size=10, italic=True)

        # Global statistics
        ws_summary["A4"] = "GLOBAL STATISTICS"
        ws_summary["A4"].font = Font(size=12, bold=True)

        stat_rows = [
            ("Files analyzed:", stats["total"], ""),
            ("Authentic (90-100%):", stats["authentic"], stats["authentic_pct"]),
            (
                "Probably authentic (70-89%):",
                stats["probably_authentic"],
                stats["probably_authentic_pct"],
            ),
            ("Suspicious (50-69%):", stats["suspect"], stats["suspect_pct"]),
            ("Very suspicious (<50%):", stats["fake"], stats["fake_pct"]),
        ]

        row = 5
        for label, value, pct in stat_rows:
            ws_summary[f"A{row}"] = label
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = value
            if pct:
                ws_summary[f"C{row}"] = pct
                ws_summary[f"C{row}"].font = Font(italic=True)
            row += 1

        # Statistics on duration issues
        row += 1
        ws_summary[f"A{row}"] = "DURATION ISSUES (Metadata Consistency)"
        ws_summary[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        duration_stat_rows = [
            (
                "Files with duration mismatch:",
                stats["duration_issues"],
                stats["duration_issues_pct"],
            ),
            (
                "Critical mismatch (>1 second):",
                stats["duration_issues_critical"],
                stats["duration_issues_critical_pct"],
            ),
        ]

        for label, value, pct in duration_stat_rows:
            ws_summary[f"A{row}"] = label
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = value
            if pct:
                ws_summary[f"C{row}"] = pct
                ws_summary[f"C{row}"].font = Font(italic=True)
            row += 1

        # Explanatory note
        row += 2
        ws_summary[f"A{row}"] = "ℹ️ Note on duration issues:"
        ws_summary[f"A{row}"].font = Font(bold=True, color="0066CC")
        row += 1
        ws_summary[f"A{row}"] = (
            "A mismatch between metadata duration and real duration may indicate:"
        )
        row += 1
        ws_summary[f"A{row}"] = "  • Corrupted file during encoding"
        row += 1
        ws_summary[f"A{row}"] = "  • Failed transcoding (sample loss)"
        row += 1
        ws_summary[f"A{row}"] = "  • Incorrect metadata after manual editing"
        row += 1
        ws_summary[f"A{row}"] = "  • Issue during album split/merge"
        row += 1
        ws_summary[f"A{row}"] = "Normal tolerance: <588 samples (~13ms for 44.1kHz)"

        # Column adjustment
        ws_summary.column_dimensions["A"].width = 45
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 15
