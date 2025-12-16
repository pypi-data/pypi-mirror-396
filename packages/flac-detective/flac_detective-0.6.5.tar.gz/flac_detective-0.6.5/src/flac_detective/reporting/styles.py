"""Styles and formatting for Excel reports."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Styles for headers
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Borders
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Colors for scores
SCORE_COLORS = {
    "good": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
    "suspect": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Orange
    "bad": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
}

# Column widths
COLUMN_WIDTHS = {
    "A": 60,  # Full path
    "B": 40,  # Filename
    "C": 15,  # Score
    "D": 50,  # Reason
    "E": 18,  # Frequency
    "F": 12,  # Sample Rate
    "G": 10,  # Bit Depth
    "H": 25,  # Encoder
    "I": 35,  # Duration Issue
    "J": 18,  # Metadata Duration
    "K": 18,  # Real Duration
}


def get_score_style(score: int) -> tuple[PatternFill, Font | None]:
    """Returns the appropriate style based on the score.

    Args:
        score: Quality score (0-100).

    Returns:
        Tuple (fill, font) for the cell style.
    """
    if score >= 70:
        return SCORE_COLORS["good"], None
    elif score >= 50:
        return SCORE_COLORS["suspect"], None
    else:
        return SCORE_COLORS["bad"], Font(color="FFFFFF", bold=True)
