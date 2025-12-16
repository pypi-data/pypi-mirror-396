"""
Formatters for Excel
"""

from typing import List, Optional, Tuple
from copy import copy
import logging

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from lighting_paperwork.helpers import ShowData, FontStyle

logger = logging.getLogger(__name__)

X_PADDING = 0.7
Y_PADDING = 0.7
Y_PADDING_HEADER = 0.4
HEAD_FOOT_PAD = 0.2

PAGE_HEIGHT_INCHES = 11


def page_setup(ws: Worksheet, rows_to_repeat: int = 0) -> None:
    """
    Set the page size, margins, and default view.
    """
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_options.horizontalCentered = True
    ws.sheet_view.view = "pageLayout"

    ws.page_margins = PageMargins(
        bottom=Y_PADDING,
        top=Y_PADDING + Y_PADDING_HEADER,
        left=X_PADDING,
        right=X_PADDING,
        header=HEAD_FOOT_PAD,
        footer=HEAD_FOOT_PAD,
    )
    ws.sheet_view.showGridLines = False

    # Force the title + column names to repeat
    if rows_to_repeat > 0:
        ws.print_title_rows = f"1:{rows_to_repeat}"


def add_title(ws: Worksheet, name: str, show_info: Optional[ShowData] = None) -> None:
    """
    Add header and footer to worksheet
    """
    if ws.oddHeader is None:
        raise RuntimeError("oddHeader is not writable!")
    if show_info is not None:
        # Header
        ws.oddHeader.left.text = f"&[Date]\n {show_info.revision}"
        ws.oddHeader.left.size = 12
        ws.oddHeader.right.text = f"{show_info.show_name}\nLD: {show_info.ld_name}"
        ws.oddHeader.right.size = 12
    else:
        ws.oddHeader.left.text = "&[Date]"
        ws.oddHeader.left.size = 12

    # Title
    ws.oddHeader.center.text = name
    ws.oddHeader.center.size = 22
    ws.oddHeader.center.font = "Calibri,Bold"

    # Footer
    if ws.oddFooter is None:
        raise RuntimeError("oddFooter is not writable!")
    ws.oddFooter.left.text = name
    ws.oddFooter.left.size = 12
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.right.size = 12


def set_col_widths(ws: Worksheet, width: List[int], page_width: int) -> None:
    """
    Set the widths of a page in terms of % of a full page

    Widths are provided in terms of percentages, but excel expects px
    Assume page width is 610px (experimentally derived)
    I think it's 96 ppi so 96 * usable page width?
    https://www.reddit.com/r/excel/comments/l9k99z/why_does_excel_use_different_units_of_measurement/
    """
    width_px = [w * 0.01 * 610 * (page_width / 100) for w in width]

    for i in range(1, ws.max_column + 1):
        # why tf do we divide by 7
        ws.column_dimensions[get_column_letter(i)].width = int(width_px[i - 1] / 7)


def wrap_all_cells(ws: Worksheet) -> None:
    """
    Force all cells to wrap text instead of overflow.
    """
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment


def add_section_header(
    ws: Worksheet, text: str, format: FontStyle, end_col: Optional[int] = None
) -> None:
    """Adds a section header to the bottom of the worksheet"""
    if ws.max_row == 1:
        section_row = 1
    else:
        section_row = ws.max_row + 1
    if end_col is None:
        end_col = ws.max_column

    # The order and stuff really matter for auto-height to work properly
    # No idea why, just be careful around here.
    header_cell = ws.cell(row=section_row, column=1)
    header_cell.value = text
    ws.merge_cells(
        start_row=section_row, end_row=section_row, start_column=1, end_column=end_col
    )

    header_cell.font = format.excel()
    header_cell.alignment = Alignment(horizontal="left", vertical="center")


def instr_schedule_pagebreaks(ws: Worksheet) -> None:
    """
    Goal: each position should fit on a page (or at least take up a full page otherwise)
    Assumes that the excel sheet is using the default formatting.
    This is ridiculously janky.
    Please sneeze away from this function lest something gets bumped.
    """
    pos_start_index = 0
    last_height = 0
    cur_height = 0.0

    # Note: all math here done in inches. it's hacky but also excel sucks so
    PAGE_FUDGE = 0.7  # Adjust this if you have weird overflow issues.
    PAGE_HEIGHT = (PAGE_HEIGHT_INCHES - (Y_PADDING * 2 + Y_PADDING_HEADER)) - PAGE_FUDGE

    TYPE_LINEBREAK_LEN = 30
    COLOR_LINEBREAK_LEN = 25

    for row in range(1, ws.max_row):
        # calculate how long this position is
        if ws.cell(row, 5).value is None and ws.cell(row, 1).value is None:
            # If no channel and no U#, end of position
            if last_height + cur_height > PAGE_HEIGHT:
                # we don't want to add this position to the same page, add pagebreak
                ws.row_breaks.append(Break(id=pos_start_index - 1))
                last_height = cur_height
            else:
                last_height = last_height + cur_height + 0.22

            cur_height = 0
            logger.debug(f"Row {row} is a end-of-section ({cur_height})")
        elif ws.cell(row, 2).value is None and (not ws.cell(row, 1).value.isdigit()):
            # This is a position title -> height of 0.33"
            cur_height += 0.33
            pos_start_index = row
            logger.debug(f"Row {row} is a position title ({cur_height})")
        elif ws.cell(row, 1).value == "U#":
            # This is a col label row
            cur_height += 0.22
            logger.debug(f"Row {row} is col label ({cur_height})")
        elif ws.cell(row, 1).value is None or ws.cell(row, 1).value.isdigit():
            # Channel row
            if (
                ws.cell(row, 3).value is not None 
                    and (len(ws.cell(row, 3).value) > TYPE_LINEBREAK_LEN
                    or len(ws.cell(row, 4).value) > COLOR_LINEBREAK_LEN)
            ):
                # Double height
                cur_height += 0.44
                logger.debug(f"Row {row} is double height channel ({cur_height})")
            else:
                cur_height += 0.22
                logger.debug(f"Row {row} is standard height channel ({cur_height})")
        else:
            # dunno, assume it's a standard row
            cur_height += 0.22
            logger.debug(f"Row {row} unknown type ({cur_height})")
