"""Generator for an instrument hookup"""

import logging
import re
from copy import copy
from typing import List, Self, Tuple
from os import path

import openpyxl
import numpy as np
import pandas as pd
from natsort import natsort_keygen, natsorted
from pandas.io.formats.style import Styler

from lighting_paperwork.helpers import FontStyle, excel_quirks
from lighting_paperwork.paperwork import PaperworkGenerator
from lighting_paperwork.style import default_position_style
import lighting_paperwork.excel_formatter as excel_formatter

logger = logging.getLogger(__name__)


class InstrumentSchedule(PaperworkGenerator):
    """
    Generates an instrument schedules with U#, instrument type, color,
        channel, address, etc per position.
    TODO: Accessories don't show up.
    """

    def __init__(
        self, *args, position_style: FontStyle = default_position_style, **kwargs
    ):
        super().__init__(*args, **kwargs)
        self.position_style = position_style

    col_widths = [5, 17, 36, 28, 7, 7]
    display_name = "Instrument Schedule"

    def generate_df(self) -> Self:
        filter_fields = [
            "Position",
            "Unit Number",
            "Purpose",
            "Instrument Type",
            "Wattage",
            "Color",
            "Gobo 1",
            "Channel",
            "Absolute Address",
        ]
        self.verify_filter_fields(filter_fields)

        self.df = pd.DataFrame(self.vw_export[filter_fields], columns=filter_fields)
        # Need to have a position to show up in the instrument schedule
        self.df["Position"] = self.df["Position"].replace("", np.nan)
        self.df = self.df.dropna(subset=["Position"])

        self.combine_instrtype().format_address_slash().combine_gelgobo().abbreviate_col_names()
        self.df = self.df.sort_values(
            by=["Position", "U#", "Purpose"], key=natsort_keygen()
        )

        self.df = self.df[
            [
                "Position",
                "U#",
                "Purpose",
                "Instr Type & Load",
                "Color & Gobo",
                "Chan",
                "Addr",
            ]
        ]

        return self

    def split_by_position(self) -> List[Tuple[str, pd.DataFrame]]:
        """
        Split dataframe into multiple dataframes, one per position.
        """
        # Step one: sort position names
        unique_vals = self.df["Position"].unique()

        # Sort by Cat (descending), Elec (ascending), other
        elec_list = natsorted([x for x in unique_vals if re.match(r"^Elec\s\d", x)])
        lx_list = natsorted([x for x in unique_vals if re.match(r"^LX\d", x)])
        cat_list = natsorted(
            [x for x in unique_vals if re.match(r"^Cat\s\d", x)], reverse=True
        )
        # TODO: might be nice to force a linebreak between categories
        foh_list = natsorted(
            [
                x
                for x in unique_vals
                if re.match(r"^FOH\s\d", x) or re.match(r"^FOH$", x)
            ]
        )

        # This will need to be tweaked per venue
        box_booms_ds = natsorted(
            [x for x in unique_vals if re.match(r"^DS[RL]\sBox Boom", x)]
        )
        box_booms_us = natsorted(
            [x for x in unique_vals if re.match(r"^US[RL]\sBox Boom", x)]
        )
        booms = natsorted([x for x in unique_vals if re.match(r"^S[RL]\sBoom\s\d", x)])
        ladders = natsorted([x for x in unique_vals if re.match(r"^S[RL]\sLadder", x)])

        # The order of this is what'll make the order in the schedule!
        special_positions = (
            cat_list
            + foh_list
            + elec_list
            + lx_list
            + booms
            + box_booms_ds
            + box_booms_us
            + ladders
        )

        other_list = natsorted([x for x in unique_vals if x not in (special_positions)])

        position_names = special_positions + other_list

        # Step two: create unique dataframes by position name
        sorted_dfs = []
        for i in position_names:
            pos_df = self.df.loc[self.df["Position"] == i].copy()
            pos_df = pos_df.drop(["Position"], axis=1)
            pos_df = pos_df.rename(columns={"Channel": "Chan", "Unit Number": "U#"})
            pos_df = pos_df.sort_values(by=["U#"], key=natsort_keygen())
            sorted_dfs.append((i, pos_df))

        return sorted_dfs

    @staticmethod
    def style_data(
        df: pd.DataFrame,
        body_style: FontStyle,
        col_width: List[int],
        border_weight: float,
    ) -> pd.DataFrame:
        chan_border_style = f"{border_weight}px dashed black"
        style_df = df.copy()
        # Set borders based on channel data
        prev_row = ("", "")
        for index, data in df.iterrows():
            style_df.loc[index, :] = ""
            if prev_row[0] == "" and prev_row[1] == "":
                style_df.loc[index, :] += f"border-bottom: {chan_border_style}; "
                prev_row = (index, data)
                continue

            if data["U#"] == prev_row[1]["U#"]:
                # Same U#, remove dashed line
                style_df.loc[prev_row[0], :] += "border-bottom: none; "
                style_df.loc[index, :] += f"border-bottom: {chan_border_style}; "
            else:
                style_df.loc[index, :] += f"border-bottom: {chan_border_style}; "

            prev_row = (index, data)

        # Last row gets a solid bottom border
        style_df.loc[prev_row[0], :] += (
            f"border-bottom: {border_weight}px solid black; "
        )

        # Set font based on column
        for col_name, _ in style_df.items():
            style_df[col_name] += (
                f"{body_style.to_css()}; vertical-align: middle; "
                f"width: {col_width[style_df.columns.get_loc(col_name)]}%; "
            )

            if col_name in ["Chan", "U#", "Addr"]:
                style_df[col_name] += "text-align: center; "
            else:
                style_df[col_name] += "text-align: left; "

        return style_df

    @staticmethod
    def style_fields(
        index: pd.Series,
        header_style: FontStyle,
        col_width: List[int],
        border_weight: float,
    ) -> List[str]:
        PaperworkGenerator.verify_width(col_width)
        style = [
            f"{header_style.to_css()}; border-top: {border_weight}px solid black; "
            f"border-bottom: {border_weight}px solid black; "
            for _ in index
        ]

        for idx, val in enumerate(index):
            if val in ["Chan", "U#", "Addr"]:
                style[idx] += "text-align: center; "
            else:
                style[idx] += "text-align: left; "

        for idx, _ in enumerate(index):
            style[idx] += f"width: {col_width[idx]}%; "

        return style

    def _make_common(self):
        raise NotImplementedError("Use _style_position for instrument schedule")

    def _style_position(
        self, position: tuple[str, pd.DataFrame]
    ) -> tuple[str, pd.io.formats.style.Styler]:
        styled = Styler.from_custom_template(path.join(path.dirname(__file__), "templates"), "header_footer.tpl")(position[1])
        styled = styled.apply(
            type(self).style_data,
            axis=None,
            body_style=self.style.body,
            col_width=self.col_widths,
            border_weight=self.border_weight,
        )
        styled = styled.hide()
        styled = styled.apply_index(
            type(self).style_fields,
            header_style=self.style.field,
            col_width=self.col_widths,
            border_weight=self.border_weight,
            axis=1,
        )
        return (position[0], styled)

    def make_excel(self, excel_path: str) -> None:
        self.formatting_quirks = excel_quirks

        self.generate_df()
        positions = self.split_by_position()
        sheet_names = []

        for idx, pos in enumerate(positions):
            _, styled = self._style_position(pos)
            sheet_names.append(f"inst_sch_tmp_{idx}")
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
                styled.to_excel(writer, sheet_name=sheet_names[-1])

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.create_sheet(title=self.display_name, index=-1)

        for idx, sht_name in enumerate(sheet_names):
            excel_formatter.add_section_header(
                ws, positions[idx][0], self.position_style, len(self.df.columns) - 1
            )
            sht = wb[sht_name]
            cur_max = ws.max_row

            # remove index column
            sht.delete_cols(idx=1)

            # Copy cells over with formatting
            for i in range(1, sht.max_row + 1):
                for j in range(1, sht.max_column + 1):
                    dest_cell = ws.cell(row=cur_max + i, column=j)
                    dest_cell.value = sht.cell(row=i, column=j).value
                    if sht.cell(row=i, column=j).has_style:
                        dest_cell._style = copy(sht.cell(row=i, column=j)._style)

                    # Wrap text here to avoid messing with the headers
                    alignment = copy(dest_cell.alignment)
                    alignment.wrapText = True
                    dest_cell.alignment = alignment

            # Add an empty row at bottom
            ws.cell(ws.max_row + 1, 1).value = None

            del wb[sht_name]

        excel_formatter.add_title(ws, self.display_name, self.show_data)
        excel_formatter.page_setup(ws, 0)
        excel_formatter.set_col_widths(ws, self.col_widths, self.page_width)
        excel_formatter.instr_schedule_pagebreaks(ws)
        wb.save(excel_path)

    def make_html(self) -> str:
        self.generate_df()
        positions = self.split_by_position()

        header_html, footer_html = self.generate_header_footer("instr")
        page_style = self.generate_page_style(
            "instr", "bottom-right", self.style.marginals.to_css()
        )

        output_html = page_style
        output_html += header_html
        output_html += "<div id='inst-schedule-container'>\n"
        for pos in positions:
            position_name, styled = self._style_position(pos)
            styled = styled.set_table_attributes('class="paperwork-table"')
            styled = styled.set_table_styles(
                self.default_table_style(), overwrite=False
            )
            styled = styled.set_table_styles(
                [{"selector": "", "props": "break-inside: avoid; margin-bottom: 5mm;"}],
                overwrite=False,
            )

            header_html = self.generate_header(
                styled.uuid,
                content_left=position_name,
                style_left=self.position_style.to_css(),
            )

            output_html += styled.to_html(
                generated_header=header_html,
            )

        output_html += "\n</div>"
        output_html += footer_html
        output_html = self.wrap_table(output_html)

        logger.info("Generated instrument schedule.")

        return output_html
