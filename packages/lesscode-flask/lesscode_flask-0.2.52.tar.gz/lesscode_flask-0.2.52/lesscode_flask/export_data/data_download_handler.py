import json
import logging
import os
from datetime import datetime

from flask import current_app
from openpyxl import Workbook

from lesscode_flask.utils.file.file_utils import check_or_create_dir
from lesscode_flask.utils.helpers import app_config
from lesscode_flask.utils.oss import Ks3Oss

download_func_dict = {}


def upload_result_url(data_list: list, file_name: str = None, describe_text: str = None):
    # 创建一个工作簿和一个工作表
    wb = Workbook()
    ws = wb.active
    # 将数据写入工作表
    for row in data_list:
        row_list = []
        for x in row:
            try:
                if type(x) == list:
                    row_list.append("，".join(list(set(x))))
                elif type(x) == dict:
                    row_list.append(json.dumps(x, ensure_ascii=False))
                else:
                    row_list.append(str(x))
            except:
                row_list.append(json.dumps(x, ensure_ascii=False))
        ws.append(row_list)
    # 自适应列宽
    for index, col in enumerate(ws.columns):
        max_length = max([len(str(data[index])) for data in data_list])
        column = col[0].column_letter  # 获取列字母
        adjusted_width = min(60, (max_length * 2 + 2))  # 添加一些额外的空白以防止截断
        ws.column_dimensions[column].width = adjusted_width
    if describe_text:
        ws.insert_rows(1, amount=1)
        ws.merge_cells(f"A1:{column}1")  # 合并第1行单元格
        ws["A1"] = describe_text
    # img = Image("./static/img/img.png")
    # ws.add_image(img, "A1")
    # 保存工作簿
    file_key = f'data_download/{file_name}-{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    path = check_or_create_dir(f'{current_app.config.get("STATIC_PATH", "")}/{file_key}')
    wb.save(path)
    clear_excel_task()
    if app_config.get("DATA_DOWNLOAD_UPLOAD_KS3"):
        ks = Ks3Oss()
        bucket_name = app_config.get("KS3_CONNECT_CONFIG", {}).get("bucket_name")
        url = ks.save(key=file_key, bucket_name=bucket_name,
                      policy="public-read",
                      filename=path, content_type="filename")
        return url
    else:
        return f'{current_app.config.get("OUTSIDE_SCREEN_IP")}/static/{file_key}'


def clear_excel_task():
    delay = 3600 * 24
    excel_path = os.path.join(current_app.config.get("STATIC_PATH", ""), "data_download")
    files = get_files(excel_path)
    for f in files:
        now = datetime.now().timestamp()
        if os.path.exists(f):
            file_time = os.path.getctime(f)
            if now - file_time > delay:
                os.remove(f)
                logging.info(f'delete file({f}) success')


def get_files(path):
    files = []
    fs = os.listdir(path)
    for f in fs:
        file_path = f'{path}/{f}'
        if os.path.isfile(file_path):
            if ".gitkeep" not in file_path:
                files.append(file_path)
        else:
            ft = get_files(file_path)
            files.extend(ft)
    return files


def format_to_table_download(table_head_list=None, table_body_list=None):
    if not table_head_list and table_body_list:
        if table_body_list.get("columns"):
            table_head_list = table_body_list["columns"]
        elif table_body_list.get("column"):
            table_head_list = table_body_list["column"]
        else:
            table_head_list = []
        table_body_list = table_body_list["dataSource"]
    elif not table_body_list and table_head_list:
        table_body_list = []
    elif not table_head_list and not table_body_list:
        table_body_list = []
        table_head_list = []
    else:
        table_body_list = table_body_list["dataSource"]
    data_index_list = [one['dataIndex'] for one in table_head_list]
    title_list = [one['title'] for one in table_head_list]
    all_data_list = [title_list]
    for one in table_body_list:
        all_data_list.append([one.get(dataIndex, "") for dataIndex in data_index_list])
    # table_head_map_dict = {one['dataIndex']: one['title'] for one in table_head_list}
    #
    # sta_map_dict = {}
    # for one in table_body_list:
    #     for tag in data_index_list:
    #         if tag not in sta_map_dict:
    #             sta_map_dict[tag] = [one.get(tag, "")]
    #         else:
    #             sta_map_dict[tag].append(one.get(tag, ""))
    # sta_map_dict = {table_head_map_dict[k]: v for k, v in sta_map_dict.items()}
    return all_data_list


def convert_page(offset: int = None, size: int = None, page_num: int = None, page_size: int = None):
    if not offset and not size:
        offset = (page_num - 1) * page_size
        size = page_size
    return offset, size
