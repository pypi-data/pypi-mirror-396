"""Microsoft Excel file loader."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from prompt_amplifier.core.exceptions import LoaderError
from prompt_amplifier.loaders.base import BaseLoader
from prompt_amplifier.models.document import Document


class ExcelLoader(BaseLoader):
    """
    Load Microsoft Excel files (.xlsx, .xls).

    Requires: openpyxl (for .xlsx) or xlrd (for .xls)

    Example:
        >>> loader = ExcelLoader(sheet_name="Data")
        >>> docs = loader.load("workbook.xlsx")
    """

    def __init__(
        self,
        sheet_name: str | int | None = None,
        row_as_document: bool = False,
        header_row: int = 0,
        content_columns: list[str] | None = None,
        metadata_columns: list[str] | None = None,
        **kwargs: Any,
    ):
        """
        Initialize Excel loader.

        Args:
            sheet_name: Sheet to load (None = all sheets, int = sheet index)
            row_as_document: If True, each row becomes a document
            header_row: Row index containing headers (0-indexed)
            content_columns: Columns to use for content
            metadata_columns: Columns to extract as metadata
        """
        super().__init__(**kwargs)
        self.sheet_name = sheet_name
        self.row_as_document = row_as_document
        self.header_row = header_row
        self.content_columns = content_columns
        self.metadata_columns = metadata_columns or []
        self._check_dependency()

    def _check_dependency(self) -> None:
        """Check if openpyxl is installed."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl is required for ExcelLoader. " "Install it with: pip install openpyxl"
            )

    def load(self, source: str | Path) -> list[Document]:
        """Load an Excel file."""
        import openpyxl

        path = Path(source)

        if not path.exists():
            raise LoaderError(f"File not found: {source}")

        try:
            workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:
            raise LoaderError(f"Failed to read Excel file {source}: {e}")

        documents = []

        # Determine which sheets to process
        if self.sheet_name is None:
            sheets = workbook.sheetnames
        elif isinstance(self.sheet_name, int):
            sheets = [workbook.sheetnames[self.sheet_name]]
        else:
            sheets = [self.sheet_name]

        for sheet_name in sheets:
            try:
                sheet = workbook[sheet_name]
                docs = self._process_sheet(sheet, sheet_name, source)
                documents.extend(docs)
            except KeyError:
                raise LoaderError(f"Sheet '{sheet_name}' not found in {source}")

        workbook.close()
        return documents

    def _process_sheet(
        self,
        sheet: Any,
        sheet_name: str,
        source: str | Path,
    ) -> list[Document]:
        """Process a single sheet."""
        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        # Extract headers
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[self.header_row])]
        data_rows = rows[self.header_row + 1 :]

        if self.row_as_document:
            return self._rows_as_documents(data_rows, headers, sheet_name, source)
        else:
            return self._sheet_as_document(data_rows, headers, sheet_name, source)

    def _rows_as_documents(
        self,
        rows: list[tuple],
        headers: list[str],
        sheet_name: str,
        source: str | Path,
    ) -> list[Document]:
        """Create one document per row."""
        documents = []
        content_cols = self.content_columns or headers

        for i, row in enumerate(rows):
            row_dict = dict(zip(headers, row))

            content_parts = []
            for col in content_cols:
                if col in row_dict and row_dict[col] is not None:
                    content_parts.append(f"{col}: {row_dict[col]}")

            if not content_parts:
                continue

            content = "\n".join(content_parts)

            metadata = {
                "sheet_name": sheet_name,
                "row_index": i,
            }
            for col in self.metadata_columns:
                if col in row_dict:
                    metadata[col] = row_dict[col]

            documents.append(
                self._create_document(
                    content=content,
                    source=source,
                    source_type="excel",
                    metadata=metadata,
                )
            )

        return documents

    def _sheet_as_document(
        self,
        rows: list[tuple],
        headers: list[str],
        sheet_name: str,
        source: str | Path,
    ) -> list[Document]:
        """Create single document from sheet."""
        content_cols = self.content_columns or headers

        lines = []
        lines.append(f"Sheet: {sheet_name}")
        lines.append("")
        lines.append(" | ".join(content_cols))
        lines.append("-" * 50)

        for row in rows:
            row_dict = dict(zip(headers, row))
            values = [str(row_dict.get(col, "") or "") for col in content_cols]
            lines.append(" | ".join(values))

        content = "\n".join(lines)

        return [
            self._create_document(
                content=content,
                source=source,
                source_type="excel",
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(rows),
                    "columns": headers,
                },
            )
        ]

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]
