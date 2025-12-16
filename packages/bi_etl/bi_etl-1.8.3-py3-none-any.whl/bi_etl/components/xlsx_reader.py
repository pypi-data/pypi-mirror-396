"""
Created on Apr 2, 2015
"""
import io
import os
from typing import *
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from bi_etl.components.etlcomponent import ETLComponent
from bi_etl.scheduler.task import ETLTask

__all__ = ['XLSXReader']


class XLSXReader(ETLComponent):
    """
    XLSXReader will read rows from a Microsoft Excel XLSX formatted workbook.
    
    Parameters
    ----------
    task: ETLTask
        The  instance to register in (if not None)
    
    file_name: str
        The file_name to parse as xlsx.
        
    logical_name: str
        The logical name of this source. Used for log messages.

    Attributes
    ----------
    column_names: list
        The names to use for columns
        
    header_row: int
        The sheet row to read headers from. Default = 1.
    
    start_row: int
        The first row to parse for data. Default = header_row + 1 
    
    workbook: :class:`openpyxl.workbook.workbook.Workbook`
        The workbook that was opened.
        
    log_first_row : boolean
        Should we log progress on the first row read. *Only applies if used as a source.*
        (inherited from ETLComponent)
        
    max_rows : int, optional
        The maximum number of rows to read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    primary_key: list
        The name of the primary key column(s). Only impacts trace messages.  Default=None.
        (inherited from ETLComponent)
    
    progress_frequency: int
        How often (in seconds) to output progress messages. None for no progress messages.
        (inherited from ETLComponent)
    
    progress_message: str
        The progress message to print. Default is ``"{logical_name} row # {row_number}"``.
        Note ``logical_name`` and ``row_number`` subs.
        (inherited from ETLComponent)
        
    restkey: str
        Column name to catch extra long rows (more columns than we have column names)
        when reading values (extra values).

    restval: str
        The value to put in columns that are in the column_names but
        not present in a given row when reading (missing values).
    """
    def __init__(self,
                 task: Optional[ETLTask],
                 file_name: Union[str, Path],
                 logical_name: Optional[str] = None,
                 **kwargs
                 ):
        self.file_name = file_name
        if logical_name is None:
            try: 
                logical_name = os.path.basename(self.file_name)
            except (AttributeError, TypeError):
                logical_name = str(self.file_name)
        
        # Don't pass kwargs up. They should be set here at the end
        super().__init__(
            task=task,
            logical_name=logical_name,
        )

        # column to catch long rows (more values than columns)
        self.restkey = 'extra data past last delimiter'
        # default value for short rows (value for missing keys)    
        self.restval = None    
                  
        self.__header_row = 1    
        self.__start_row = None
        self.__active_row = None               

        self._workbook = None
        self._active_worksheet: Optional[Worksheet] = None
        self._active_worksheet_name: Optional[str] = None

        # Should be the last call of every init
        self.set_kwattrs(**kwargs)

    def __repr__(self):
        return f"XLSXReader({self.logical_name})"
    
    @property
    def header_row(self) -> int:
        """
        The sheet row to read headers from. Default = 1.
        """
        return self.__header_row

    @header_row.setter
    def header_row(self, value: int):
        self.__header_row = value
        
    @property
    def start_row(self) -> int:
        """
        The sheet row to start reading data from.
        Default = header_row + 1
        """
        if self.__start_row is not None:
            return self.__start_row
        else:
            return self.header_row + 1

    @start_row.setter
    def start_row(self, value: int):
        self.__start_row = value

    def has_workbook_init(self) -> bool:
        return self._workbook is not None

    @property
    def workbook(self):
        if self._workbook is None:
            # Work around for openpyxl close not working correctly
            with open(self.file_name, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            self._workbook = load_workbook(filename=in_mem_file, read_only=True)
            # Original open code if the openpyxl libary can properly close the file
            # self._workbook = load_workbook(filename=self.file_name, read_only=True)
        return self._workbook
    
    def set_active_worksheet_by_name(self, sheet_name: str):
        self._active_worksheet = self.workbook[sheet_name]
        self._active_worksheet_name = sheet_name
        self._column_names = None
        self._full_iteration_header = None
        
    def set_active_worksheet_by_number(self, sheet_number: int):
        """
        Change to an existing worksheet based on the sheet number.
        """
        sheet_names = self.get_sheet_names()
        if len(sheet_names) >= (sheet_number + 1):
            sheet_name = sheet_names[sheet_number]
        else:
            raise ValueError(f"{self} does not have a worksheet numbered {sheet_number}")
        self.set_active_worksheet_by_name(sheet_name)
    
    @property
    def active_worksheet(self):
        if self._active_worksheet is None:
            self.set_active_worksheet_by_number(0)
        return self._active_worksheet

    @property
    def active_worksheet_name(self) -> str:
        # if self._active_worksheet_name is None:
        #     self.set_active_worksheet_by_number(0)
        # return self._active_worksheet_name
        return self.active_worksheet.title
    
    def get_sheet_names(self):
        return self.workbook.sheetnames
    
    def get_sheet_by_name(self, name):
        """Returns a worksheet by its name.

        Parameters
        ----------
        name: str
            The name of the worksheet to look for
            
        Returns
        -------
        openpyxl.worksheet.worksheet.Worksheet
            Worksheet object, or None if no worksheet has the name specified.

        """
        try:
            return self.workbook[name]
        except KeyError:
            return
    
    @property
    def line_num(self):
        """
        The current line number in the source file.
        line_num differs from rows_read in that rows_read deals with rows that would be returned to the caller
        """
        return self.__active_row
    
    def _obtain_column_names(self):
        try:
            row = self.read_header_row()                
            self._column_names = row
            if self.trace_data:
                self.log.debug(f"Column names read: {self._column_names}")
        except StopIteration:
            pass
    
    @staticmethod
    def _get_cell_value(cell):
        value = cell.value
        if hasattr(value, 'strip'):
            value = value.strip()
            if value == '':
                value = None
        elif isinstance(value, datetime):
            # Excel time values of 12:00:00 AM come in as 1899-12-30 instead
            if value == datetime(1899, 12, 30):
                value = time(12, 0, 0)
        return value
    
    @staticmethod
    def _get_cell_values(row_cells) -> list:
        # Convert empty strings to None to be consistent with DB reads
        return list(map(XLSXReader._get_cell_value, row_cells))
        
    def read_header_row(self):
        # See https://openpyxl.readthedocs.org/en/latest/tutorial.html
        # noinspection PyTypeChecker
        row = next(self.active_worksheet.iter_rows(
            min_col=1,
            min_row=self.header_row,
            max_col=None,
            max_row=self.header_row,
            )
        )
        column_names = [value or f'un_named_col_{col_num}' for col_num, value in enumerate(XLSXReader._get_cell_values(row))]
        return column_names

    def _raw_rows(self):
        # See https://openpyxl.readthedocs.org/en/latest/tutorial.html
        self.__active_row = self.start_row
        len_column_names = len(self.column_names)
        this_iteration_header = self.full_iteration_header
        for row in self.active_worksheet.iter_rows(min_row=self.start_row):
            if len(row) > 0:
                self.__active_row = row[0].row
            else:
                self.__active_row += 1
            row_values = XLSXReader._get_cell_values(row)           
            d = self.Row(data=row_values[:len_column_names], iteration_header=this_iteration_header)
            len_column_names = len(self.column_names)
            len_row = len(row_values)
            if len_column_names < len_row:
                if self.restkey is not None:
                    # Note: Adding restkey to the row will create a new iteration header
                    #      (shared by subsequent rows with extra values)
                    d[self.restkey] = row_values[len_column_names:]
            elif len_column_names > len_row:
                # This could be done in a faster way, but hopefully is rare so not worth optimizing
                for key in self.column_names[len_row]:
                    d[key] = self.restval
            yield d 
            
    def close(self, error: bool = False):
        if not self.is_closed:
            if self._workbook is not None:
                self._workbook.close()
                self._workbook = None
            super(XLSXReader, self).close(error=error)
