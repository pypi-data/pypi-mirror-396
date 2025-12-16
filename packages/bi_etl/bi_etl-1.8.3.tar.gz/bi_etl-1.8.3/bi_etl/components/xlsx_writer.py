"""
Created on Apr 2, 2015
"""
import os
import re
import sys
from pathlib import Path
from typing import *

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as WorksheetTable, TableStyleInfo

from bi_etl.components.row.row import Row
from bi_etl.components.xlsx_reader import XLSXReader
from bi_etl.scheduler.task import ETLTask
from bi_etl.statistics import Statistics

__all__ = ['XLSXWriter']


class XLSXWriter(XLSXReader):
    """
    XLSXWriter will write rows to a Microsoft Excel XLSX formatted workbook.
    
    Parameters
    ----------
    task: ETLTask
        The  instance to register in (if not None)
    
    file_name: str
        The file_name to parse as xlsx.
        
    logical_name: str
        The logical name of this source. Used for log messages.

    write_only: bool
        Should we use the faster write only mode ?
        (only supports brand-new files not adding data to an existing file)

    Attributes
    ----------
    column_names: list
        The names to use for columns
        
    header_row: int
        The sheet row to read headers from. Default = 1.
    
    start_row: int
        The first row to parse for data. Default = header_row + 1 
    
    workbook: :class:`openpyxl.workbook.workbook.Workbook`
        The workbook that was opened.
        
    log_first_row : boolean
        Should we log progress on the first row read. *Only applies if used as a source.*
        (inherited from ETLComponent)
        
    max_rows : int, optional
        The maximum number of rows to read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    primary_key: list
        The name of the primary key column(s). Only impacts trace messages.  Default=None.
        (inherited from ETLComponent)
    
    progress_frequency: int
        How often (in seconds) to output progress messages. None for no progress messages.
        (inherited from ETLComponent)
    
    progress_message: str
        The progress message to print. Default is ``"{logical_name} row # {row_number}"``.
        Note ``logical_name`` and ``row_number`` subs.
        (inherited from ETLComponent)
        
    restkey: str
        Column name to catch extra long rows (more columns than we have column names)
        when reading values (extra values).
        
    restval: str
        The value to put in columns that are in the column_names but 
        not present in a given row when reading (missing values).
    """
    def __init__(self,
                 task: Optional[ETLTask],
                 file_name: Union[str, Path],
                 logical_name: str = None,
                 write_only: bool = True,
                 **kwargs
                 ):
        super().__init__(
            task=task,
            file_name=file_name,
            logical_name=logical_name,
        )
        self.headers_written = False
        self.write_only = write_only
        self._insert_cnt = 0
        self._insert_cnt_this_sheet = 0

        # Matches invalid XML1.0 unicode characters, like control characters:
        # http://www.w3.org/TR/2006/REC-xml-20060816/#charsets
        # http://stackoverflow.com/questions/1707890/fast-way-to-filter-illegal-xml-unicode-chars-in-python

        _illegal_unichrs = [(0x00, 0x08), (0x0B, 0x0C), (0x0E, 0x1F),
                            (0x7F, 0x84), (0x86, 0x9F),
                            (0xFDD0, 0xFDDF), (0xFFFE, 0xFFFF)]
        if sys.maxunicode >= 0x10000:  # not narrow build
            _illegal_unichrs.extend([(0x1FFFE, 0x1FFFF), (0x2FFFE, 0x2FFFF),
                                     (0x3FFFE, 0x3FFFF), (0x4FFFE, 0x4FFFF),
                                     (0x5FFFE, 0x5FFFF), (0x6FFFE, 0x6FFFF),
                                     (0x7FFFE, 0x7FFFF), (0x8FFFE, 0x8FFFF),
                                     (0x9FFFE, 0x9FFFF), (0xAFFFE, 0xAFFFF),
                                     (0xBFFFE, 0xBFFFF), (0xCFFFE, 0xCFFFF),
                                     (0xDFFFE, 0xDFFFF), (0xEFFFE, 0xEFFFF),
                                     (0xFFFFE, 0xFFFFF), (0x10FFFE, 0x10FFFF)])

        _illegal_ranges = ["%s-%s" % (chr(low), chr(high))
                           for (low, high) in _illegal_unichrs]
        self._illegal_xml_chars_RE = re.compile(u'[%s]' % u''.join(_illegal_ranges))

        # Should be the last call of every init
        self.set_kwattrs(**kwargs)

    @property
    def __repr__(self):
        return f"XLSXWriter({self.logical_name})"

    @property
    def rows_inserted(self):
        return self._insert_cnt

    @property
    def rows_inserted_this_sheet(self):
        return self._insert_cnt_this_sheet

    @property
    def workbook(self):
        if self._workbook is None:
            if self.write_only:
                self._workbook = Workbook(write_only=True)
            else:
                if os.path.isfile(self.file_name):
                    self._workbook = load_workbook(filename=self.file_name, read_only=False)
                else:
                    self._workbook = Workbook()
                    del self._workbook['Sheet']
        return self._workbook

    def set_active_worksheet_by_name(self, sheet_name):
        self._active_worksheet_name = sheet_name
        if sheet_name not in self.workbook:
            self._active_worksheet = self.workbook.create_sheet(sheet_name)
            self._insert_cnt_this_sheet = 0
            if self.headers_written:
                self._column_names = None
                self.headers_written = False
        else:
            self._active_worksheet = self.workbook[sheet_name]
            self._column_names = None
            self._insert_cnt_this_sheet = 0

    def set_active_worksheet_by_number(self, sheet_number: int):
        """
        Change to an existing worksheet based on the sheet number.
        """
        try:
            super().set_active_worksheet_by_number(sheet_number)
        except ValueError as e:
            raise ValueError(f"{e}. Use set_active_worksheet_by_name to create new sheets.")

    def _obtain_column_names(self):
        raise ValueError(f'Column names must be explicitly set on {self}')

    def set_columns_and_widths(self, columns_dict: Dict[str, float]):
        """
        Set the column names and widths at the same time.
        See :py:meth:`set_widths`.
        """
        self.column_names = columns_dict.keys()
        self.set_widths(columns_dict.values())

    def write_header(self):
        """
        Write the header row.
        """
        if self.column_names is None:
            raise ValueError("insert called before column_names set (or possibly column_names needs to be set after set_active_worksheet_by_name call.")
        self.active_worksheet.append(self.column_names)
        self.headers_written = True

    def insert_row(
        self,
        source_row: Row,  # Must be a single row
        additional_insert_values: dict = None,
        stat_name: str = 'insert',
        parent_stats: Statistics = None,
    ) -> Row:
        """
        Inserts a row into the database (batching rows as batch_size)

        Parameters
        ----------
        source_row:
            The row with values to insert
        additional_insert_values:
            Values to add / override in the row before inserting.
        stat_name:
            Name of this step for the ETLTask statistics.
        parent_stats:
            Optional Statistics object to nest this steps statistics in.
            Default is to place statistics in the ETLTask level statistics.

        Returns
        -------
        new_row
        """
        stats = self.get_stats_entry(stat_name, parent_stats=parent_stats)
        stats.timer.start()

        # row_values = [source_row[col] for col in self.column_names]

        if not self.headers_written:
            self.write_header()

        assert len(source_row) > 0, f"Empty row passed into {self} insert"

        values = []
        for column_name in self.column_names:
            if additional_insert_values and column_name in additional_insert_values:
                col_value = additional_insert_values[column_name]
            else:
                try:
                    col_value = source_row[column_name]
                except KeyError:
                    col_value = None
            if isinstance(col_value, str):
                col_value = self._illegal_xml_chars_RE.sub('\\?', col_value)
            values.append(col_value)

        new_row = self.row_object(iteration_header=self.full_iteration_header)
        new_row.update_from_values(values)

        assert len(values) > 0, f"No values with column names from {self} found in {source_row}"

        self.active_worksheet.append(values)

        self._insert_cnt += 1
        self._insert_cnt_this_sheet += 1

        stats.timer.stop()

        return new_row

    def set_widths(self, column_widths: Iterable[float]):
        """
        Set the column widths in the xlsx for the currently active worksheet.
        """
        for i, column_width in enumerate(column_widths):
            self.active_worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    def make_table_from_inserted_data(self, style_name: str = 'TableStyleMedium2'):
        """
        Format the newly inserted worksheet data as a table.

        Parameters
        ----------

        style_name:
            The name of the Excel style to apply to the table.
        """
        if self._insert_cnt_this_sheet == 0:
            return
        last_col_letter = get_column_letter(len(self.column_names))

        table_ref = f"A1:{last_col_letter}{self._insert_cnt_this_sheet + 1}"
        self.log.debug(f"Adding table to worksheet {self.active_worksheet_name} for range {table_ref}")

        table = WorksheetTable(
            displayName=f"Tbl_{self.active_worksheet_name}",
            ref=table_ref
        )
        style = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # noinspection PyProtectedMember
        table._initialise_columns()
        for column, value in zip(table.tableColumns, self.column_names):
            column.name = value
        self.active_worksheet.add_table(table)

    def insert(
        self,
        source_row: Union[Row, list],  # Could also be a whole list of rows
        parent_stats: Statistics = None,
        **kwargs
    ):
        """
        Insert a row or list of rows in the table.

        Parameters
        ----------
        source_row:
            Row(s) to insert
        parent_stats:
            Optional Statistics object to nest this steps statistics in.
            Default is to place statistics in the ETLTask level statistics.
        """

        if isinstance(source_row, list):
            for row in source_row:
                self.insert_row(
                    row,
                    parent_stats=parent_stats,
                    **kwargs
                )
        else:
            self.insert_row(
                 source_row,
                 parent_stats=parent_stats,
                 **kwargs
             )
            
    def close(self, error: bool = False):
        """
        Close the xlsx file, saving first if ``error`` is false.

        Parameters
        ----------
        error:
            Did we run into an error during processing?
            Errors cause a rollback, which skips the save of the file.
        """
        if self.has_workbook_init():
            if not error:
                self.workbook.save(filename=self.file_name)
            else:
                self.log.info(f"{self}.close not saving due to error")
            self.workbook.close()
        super().close(error=error)
