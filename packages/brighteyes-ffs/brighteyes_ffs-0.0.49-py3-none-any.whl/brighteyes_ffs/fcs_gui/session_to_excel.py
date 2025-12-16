from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .restore_session import restorelib
import os


def ffs2excel(ffs_file, fname_xlsx):
    """
    Convert .ffs file to an excel sheet with a summary of the fit results

    Parameters
    ----------
    ffs_file : path
        Path to .ffs file
    fname_xlsx : path
        Path to .xslx file.

    Returns
    -------
    None.

    """
    lib = restorelib(ffs_file)
    lib2excel(lib, fname_xlsx)


def lib2excel(lib, fname_xlsx):
    """
    Convert a GUI session to an excel sheet.

    Parameters
    ----------
    lib : GUI analysis object
        Same as output from restore_session.
    fname_xlsx : path
        Path to .xlsx file, should end with .xlsx

    Returns
    -------
    None.

    """
    wb = Workbook()
    ws = wb.active
    
    current_line = 1
    
    num_im = lib.num_images
    for im in range(num_im):
        # go through all images
        im_obj = lib.lib[0]
        num_ffs = im_obj.num_files
        for f in range(num_ffs):
            current_column = 'A'
            # go through all fcs files
            ffs_obj = im_obj.ffs_list[f]
            fname = ffs_obj.fname
            
            folder_name = os.path.dirname(fname)  # Get folder path
            file_name = os.path.basename(fname) # Get file without folder
            
            ws[current_column + str(current_line)] = folder_name
            current_column = next_letter(current_column)
            ws[current_column + str(current_line)] = file_name
            
            num_analyses = len(ffs_obj.analysis_list)
            for a in range(num_analyses):
                # go through all correlation analyses
                column_corr = next_letter(current_column)
                anal_obj = ffs_obj.analysis_list[a]
                mode = anal_obj.mode # e.g. spot-variation fcs
                ws[column_corr + str(current_line)] = mode
                num_fits = len(anal_obj.fits)
                for g in range(num_fits):
                    # go through all fits
                    column_fit = next_letter(column_corr)
                    fit_obj = anal_obj.fits[g]
                    num_curves = len(fit_obj.fit_all_curves)
                    for h in range(num_curves):
                        singlefit_obj = fit_obj.fit_all_curves[h]
                        fitlabel = singlefit_obj.fitfunction_label
                        if h == 0:
                            ws[column_fit + str(current_line)] = fitlabel
                        # go through individual curve fits
                        column_fitresult = next_letter(column_fit)
                        fit_params = singlefit_obj.paramidx
                        for v in range(len(fit_params)):
                            ws[column_fitresult + str(current_line)] = singlefit_obj.startvalues[fit_params[v]]
                            if singlefit_obj.fitarray[fit_params[v]]:
                                ws[column_fitresult + str(current_line)].fill = green_fill()
                            column_fitresult = next_letter(column_fitresult)
                        current_line += 1
    wb.save(fname_xlsx)
            
def next_letter(letter):
    return chr(ord(letter) + 1)


def green_fill():
    return PatternFill(start_color="FFB5E6A2", end_color="FFB5E6A2", fill_type="solid")