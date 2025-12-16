"""
Change Log Module for Excel MCP

Provides automatic change tracking for all Excel modifications.
Creates a "Change Log" sheet with reconstruction-capable entries.

Log Entry Format:
| DateTime | Sheet | Range | Before | Action | After | Reason | validated_after |
"""

import json
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel-mcp")

CHANGE_LOG_SHEET = "Change Log"
HEADERS = ["DateTime", "Sheet", "Range", "Before", "Action", "After", "Reason", "validated_after"]


@dataclass
class ChangeLogEntry:
    """Represents a single change log entry."""
    datetime: str
    sheet: str
    range: str
    before: str
    action: str
    after: str
    reason: str
    validated_after: str


def get_current_datetime() -> str:
    """Get current datetime formatted as dd/mm/yyyy hh:mm:ss."""
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def ensure_log_sheet_exists(wb) -> Any:
    """
    Create Change Log sheet with headers if it doesn't exist.
    
    Args:
        wb: openpyxl Workbook object
        
    Returns:
        The Change Log worksheet
    """
    if CHANGE_LOG_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(CHANGE_LOG_SHEET)
        
        # Add headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            "A": 20,  # DateTime
            "B": 15,  # Sheet
            "C": 15,  # Range
            "D": 40,  # Before
            "E": 50,  # Action
            "F": 40,  # After
            "G": 50,  # Reason
            "H": 40,  # validated_after
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        logger.info(f"Created '{CHANGE_LOG_SHEET}' sheet with headers")
        return ws
    
    return wb[CHANGE_LOG_SHEET]


def get_next_log_row(ws) -> int:
    """Get the next empty row in the Change Log sheet."""
    return ws.max_row + 1


def capture_range_values_from_file(filepath: str, sheet_name: str, start_cell: str, end_cell: Optional[str] = None) -> str:
    """
    Capture range values by loading the file fresh from disk.
    This ensures we get the actual saved values, not cached data.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (e.g., "A1")
        end_cell: Ending cell (e.g., "C5"), optional for single cell
        
    Returns:
        JSON string representation of the range values.
    """
    import os
    
    if not os.path.exists(filepath):
        logger.warning(f"File does not exist: {filepath}")
        return json.dumps([])
    
    try:
        # Load workbook fresh from disk
        # data_only=True gets calculated values instead of formulas
        # We'll try both ways to get the best representation
        wb_formulas = load_workbook(filepath, data_only=False)
        
        if sheet_name not in wb_formulas.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            wb_formulas.close()
            return json.dumps([])
        
        ws = wb_formulas[sheet_name]
        result = capture_range_values(ws, start_cell, end_cell)
        wb_formulas.close()
        
        return result
        
    except Exception as e:
        logger.error(f"Error capturing range values from file: {e}")
        return json.dumps({"error": str(e)})


def capture_range_values(ws, start_cell: str, end_cell: Optional[str] = None) -> str:
    """
    Capture all cell values in a range as a JSON string.
    
    Args:
        ws: openpyxl Worksheet object
        start_cell: Starting cell (e.g., "A1")
        end_cell: Ending cell (e.g., "C5"), optional for single cell
        
    Returns:
        JSON string representation of the range values.
        - Numeric values are kept as-is
        - Formulas are prefixed with ' for display (e.g., '=SUM(A1:A10))
        - Text values are kept as-is
    """
    from .cell_utils import parse_cell_range
    
    if end_cell is None:
        end_cell = start_cell
    
    try:
        # parse_cell_range expects two separate arguments, not a colon-separated string
        coords = parse_cell_range(start_cell, end_cell)
        start_row, start_col, end_row, end_col = coords[0], coords[1], coords[2], coords[3]
        
        # If end coordinates are None (single cell), use start coordinates
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
            
    except Exception as e:
        logger.warning(f"Failed to parse range {start_cell}:{end_cell}: {e}")
        return json.dumps({"error": f"Invalid range: {start_cell}:{end_cell}"})
    
    values = []
    for row in range(start_row, end_row + 1):
        row_values = []
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            
            # Log for debugging
            logger.debug(f"Cell {get_column_letter(col)}{row}: value={value}, type={type(value)}, data_type={getattr(cell, 'data_type', 'N/A')}")
            
            # Check if cell has a formula
            is_formula = False
            formula_str = None
            
            # Method 1: Check data_type
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                is_formula = True
                formula_str = value
            
            # Method 2: Check if value starts with '=' (string formula)
            if isinstance(value, str) and value.startswith('='):
                is_formula = True
                formula_str = value
            
            # Method 3: Check for formula attribute (some openpyxl versions)
            if hasattr(cell, 'value') and cell.value is not None:
                cell_val = cell.value
                if hasattr(cell_val, 'formula'):
                    is_formula = True
                    formula_str = f"={cell_val.formula}" if not str(cell_val.formula).startswith('=') else cell_val.formula
            
            if is_formula and formula_str:
                # Prefix formula with ' so it displays as text in Excel
                if str(formula_str).startswith('='):
                    row_values.append(f"'{formula_str}")
                else:
                    row_values.append(f"'={formula_str}")
            elif value is None:
                # Keep None as empty string for cleaner display
                row_values.append("")
            elif isinstance(value, (int, float)):
                # Keep numeric values as-is
                row_values.append(value)
            else:
                # Text and other values
                row_values.append(value)
                
        values.append(row_values)
    
    return json.dumps(values, default=str)


def calculate_end_cell_from_data(start_cell: str, data: List[List]) -> str:
    """
    Calculate the end cell based on start cell and data dimensions.
    
    Args:
        start_cell: Starting cell (e.g., "A1")
        data: 2D list of data
        
    Returns:
        End cell address (e.g., "C5")
    """
    from .cell_utils import parse_cell_range
    
    if not data or not data[0]:
        return start_cell
    
    # parse_cell_range with single cell returns (row, col, None, None)
    coords = parse_cell_range(start_cell)
    start_row, start_col = coords[0], coords[1]
    
    num_rows = len(data)
    num_cols = max(len(row) for row in data) if data else 1
    
    end_row = start_row + num_rows - 1
    end_col = start_col + num_cols - 1
    
    return f"{get_column_letter(end_col)}{end_row}"


def log_change_entry(filepath: str, entry: ChangeLogEntry) -> None:
    """
    Append a change log entry to the Change Log sheet.
    
    Args:
        filepath: Path to the Excel file
        entry: ChangeLogEntry object with all fields
    """
    try:
        wb = load_workbook(filepath)
        log_ws = ensure_log_sheet_exists(wb)
        
        next_row = get_next_log_row(log_ws)
        
        # Write entry data
        row_data = [
            entry.datetime,
            entry.sheet,
            entry.range,
            entry.before,
            entry.action,
            entry.after,
            entry.reason,
            entry.validated_after,
        ]
        
        # Style for data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, value in enumerate(row_data, 1):
            cell = log_ws.cell(row=next_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        wb.save(filepath)
        wb.close()
        
        logger.info(f"Logged change to '{CHANGE_LOG_SHEET}': {entry.action[:50]}...")
        
    except Exception as e:
        logger.error(f"Failed to log change entry: {e}")
        # Don't raise - logging failure shouldn't break the operation


def wrap_with_change_log(
    filepath: str,
    sheet_name: str,
    range_str: str,
    action: str,
    expected_after: str,
    reason: str,
    execute_fn,
    *args,
    **kwargs
) -> Any:
    """
    Wrapper function that handles change logging around an Excel operation.
    
    This function:
    1. Captures DateTime
    2. Captures 'before' state from the file
    3. Executes the operation
    4. Captures 'validated_after' state from the file
    5. Logs the complete entry
    
    Args:
        filepath: Path to Excel file
        sheet_name: Sheet being modified
        range_str: Range being modified (e.g., "A1:C5")
        action: Description of what's being done (from Claude)
        expected_after: Expected result (from Claude)
        reason: Why this change is being made (from Claude)
        execute_fn: The function to execute
        *args, **kwargs: Arguments to pass to execute_fn
        
    Returns:
        Result from execute_fn
    """
    import os
    
    # 1. Capture DateTime
    dt = get_current_datetime()
    
    # 2. Capture 'before' state
    before_values = "[]"
    
    # Parse range
    if ":" in range_str:
        start_cell, end_cell = range_str.split(":")
    else:
        start_cell = range_str
        end_cell = range_str
    
    try:
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                before_values = capture_range_values(ws, start_cell, end_cell)
            wb.close()
    except Exception as e:
        logger.warning(f"Could not capture 'before' state: {e}")
        before_values = json.dumps({"error": str(e)})
    
    # 3. Execute the operation
    result = execute_fn(*args, **kwargs)
    
    # 4. Capture 'validated_after' state
    validated_after = "[]"
    try:
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                validated_after = capture_range_values(ws, start_cell, end_cell)
            wb.close()
    except Exception as e:
        logger.warning(f"Could not capture 'validated_after' state: {e}")
        validated_after = json.dumps({"error": str(e)})
    
    # 5. Log the change
    entry = ChangeLogEntry(
        datetime=dt,
        sheet=sheet_name,
        range=range_str,
        before=before_values,
        action=action,
        after=expected_after,
        reason=reason,
        validated_after=validated_after,
    )
    
    log_change_entry(filepath, entry)
    
    return result

