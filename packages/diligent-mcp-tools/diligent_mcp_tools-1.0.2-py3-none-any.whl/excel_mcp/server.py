import logging
import os
import asyncio
from typing import Any, List, Dict, Optional

from mcp.server.fastmcp import FastMCP

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
)

# Import change logging module
from excel_mcp.change_log import (
    wrap_with_change_log,
    calculate_end_cell_from_data,
    ChangeLogEntry,
    get_current_datetime,
    capture_range_values,
    capture_range_values_from_file,
    log_change_entry,
)

# Import browser automation functions (Playwright)
from excel_mcp.browser import (
    navigate_to_url_impl,
    take_screenshot_impl,
    click_element_impl,
    fill_input_impl,
    get_text_content_impl,
    wait_for_selector_impl,
    evaluate_javascript_impl,
    get_page_content_impl,
    select_option_impl,
    check_checkbox_impl,
    hover_element_impl,
    press_key_impl,
    get_element_attribute_impl,
    scroll_page_impl,
    get_cookies_impl,
    set_cookie_impl,
    clear_cookies_impl,
    cleanup_browser
)

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging - use user's home directory or temp for log file
def _get_log_file_path() -> str:
    """Get a writable path for the log file."""
    # Try user's home directory first
    home_log = os.path.join(os.path.expanduser("~"), ".diligent-mcp", "excel-mcp.log")
    try:
        os.makedirs(os.path.dirname(home_log), exist_ok=True)
        return home_log
    except (OSError, PermissionError):
        pass
    # Fallback to temp directory
    import tempfile
    return os.path.join(tempfile.gettempdir(), "excel-mcp.log")

LOG_FILE = _get_log_file_path()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")

# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files. All data-modifying tools require 'action', 'expected_after', and 'reason' parameters for change logging."
)

def get_excel_path(filename: str) -> str:
    """Get full path to Excel file."""
    if os.path.isabs(filename):
        return filename
    if EXCEL_FILES_PATH is None:
        raise ValueError(f"Invalid filename: {filename}, must be an absolute path when not in SSE mode")
    return os.path.join(EXCEL_FILES_PATH, filename)


# =============================================================================
# TOOLS WITH CHANGE LOGGING (Data-modifying tools)
# =============================================================================

@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str,
    action: str,
    expected_after: str,
    reason: str,
) -> str:
    """
    Write data to Excel worksheet WITH CHANGE LOGGING.

    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Name of worksheet to write to
    - data: List of lists containing data to write (rows)
    - start_cell: Cell to start writing to (e.g., "A1")
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: Plain text description of what you're doing.
              Example: "Writing new employee records" or "Updating Q4 revenue figures"
    - expected_after: THE EXACT VALUES you are writing to the range as a JSON array.
              THIS MUST BE THE ACTUAL DATA, NOT A DESCRIPTION!
              Example: [[100, 200, 300], [400, 500, 600]]
              Example: [["John", "Doe", 50000], ["Jane", "Smith", 60000]]
              Example: [["'=A1*2"], ["'=A2*2"]] (for formulas, prefix with ')
    - reason: Why this change is part of the larger workflow.
              Example: "Adding new hire from HR onboarding" or "Correcting data entry error"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl.utils import get_column_letter
        from excel_mcp.cell_utils import parse_cell_range
        
        # Calculate end cell from data dimensions
        end_cell = calculate_end_cell_from_data(start_cell, data)
        range_str = f"{start_cell}:{end_cell}"
        
        # Capture before state - load fresh from disk
        dt = get_current_datetime()
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        logger.info(f"Captured BEFORE state for {range_str}: {before_values[:100]}...")
        
        # Execute the write operation
        result = write_data(full_path, sheet_name, data, start_cell)
        
        # Capture after state (validated_after) - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        logger.info(f"Captured AFTER state for {range_str}: {validated_after[:100]}...")
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise


@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
    action: str,
    expected_after: str,
    reason: str,
) -> str:
    """
    Apply Excel formula to cell WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - cell: Target cell reference (e.g., "A1")
    - formula: Excel formula to apply (e.g., "=SUM(A1:A10)")
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What formula operation you're performing.
              Example: "Adding SUM formula for total revenue"
    - expected_after: THE EXACT FORMULA you are writing, prefixed with ' for display.
              THIS MUST BE THE ACTUAL FORMULA, NOT A DESCRIPTION!
              Example: [["'=SUM(B2:B10)"]]
              Example: [["'=VLOOKUP(A1, Data!A:B, 2, FALSE)"]]
    - reason: Why this formula is needed.
              Example: "Calculating monthly totals for report"
    """
    try:
        full_path = get_excel_path(filepath)
        
        range_str = f"{cell}:{cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, cell, cell)
        logger.info(f"Captured BEFORE state for {range_str}: {before_values}")
        
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # Apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, cell, cell)
        logger.info(f"Captured AFTER state for {range_str}: {validated_after}")
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise


@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    action: str,
    expected_after: str,
    reason: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """
    Apply formatting to a range of cells WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_cell: Starting cell of range
    - end_cell: Ending cell of range (optional for single cell)
    - Formatting options: bold, italic, underline, font_size, font_color, bg_color, etc.
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What formatting you're applying.
              Example: "Making headers bold with yellow background"
    - expected_after: THE CELL VALUES in the range (formatting doesn't change values).
              THIS MUST BE THE ACTUAL VALUES, NOT A DESCRIPTION!
              Example: [["Header1", "Header2"]] (the values stay the same, only style changes)
    - reason: Why this formatting is needed.
              Example: "Highlighting important data per style guide"
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        effective_end_cell = end_cell or start_cell
        range_str = f"{start_cell}:{effective_end_cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, effective_end_cell)
        
        # Apply formatting
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,
            font_color=font_color,
            bg_color=bg_color,
            border_style=border_style,
            border_color=border_color,
            number_format=number_format,
            alignment=alignment,
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,
            conditional_format=conditional_format
        )
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, effective_end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise


@mcp.tool()
def format_range_matrix(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    styles: List[List[Optional[Dict[str, Any]]]],
    action: str,
    expected_after: str,
    reason: str,
) -> str:
    """
    Apply formatting to cells using a 2D style matrix WITH CHANGE LOGGING.
    
    Style properties per cell:
    - font: {bold, italic, underline, size, color, strike}
    - fill: {color} - background color as hex
    - border: {style, color}
    - numFmt: number format string
    - alignment: {horizontal, vertical, wrapText}
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Name of worksheet
    - start_cell: Top-left cell of the range
    - styles: 2D array of style objects (use null for cells to skip)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What formatting you're applying.
              Example: "Applying conditional formatting to data cells"
    - expected_after: Description of the formatting.
              Example: "Header row bold, data cells with borders"
    - reason: Why this formatting is needed.
              Example: "Preparing report for presentation"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from excel_mcp.cell_utils import parse_cell_range
        
        # Calculate end cell from styles dimensions
        num_rows = len(styles)
        num_cols = max(len(row) for row in styles) if styles else 1
        coords = parse_cell_range(f"{start_cell}:{start_cell}")
        end_row = coords[0] + num_rows - 1
        end_col = coords[1] + num_cols - 1
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        range_str = f"{start_cell}:{end_cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Apply formatting
        wb = load_workbook(full_path)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
        ws = wb[sheet_name]
        
        start_row, start_col = coords[0], coords[1]
        cells_formatted = 0
        
        for row_idx, row_styles in enumerate(styles):
            for col_idx, cell_style in enumerate(row_styles):
                if cell_style is None:
                    continue
                    
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                
                if "font" in cell_style:
                    font_dict = cell_style["font"]
                    cell.font = Font(
                        bold=font_dict.get("bold"),
                        italic=font_dict.get("italic"),
                        underline="single" if font_dict.get("underline") else None,
                        size=font_dict.get("size"),
                        color=font_dict.get("color"),
                        strike=font_dict.get("strike")
                    )
                
                if "fill" in cell_style:
                    fill_dict = cell_style["fill"]
                    color = fill_dict.get("color", "FFFFFF")
                    if color.startswith("#"):
                        color = color[1:]
                    if len(color) == 6:
                        color = "FF" + color
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                if "border" in cell_style:
                    border_dict = cell_style["border"]
                    style = border_dict.get("style", "thin")
                    color = border_dict.get("color", "000000")
                    if color.startswith("#"):
                        color = color[1:]
                    side = Side(style=style, color=color)
                    cell.border = Border(top=side, bottom=side, left=side, right=side)
                
                if "numFmt" in cell_style:
                    cell.number_format = cell_style["numFmt"]
                
                if "alignment" in cell_style:
                    align_dict = cell_style["alignment"]
                    cell.alignment = Alignment(
                        horizontal=align_dict.get("horizontal"),
                        vertical=align_dict.get("vertical"),
                        wrap_text=align_dict.get("wrapText", False)
                    )
                
                cells_formatted += 1
        
        wb.save(full_path)
        wb.close()
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return f"Successfully formatted {cells_formatted} cells starting at {start_cell}"
        
    except Exception as e:
        logger.error(f"Error formatting range matrix: {e}")
        raise


@mcp.tool()
def merge_cells(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    action: str,
    expected_after: str,
    reason: str,
) -> str:
    """
    Merge a range of cells WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_cell: Starting cell of range
    - end_cell: Ending cell of range
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What you're merging.
              Example: "Merging header cells A1:D1"
    - expected_after: Description of merged result.
              Example: "Single merged cell spanning A1:D1"
    - reason: Why merging is needed.
              Example: "Creating centered title row"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        
        range_str = f"{start_cell}:{end_cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise


@mcp.tool()
def unmerge_cells(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    action: str,
    expected_after: str,
    reason: str,
) -> str:
    """
    Unmerge a range of cells WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_cell: Starting cell of range
    - end_cell: Ending cell of range
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What you're unmerging.
              Example: "Unmerging cells A1:D1"
    - expected_after: Description of unmerged result.
              Example: "Separate cells A1, B1, C1, D1"
    - reason: Why unmerging is needed.
              Example: "Need individual cell access for data entry"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        
        range_str = f"{start_cell}:{end_cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise


@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    action: str,
    expected_after: str,
    reason: str,
    target_sheet: Optional[str] = None,
) -> str:
    """
    Copy a range of cells to another location WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Source worksheet name
    - source_start: Starting cell of source range
    - source_end: Ending cell of source range
    - target_start: Starting cell for paste
    - target_sheet: Optional target worksheet name
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What you're copying.
              Example: "Copying data from A1:C10 to E1"
    - expected_after: Description of copied data.
              Example: "Duplicate of source range at new location"
    - reason: Why copying is needed.
              Example: "Creating backup of original data"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from excel_mcp.cell_utils import parse_cell_range
        from excel_mcp.sheet import copy_range_operation
        
        # Calculate target end cell
        source_coords = parse_cell_range(f"{source_start}:{source_end}")
        target_coords = parse_cell_range(f"{target_start}:{target_start}")
        
        source_rows = source_coords[2] - source_coords[0] + 1
        source_cols = source_coords[3] - source_coords[1] + 1
        
        target_end_row = target_coords[0] + source_rows - 1
        target_end_col = target_coords[1] + source_cols - 1
        target_end = f"{get_column_letter(target_end_col)}{target_end_row}"
        
        effective_target_sheet = target_sheet or sheet_name
        range_str = f"{target_start}:{target_end}"
        dt = get_current_datetime()
        
        # Capture before state (of target range) - load fresh from disk
        before_values = capture_range_values_from_file(full_path, effective_target_sheet, target_start, target_end)
        
        result = copy_range_operation(
            full_path, sheet_name, source_start, source_end,
            target_start, effective_target_sheet
        )
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, effective_target_sheet, target_start, target_end)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=effective_target_sheet,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise


@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    action: str,
    expected_after: str,
    reason: str,
    shift_direction: str = "up",
) -> str:
    """
    Delete a range of cells and shift remaining cells WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_cell: Starting cell of range
    - end_cell: Ending cell of range
    - shift_direction: Direction to shift cells ("up" or "left")
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What you're deleting.
              Example: "Deleting rows 5-10 and shifting up"
    - expected_after: Description of result.
              Example: "Cells below shifted up to fill gap"
    - reason: Why deletion is needed.
              Example: "Removing obsolete data entries"
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        
        range_str = f"{start_cell}:{end_cell}"
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        result = delete_range_operation(
            full_path, sheet_name, start_cell, end_cell, shift_direction
        )
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise


@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    action: str,
    expected_after: str,
    reason: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9",
) -> str:
    """
    Create a native Excel table from a range WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Sheet name where table is created
    - data_range: Range for the table (e.g., "A1:D10")
    - table_name: Optional unique name for the table
    - table_style: Visual style (default: TableStyleMedium9)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What table you're creating.
              Example: "Creating table from sales data A1:D50"
    - expected_after: Description of the table.
              Example: "Formatted table with headers and filter dropdowns"
    - reason: Why table is needed.
              Example: "Enable sorting and filtering for data analysis"
    """
    try:
        full_path = get_excel_path(filepath)
        
        # Parse data_range
        if ":" in data_range:
            start_cell, end_cell = data_range.split(":")
        else:
            start_cell = end_cell = data_range
        
        range_str = data_range
        dt = get_current_datetime()
        
        # Capture before state - load fresh from disk
        before_values = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        
        # Capture after state - load fresh from disk
        validated_after = capture_range_values_from_file(full_path, sheet_name, start_cell, end_cell)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=before_values,
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=validated_after,
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise


@mcp.tool()
def insert_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    action: str,
    expected_after: str,
    reason: str,
    count: int = 1,
) -> str:
    """
    Insert rows starting at specified row WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_row: Row number where to start inserting (1-based)
    - count: Number of rows to insert (default: 1)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What rows you're inserting.
              Example: "Inserting 3 rows at row 5"
    - expected_after: Description of result.
              Example: "3 new blank rows at position 5-7"
    - reason: Why rows are needed.
              Example: "Making space for new data entries"
    """
    try:
        full_path = get_excel_path(filepath)
        
        range_str = f"Row {start_row} to Row {start_row + count - 1}"
        dt = get_current_datetime()
        
        result = insert_row(full_path, sheet_name, start_row, count)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before="[existing rows]",
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=f"[{count} new rows inserted at row {start_row}]",
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise


@mcp.tool()
def insert_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    action: str,
    expected_after: str,
    reason: str,
    count: int = 1,
) -> str:
    """
    Insert columns starting at specified column WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_col: Column number where to start inserting (1-based, 1=A)
    - count: Number of columns to insert (default: 1)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What columns you're inserting.
              Example: "Inserting 2 columns at column C"
    - expected_after: Description of result.
              Example: "2 new blank columns at position C-D"
    - reason: Why columns are needed.
              Example: "Adding space for new data fields"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl.utils import get_column_letter
        
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(start_col + count - 1)
        range_str = f"Column {start_letter} to Column {end_letter}"
        dt = get_current_datetime()
        
        result = insert_cols(full_path, sheet_name, start_col, count)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before="[existing columns]",
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=f"[{count} new columns inserted at column {start_letter}]",
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise


@mcp.tool()
def delete_sheet_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    action: str,
    expected_after: str,
    reason: str,
    count: int = 1,
) -> str:
    """
    Delete rows starting at specified row WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_row: Row number where to start deleting (1-based)
    - count: Number of rows to delete (default: 1)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What rows you're deleting.
              Example: "Deleting rows 5-7"
    - expected_after: Description of result.
              Example: "Rows removed, below rows shifted up"
    - reason: Why rows are being deleted.
              Example: "Removing duplicate or obsolete entries"
    """
    try:
        full_path = get_excel_path(filepath)
        
        range_str = f"Row {start_row} to Row {start_row + count - 1}"
        dt = get_current_datetime()
        
        result = delete_rows(full_path, sheet_name, start_row, count)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=f"[{count} rows at position {start_row}]",
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=f"[{count} rows deleted, remaining rows shifted up]",
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise


@mcp.tool()
def delete_sheet_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    action: str,
    expected_after: str,
    reason: str,
    count: int = 1,
) -> str:
    """
    Delete columns starting at specified column WITH CHANGE LOGGING.
    
    PARAMETERS:
    - filepath: Path to Excel file
    - sheet_name: Target worksheet name
    - start_col: Column number where to start deleting (1-based, 1=A)
    - count: Number of columns to delete (default: 1)
    
    REQUIRED FOR CHANGE LOG (Claude MUST provide these):
    - action: What columns you're deleting.
              Example: "Deleting columns C-D"
    - expected_after: Description of result.
              Example: "Columns removed, right columns shifted left"
    - reason: Why columns are being deleted.
              Example: "Removing unused data fields"
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl.utils import get_column_letter
        
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(start_col + count - 1)
        range_str = f"Column {start_letter} to Column {end_letter}"
        dt = get_current_datetime()
        
        result = delete_cols(full_path, sheet_name, start_col, count)
        
        # Log the change
        entry = ChangeLogEntry(
            datetime=dt,
            sheet=sheet_name,
            range=range_str,
            before=f"[{count} columns at position {start_letter}]",
            action=action,
            after=expected_after,
            reason=reason,
            validated_after=f"[{count} columns deleted, remaining columns shifted left]",
        )
        log_change_entry(full_path, entry)
        
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise


# =============================================================================
# READ-ONLY TOOLS (No change logging needed)
# =============================================================================

@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise


@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """Read data from Excel worksheet with cell metadata including validation rules."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(full_path, sheet_name, start_cell, end_cell)
        if not result or not result.get("cells"):
            return "No data found in specified range"
        import json
        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise


@mcp.tool()
def read_data_with_styles(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    include_formulas: bool = False
) -> str:
    """Read data from Excel worksheet with FULL STYLE information per cell."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_styles
        result = read_excel_range_with_styles(
            full_path, sheet_name, start_cell, end_cell,
            include_values=True, include_formulas=include_formulas
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
        import json
        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error reading data with styles: {e}")
        raise


@mcp.tool()
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise


@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise


@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise


@mcp.tool()
def get_data_validation_info(filepath: str, sheet_name: str) -> str:
    """Get all data validation rules in a worksheet."""
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
        import json
        return json.dumps({"sheet_name": sheet_name, "validation_rules": validations}, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise


# =============================================================================
# TOOLS WITHOUT CHANGE LOGGING (No cell/range operations)
# =============================================================================

@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise


@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise


@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path, sheet_name=sheet_name, data_range=data_range,
            chart_type=chart_type, target_cell=target_cell,
            title=title, x_axis=x_axis, y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise


@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path, sheet_name=sheet_name, data_range=data_range,
            rows=rows, values=values, columns=columns or [], agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise


@mcp.tool()
def copy_worksheet(filepath: str, source_sheet: str, target_sheet: str) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise


@mcp.tool()
def delete_worksheet(filepath: str, sheet_name: str) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise


@mcp.tool()
def rename_worksheet(filepath: str, old_name: str, new_name: str) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise


# =============================================================================
# SERVER RUNNERS
# =============================================================================

def run_sse():
    """Run Excel MCP server in SSE mode."""
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")


def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")


# =============================================================================
# BROWSER AUTOMATION TOOLS (Playwright)
# =============================================================================

@mcp.tool()
async def browser_navigate(
    url: str,
    wait_until: str = "load",
    timeout: int = 30000
) -> str:
    """
    Navigate to a URL using the browser.
    
    Args:
        url: URL to navigate to
        wait_until: When to consider navigation successful ("load", "domcontentloaded", "networkidle")
        timeout: Navigation timeout in milliseconds
        
    Returns:
        JSON string with navigation result
    """
    try:
        import json
        result = await navigate_to_url_impl(url, wait_until, timeout)
        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error navigating: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_screenshot(
    filepath: str,
    full_page: bool = False,
    selector: Optional[str] = None
) -> str:
    """
    Take a screenshot of the current page or element.
    
    Args:
        filepath: Path where to save the screenshot
        full_page: Whether to capture the full scrollable page
        selector: Optional CSS selector for specific element screenshot
        
    Returns:
        Success message with file path
    """
    try:
        result = await take_screenshot_impl(filepath, full_page, selector)
        return result["message"]
    except Exception as e:
        logger.error(f"Error taking screenshot: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_click(
    selector: str,
    button: str = "left",
    click_count: int = 1,
    timeout: int = 30000
) -> str:
    """
    Click an element on the page.
    
    Args:
        selector: CSS selector for the element to click
        button: Mouse button to use ("left", "right", "middle")
        click_count: Number of clicks (1 for single, 2 for double)
        timeout: Timeout in milliseconds
        
    Returns:
        Success message
    """
    try:
        result = await click_element_impl(selector, button, click_count, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error clicking element: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_fill_input(
    selector: str,
    value: str,
    timeout: int = 30000
) -> str:
    """
    Fill an input field with text.
    
    Args:
        selector: CSS selector for the input element
        value: Text to fill into the input
        timeout: Timeout in milliseconds
        
    Returns:
        Success message
    """
    try:
        result = await fill_input_impl(selector, value, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error filling input: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_get_text(
    selector: str,
    timeout: int = 30000
) -> str:
    """
    Get text content of an element.
    
    Args:
        selector: CSS selector for the element
        timeout: Timeout in milliseconds
        
    Returns:
        Text content of the element
    """
    try:
        import json
        result = await get_text_content_impl(selector, timeout)
        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error getting text: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_wait_for_element(
    selector: str,
    state: str = "visible",
    timeout: int = 30000
) -> str:
    """
    Wait for an element to be in a specific state.
    
    Args:
        selector: CSS selector for the element
        state: State to wait for ("attached", "detached", "visible", "hidden")
        timeout: Timeout in milliseconds
        
    Returns:
        Success message when element reaches the specified state
    """
    try:
        result = await wait_for_selector_impl(selector, state, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error waiting for element: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_execute_js(script: str) -> str:
    """
    Execute JavaScript code in the browser page context.
    
    Args:
        script: JavaScript code to execute
        
    Returns:
        Result of the JavaScript execution
    """
    try:
        import json
        result = await evaluate_javascript_impl(script)
        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error executing JavaScript: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_get_page_content() -> str:
    """
    Get the full HTML content of the current page.
    
    Returns:
        JSON string with URL, title, and HTML content
    """
    try:
        import json
        result = await get_page_content_impl()
        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error getting page content: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_select_option(
    selector: str,
    value: Optional[str] = None,
    label: Optional[str] = None,
    index: Optional[int] = None,
    timeout: int = 30000
) -> str:
    """
    Select an option in a dropdown/select element.
    
    Args:
        selector: CSS selector for the select element
        value: Option value to select (optional)
        label: Option label to select (optional)
        index: Option index to select (optional, 0-based)
        timeout: Timeout in milliseconds
        
    Note: Must provide one of value, label, or index
        
    Returns:
        Success message
    """
    try:
        result = await select_option_impl(selector, value, label, index, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error selecting option: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_check_checkbox(
    selector: str,
    checked: bool = True,
    timeout: int = 30000
) -> str:
    """
    Check or uncheck a checkbox.
    
    Args:
        selector: CSS selector for the checkbox element
        checked: True to check, False to uncheck
        timeout: Timeout in milliseconds
        
    Returns:
        Success message
    """
    try:
        result = await check_checkbox_impl(selector, checked, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error checking checkbox: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_hover(
    selector: str,
    timeout: int = 30000
) -> str:
    """
    Hover over an element.
    
    Args:
        selector: CSS selector for the element
        timeout: Timeout in milliseconds
        
    Returns:
        Success message
    """
    try:
        result = await hover_element_impl(selector, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error hovering element: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_press_key(
    key: str,
    selector: Optional[str] = None,
    timeout: int = 30000
) -> str:
    """
    Press a keyboard key.
    
    Args:
        key: Key to press (e.g., "Enter", "Escape", "a", "Control+A", "ArrowDown")
        selector: Optional CSS selector to focus before pressing key
        timeout: Timeout in milliseconds
        
    Returns:
        Success message
    """
    try:
        result = await press_key_impl(key, selector, timeout)
        return result["message"]
    except Exception as e:
        logger.error(f"Error pressing key: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_get_attribute(
    selector: str,
    attribute: str,
    timeout: int = 30000
) -> str:
    """
    Get an attribute value from an element.
    
    Args:
        selector: CSS selector for the element
        attribute: Attribute name (e.g., "href", "class", "id", "value")
        timeout: Timeout in milliseconds
        
    Returns:
        JSON string with attribute value
    """
    try:
        import json
        result = await get_element_attribute_impl(selector, attribute, timeout)
        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error getting attribute: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_scroll(
    direction: str = "down",
    amount: Optional[int] = None
) -> str:
    """
    Scroll the page in a specified direction.
    
    Args:
        direction: Scroll direction ("up", "down", "left", "right")
        amount: Amount to scroll in pixels (defaults to viewport height/width)
        
    Returns:
        Success message
    """
    try:
        result = await scroll_page_impl(direction, amount)
        return result["message"]
    except Exception as e:
        logger.error(f"Error scrolling page: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_get_cookies() -> str:
    """
    Get all cookies from the current browser context.
    
    Returns:
        JSON string with list of cookies
    """
    try:
        import json
        result = await get_cookies_impl()
        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error getting cookies: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_set_cookie(
    name: str,
    value: str,
    url: Optional[str] = None,
    domain: Optional[str] = None,
    path: str = "/",
    expires: Optional[float] = None,
    http_only: bool = False,
    secure: bool = False,
    same_site: str = "Lax"
) -> str:
    """
    Set a cookie in the browser context.
    
    Args:
        name: Cookie name
        value: Cookie value
        url: Optional URL (requires protocol and domain)
        domain: Optional domain (e.g., ".example.com")
        path: Cookie path (default: "/")
        expires: Optional expiration timestamp (Unix time)
        http_only: HttpOnly flag (default: False)
        secure: Secure flag (default: False)
        same_site: SameSite attribute ("Strict", "Lax", or "None", default: "Lax")
        
    Returns:
        Success message
    """
    try:
        result = await set_cookie_impl(name, value, url, domain, path, expires, http_only, secure, same_site)
        return result["message"]
    except Exception as e:
        logger.error(f"Error setting cookie: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_clear_cookies() -> str:
    """
    Clear all cookies from the browser context.
    
    Returns:
        Success message
    """
    try:
        result = await clear_cookies_impl()
        return result["message"]
    except Exception as e:
        logger.error(f"Error clearing cookies: {e}")
        return f"Error: {str(e)}"


@mcp.tool()
async def browser_cleanup() -> str:
    """
    Clean up browser resources (close browser, context, and page).
    Use this when you're done with browser automation to free up resources.
    
    Returns:
        Success message
    """
    try:
        await cleanup_browser()
        return "Browser cleanup completed successfully"
    except Exception as e:
        logger.error(f"Error cleaning up browser: {e}")
        return f"Error: {str(e)}"


# =============================================================================
# SERVER RUNNERS
# =============================================================================

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
