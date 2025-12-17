from eaasy import Common, BaseEntity, limiter
from eaasy.extensions.helper import verify_oidc
from typing import Any, Iterable, Literal
from flask_restx import Namespace, Resource, Model, OrderedModel, fields
from flask_restx.reqparse import RequestParser
from flask_restx._http import HTTPStatus
from sqlalchemy.orm.attributes import InstrumentedAttribute
from eaasy.domain.exceptions import ErrorResponse
from flask import request, send_file
from datetime import datetime, timezone
from flask_oidc import OpenIDConnect
from authlib.integrations.flask_oauth2 import current_token
from io import BytesIO
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage

TokenProperties = Literal['sub', 'email_verified', 'email', 'name', 'preferred_username', 'given_name', 'family_name', 'username'] 


def _to_serializable_dict(item: Any) -> dict[str, Any]:
    if isinstance(item, dict):
        return {k: v for k, v in item.items() if not k.startswith('_')}
    obj = getattr(item, '__dict__', {})
    return {k: v for k, v in obj.items() if not k.startswith('_')}


def _collect_headers_from_records(records: Iterable[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for record in records:
        for key in record.keys():
            if key not in headers:
                headers.append(key)
    return headers


def _build_workbook(headers: list[str], rows: Iterable[list[Any]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    if headers:
        worksheet.append(headers) # type: ignore
    for row in rows:
        worksheet.append(row) # type: ignore
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _read_workbook_records(stream: BytesIO, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(stream, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True)) # type: ignore
    if not rows: # pragma: no cover
        return [], []

    headers_raw = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    headers = [header for header in headers_raw if header]
    normalized_headers = [header if header else None for header in headers_raw]

    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        row_dict: dict[str, Any] = {}
        for header, value in zip(normalized_headers, values):
            if header:
                row_dict[header] = value
        if any(value is not None for value in row_dict.values()):
            records.append(row_dict)

    return headers, records


def build_model(entity: Common | Any, namespace: Namespace | None = None) -> tuple[Namespace, Model | OrderedModel]:
    ns = namespace if namespace is not None \
        else Namespace(entity.__name__, description=f"{entity.__name__} operations", path=f"/{entity.__name__.lower()}")
    model_name = entity.__name__
    model_attributes = {}

    attributes = [x for x in entity.__dict__.keys() if not x.startswith('_')]

    for attr in attributes:
        attribute: InstrumentedAttribute = entity.__dict__[attr]
        if not attr.startswith('_'):
            if f'{attribute.type}' == 'INTEGER':
                model_attributes[attr] = fields.Integer(
                    required=not attribute.nullable, 
                    description=attr,
                    default=0 if not attribute.nullable else None,
                    example=0)
            elif 'NUMERIC' in f'{attribute.type}':
                model_attributes[attr] = fields.Float(
                    required=not attribute.nullable, 
                    description=attr,
                    default=0.0 if not attribute.nullable else None,
                    example=0.0)
            elif f'{attribute.type}' == 'VARCHAR':
                model_attributes[attr] = fields.String(
                    required=not attribute.nullable, 
                    description=attr,
                    default='' if not attribute.nullable else None,
                    example='string')
            elif f'{attribute.type}' == 'DATETIME':
                model_attributes[attr] = fields.DateTime(
                    required=not attribute.nullable, 
                    description=attr,
                    default=datetime.strftime(datetime.now(timezone.utc), '%Y-%m-%dT%H:%M:%S') if not attribute.nullable else None,
                    example=datetime.strftime(datetime.now(timezone.utc), '%Y-%m-%dT%H:%M:%S'))
            elif f'{attribute.type}' == 'BOOLEAN':
                model_attributes[attr] = fields.Boolean(
                    required=not attribute.nullable, 
                    description=attr,
                    default=False,
                    example=False)
            else:
                print(f"\033[33mType {attribute.type} not supported\033[0m")

    model = ns.model(model_name, model_attributes) 

    return ns, model

def build_dynamic_class(name: str, base_classes: list[type], **kwargs) -> type:
    return type(name, tuple(base_classes), kwargs)

def build_resource(
        entity: BaseEntity, 
        namespace: Namespace, 
        get_model: Model | OrderedModel,
        upsert_model: Model | OrderedModel,
        **kwargs) -> None:
    
    def _log(level: Literal['debug', 'info', 'warning', 'error'], message: str):
        logger = kwargs.get('logger', None)
        if logger is not None: # pragma: no cover
            getattr(logger, level)(message)

    def callback(name: str, data: Any = None):
        on_event = kwargs.get(name, None)
        if on_event is not None:
            try:
                on_event(data)
            except Exception as e:
                _log('error', f'Error on {name} event: {e}')
                on_callback_fail = kwargs.get(f'{name}_fail', None)
                if on_callback_fail is not None:
                    on_callback_fail(e)
                else: 
                    raise Exception({
                        'status_code': 500,
                        'message': str(e),
                        'data': data
                    })

    def get_limit_rate(name: str) -> str | None:
        limiter = kwargs.get('limit', None)
        return limiter if limiter is not None else kwargs.get(name, None)
    
    def get_oidc(name: str) -> OpenIDConnect | None:
        oidc = kwargs.get(name, None)
        return oidc if oidc is not None else kwargs.get('oidc', None)
    
    def get_current_user_info(property: TokenProperties) -> str: # pragma: no cover
        return dict(current_token)[property] # type: ignore
    
    custom_wrapper = kwargs.get('custom_wrapper', None)
    if custom_wrapper is not None:
        assert callable(custom_wrapper)
    else:
        custom_wrapper = lambda f: f
    
    class DynamicResourceGetAll(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('get_all_oidc'))
        @limiter.limit(get_limit_rate('get_all_limit') or "")
        @namespace.marshal_with(get_model, as_list=True)
        def get(self):
            all_entities = entity.get_all()
            sort_by = kwargs.get('sort_by', None)
            sorting = kwargs.get('sorting', 'asc')

            if len(all_entities) == 0:
                return all_entities, HTTPStatus.OK

            if sort_by is not None and sort_by in all_entities[0].__dict__.keys():
                all_entities = sorted(all_entities, key=lambda x: getattr(x, sort_by), reverse=(sorting=='desc'))
            elif sort_by is not None: # pragma: no cover
                _log('warning', f'Cannot sort by {sort_by}, attribute not found in entity {entity.__name__}')

            return all_entities, HTTPStatus.OK

    class DynamicResourceCreate(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('post_oidc'))
        @limiter.limit(get_limit_rate('post_limit') or "")
        @namespace.expect(upsert_model)
        @namespace.marshal_with(get_model)
        def post(self):
            try:
                data = entity.create(**namespace.payload)
                callback('on_post', data=data)
                
                return data, HTTPStatus.CREATED
            except Exception as e:
                _log('error', f'Error creating {entity.__name__}: {e}')
                self.error_response(e.args)

    class DynamicResourceGetById(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('get_by_id_oidc'))
        @limiter.limit(get_limit_rate('get_by_id_limit') or "")
        @namespace.marshal_with(get_model)
        def get(self, id):
            try:
                return entity.get_by_id(id), HTTPStatus.OK
            except Exception as e:
                _log('error', f'Error getting {entity.__name__} by id {id}: {e}')
                self.error_response(e.args)

    class DynamicResourceEditById(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('put_oidc'))
        @limiter.limit(get_limit_rate('put_limit') or "")
        @namespace.expect(upsert_model)
        @namespace.marshal_with(get_model)
        def put(self, id):
            try:
                data = entity.update(id, **namespace.payload)
                callback('on_put', data=data)

                return data, HTTPStatus.OK
            except Exception as e:
                _log('error', f'Error updating {entity.__name__} id {id}: {e}')
                self.error_response(e.args)

    class DynamicResourceDeleteById(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('delete_all_oidc'))
        @limiter.limit(get_limit_rate('delete_limit') or "")
        def delete(self, id):
            try:
                entity.delete(id)
                callback('on_delete')
                return '', HTTPStatus.NO_CONTENT
            except Exception as e:
                _log('error', f'Error deleting {entity.__name__} id {id}: {e}')
                self.error_response(e.args)

    def identity_decorator(func):
        return func

    file_upload_parser: RequestParser | None = None
    if kwargs.get('file_import', False) or kwargs.get('file_preview', False):
        file_upload_parser = namespace.parser()
        file_upload_parser.add_argument(
            kwargs.get('file_field', 'file'),
            type=FileStorage,
            location='files',
            required=True,
            help='Excel (.xlsx) file'
        )

    file_expect = namespace.expect(file_upload_parser) if file_upload_parser is not None else identity_decorator

    class DynamicResourceFileExport(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('file_export_oidc'))
        @limiter.limit(get_limit_rate('file_export_limit') or "")
        def get(self):
            try:
                entities = entity.get_all()
                data = [_to_serializable_dict(item) for item in entities]
                headers = kwargs.get('file_headers')
                if headers is None:
                    headers = _collect_headers_from_records(data)
                rows = [[record.get(header) for header in headers] for record in data]
                workbook_stream = _build_workbook(headers, rows)
                filename = kwargs.get('file_name', f"{entity.__name__}.xlsx")
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                return send_file(
                    workbook_stream,
                    mimetype=mimetype,
                    as_attachment=True,
                    download_name=filename,
                )
            except Exception as e:
                _log('error', f'Error exporting {entity.__name__} to file: {e}')
                self.error_response(e.args)

    class DynamicResourceFileImport(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('file_import_oidc'))
        @limiter.limit(get_limit_rate('file_import_limit') or "")
        @file_expect
        def post(self):
            upload_field = kwargs.get('file_field', 'file')
            uploaded_file = request.files.get(upload_field)

            if uploaded_file is None or uploaded_file.filename == '':
                return {'message': 'Missing file upload'}, HTTPStatus.BAD_REQUEST

            try:
                file_stream = BytesIO(uploaded_file.read())
                file_stream.seek(0)
                headers, records = _read_workbook_records(file_stream, kwargs.get('file_sheet'))

                unique_fields = kwargs.get('file_unique_fields')
                if unique_fields is None: # pragma: no cover
                    unique_fields = ['id'] if 'id' in headers else []

                created = 0
                updated = 0

                existing_index: dict[tuple[Any, ...], dict[str, Any]] = {}
                if unique_fields:
                    existing_entities = entity.get_all()
                    existing_records = []
                    for item in existing_entities:
                        mapped = _to_serializable_dict(item)
                        if isinstance(mapped, dict):
                            existing_records.append(mapped)
                    for record in existing_records:
                        if not hasattr(record, 'get'): # pragma: no cover
                            continue
                        key = tuple(record.get(field) for field in unique_fields)
                        if None not in key:
                            existing_index[key] = record

                for record in records:
                    if unique_fields:
                        key = tuple(record.get(field) for field in unique_fields)
                        if None not in key and key in existing_index:
                            existing = existing_index[key]
                            if hasattr(existing, 'get'):
                                target_id = existing.get('id') or record.get('id')
                            else: # pragma: no cover
                                target_id = record.get('id')
                            if target_id is not None:
                                entity.update(target_id, **record)
                                updated += 1
                                continue
                    entity.create(**record)
                    created += 1

                return {'created': created, 'updated': updated}, HTTPStatus.OK
            except Exception as e: # pragma: no cover
                _log('error', f'Error importing {entity.__name__} from file: {e}')
                self.error_response(e.args)

    class DynamicResourceFilePreview(Resource, ErrorResponse):
        @custom_wrapper
        @verify_oidc(get_oidc('file_preview_oidc'))
        @limiter.limit(get_limit_rate('file_preview_limit') or "")
        @file_expect
        def post(self):
            upload_field = kwargs.get('file_field', 'file')
            uploaded_file = request.files.get(upload_field)

            if uploaded_file is None or uploaded_file.filename == '':
                return {'message': 'Missing file upload'}, HTTPStatus.BAD_REQUEST

            try:
                file_stream = BytesIO(uploaded_file.read())
                file_stream.seek(0)
                headers, records = _read_workbook_records(file_stream, kwargs.get('file_sheet'))
                return {'headers': headers, 'rows': records}, HTTPStatus.OK
            except Exception as e: # pragma: no cover
                _log('error', f'Error generating preview for {entity.__name__}: {e}')
                self.error_response(e.args)

    get_all = kwargs.get('get_all', True)
    post = kwargs.get('post', True)

    get_by_id = kwargs.get('get_by_id', True)
    put = kwargs.get('put', True)
    delete = kwargs.get('delete', True)

    all = []
    if get_all: all.append(DynamicResourceGetAll)
    if post: all.append(DynamicResourceCreate)

    if len(all) > 0:
        DynamicResource = build_dynamic_class('DynamicResource', all)
        namespace.add_resource(DynamicResource, '/')

    by_id = []
    if get_by_id: by_id.append(DynamicResourceGetById)
    if put: by_id.append(DynamicResourceEditById)
    if delete: by_id.append(DynamicResourceDeleteById)

    if len(by_id) > 0:
        DynamicResourceById = build_dynamic_class('DynamicResourceById', by_id)
        namespace.add_resource(DynamicResourceById, '/<int:id>/')

    file_resources = []
    if kwargs.get('file_export', False):
        file_resources.append(DynamicResourceFileExport)
    if kwargs.get('file_import', False):
        file_resources.append(DynamicResourceFileImport)

    if len(file_resources) > 0:
        DynamicResourceFile = build_dynamic_class('DynamicResourceFile', file_resources)
        namespace.add_resource(DynamicResourceFile, '/file/')

    if kwargs.get('file_preview', False):
        namespace.add_resource(DynamicResourceFilePreview, '/file/preview/')
