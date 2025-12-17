import io
from unittest import TestCase
from unittest.mock import patch, MagicMock
from eaasy.extensions import build_model, build_resource, build_dynamic_class
from flask_restx import Api, fields, Namespace, Model
from flask_restx.reqparse import RequestParser
from eaasy.domain.database import BaseEntity
from flask import Flask
from openpyxl import load_workbook, Workbook

class MockProperty:
    def __init__(self, name, type, nullable):
        self.__dict__["name"] = name
        self.__dict__["type"] = type
        self.__dict__["nullable"] = nullable

class TestEntity:
    __name__ = "TestEntity"
    __dict__ = {
        "id": MockProperty("id", "INTEGER", False),
        "name": MockProperty("name", "VARCHAR", True),
        "email": MockProperty("email", "VARCHAR", False),
        "height": MockProperty("height", "NUMERIC", True),
        "weight": MockProperty("weight", "NUMERIC", False),
        "created_at": MockProperty("created_at", "DATETIME", False),
        "deleted_at": MockProperty("deleted_at", "DATETIME", True),
        "is_active": MockProperty("is_active", "BOOLEAN", True),
        "unsupported_field": MockProperty("unsupported_field", "UNSUPPORTED", True)
    }

class TestBaseEntity(BaseEntity):
    __name__ = "TestBaseEntity"
    __tablename__ = "test_base_entity"
    def __init__(self, **kwargs):
        # Set the attributes as class properties
        for key, value in kwargs.items():
            setattr(self, key, value)

@patch("builtins.print")
class TestBuilder(TestCase):

    ### Model building tests

    def test_building_model_returns_expected_namespace(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        namespace, _ = build_model(test_class)

        # Assert
        self.assertIsInstance(namespace, Namespace)
        self.assertEqual(namespace.name, test_class.__name__)
        self.assertEqual(namespace.description, "TestEntity operations")
        self.assertEqual(namespace.path, "/testentity")

    def test_building_model_returns_expected_properties(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        self.assertIsInstance(model, Model)
        self.assertIn("id", model)
        self.assertIn("name", model)
        self.assertIn("email", model)
        self.assertIn("height", model)
        self.assertIn("created_at", model)
        self.assertIn("is_active", model)

    def test_building_model_returns_expected_integer(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        id_field = model["id"]
        self.assertIsInstance(id_field, fields.Integer)
        self.assertTrue(id_field.required)
        self.assertEqual(id_field.default, 0)
        self.assertEqual(id_field.example, 0)

    def test_building_model_returns_expected_nullable_float(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        weight_field = model["height"]
        self.assertIsInstance(weight_field, fields.Float)
        self.assertFalse(weight_field.required)
        self.assertIsNone(weight_field.default)

    def test_building_model_returns_expected_required_float(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        height_field = model["weight"]
        self.assertIsInstance(height_field, fields.Float)
        self.assertTrue(height_field.required)
        self.assertEqual(height_field.default, 0.0)
        self.assertEqual(height_field.example, 0.0)

    def test_building_model_returns_expected_nullable_string(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        name_field = model["name"]
        self.assertIsInstance(name_field, fields.String)
        self.assertFalse(name_field.required)
        self.assertIsNone(name_field.default)

    def test_building_model_returns_expected_required_string(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        email_field = model["email"]
        self.assertIsInstance(email_field, fields.String)
        self.assertTrue(email_field.required)
        self.assertEqual(email_field.default, "")

    def test_building_model_returns_expected_nullable_datetime(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        created_at_field = model["deleted_at"]
        self.assertIsInstance(created_at_field, fields.DateTime)
        self.assertFalse(created_at_field.required)
        self.assertIsNone(created_at_field.default)

    def test_building_model_returns_expected_required_datetime(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        deleted_at_field = model["created_at"]
        self.assertIsInstance(deleted_at_field, fields.DateTime)
        self.assertTrue(deleted_at_field.required)
        self.assertIsNotNone(deleted_at_field.default)

    def test_building_model_returns_expected_boolean(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        is_active_field = model["is_active"]
        self.assertIsInstance(is_active_field, fields.Boolean)
        self.assertFalse(is_active_field.required)
        self.assertEqual(is_active_field.default, False)

    def test_building_model_logs_unsupported_types(self, *_):
        # Arrange
        test_class = TestEntity()
        
        # Act
        _, model = build_model(test_class)

        # Assert
        self.assertNotIn("unsupported_field", model)

    ### Resource building tests

    def test_building_resource_returns_namespace_with_no_resources(self, *_):
        # Arrange
        test_class = TestBaseEntity()
        namespace, _ = build_model(test_class)
        resources = len(namespace.resources)
        
        # Assert
        self.assertEqual(0, resources)

    def test_building_resource_returns_namespace_with_get_resource(self, *_):
        # Arrange
        test_class = TestBaseEntity()
        namespace, model = build_model(test_class)
        
        # Act
        build_resource(test_class, namespace, model, model)
        
        # Assert
        found_resource = [x for x in namespace.resources if x.urls[0] == "/"]
        self.assertEqual(1, len(found_resource))

    def test_building_resource_returns_namespace_with_get_by_id_resource(self, *_):
        # Arrange
        test_class = TestBaseEntity()
        namespace, model = build_model(test_class)
        
        # Act
        build_resource(test_class, namespace, model, model)
        
        # Assert
        found_resource = [x for x in namespace.resources if x.urls[0] == "/<int:id>/"]
        self.assertEqual(1, len(found_resource))

    def test_building_resource_with_wrapper_applies_wrapper_to_methods(self, *_):
        # Arrange
        test_class = TestBaseEntity()
        namespace, model = build_model(test_class)
        
        # Act
        build_resource(test_class, namespace, model, model, custom_wrapper=lambda f: (lambda *a, **k: f(*a, **k)))
        
        # Assert
        get_all_resource = [x for x in namespace.resources if x.urls[0] == "/"][0]
        get_by_id_resource = [x for x in namespace.resources if x.urls[0] == "/<int:id>/"][0]
        self.assertTrue(hasattr(get_all_resource.resource, "get"))
        self.assertTrue(hasattr(get_by_id_resource.resource, "get"))
        self.assertTrue(callable(get_all_resource.resource.get))
        self.assertTrue(callable(get_by_id_resource.resource.get))

    def setup_namespace(self, name: str):
        app = Flask(__name__)
        app.testing = True
        app.config["TESTING"] = True
        api = Api(app)
        namespace = Namespace(name)
        api.add_namespace(namespace)

        mock_entity = MagicMock()
        mock_entity.get_all.return_value = []
        mock_entity.get_by_id.return_value = {}
        mock_entity.create.return_value = {}
        mock_entity.update.return_value = {}
        mock_entity.__name__ = "TestBaseEntity"

        return app, namespace, mock_entity

    def build_workbook(self, rows: list[list[object]]):
        workbook = Workbook()
        worksheet = workbook.active
        for row in rows:
            worksheet.append(row) # type: ignore
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream
    
    def test_getting_all_returns_not_found_when_resource_is_not_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, *_ = self.setup_namespace(ns_name)
        
        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/')

        # Assert
        self.assertEqual(404, response.status_code)

    def test_exporting_returns_excel_file_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = [
            TestBaseEntity(id=1, name="Alpha", email="alpha@example.com"),
            TestBaseEntity(id=2, name="Beta", email="beta@example.com"),
        ]
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_export=True,
            file_headers=['id', 'name', 'email'],
        )

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/file/')

        # Assert
        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response.headers['Content-Type']
        )
        self.assertIn('attachment; filename=TestBaseEntity.xlsx', response.headers['Content-Disposition'])

        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook.active

        header_values = [cell.value for cell in sheet[1]] # type: ignore
        first_row_values = [cell.value for cell in sheet[2]] # type: ignore
        second_row_values = [cell.value for cell in sheet[3]] # type: ignore

        self.assertEqual(['id', 'name', 'email'], header_values)
        self.assertEqual([1, 'Alpha', 'alpha@example.com'], first_row_values)
        self.assertEqual([2, 'Beta', 'beta@example.com'], second_row_values)

    def test_importing_creates_new_rows_when_unique_not_found(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = []
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_import=True,
            file_unique_fields=['email'],
        )
        workbook_stream = self.build_workbook([
            ['email', 'name'],
            ['alpha@example.com', 'Alpha'],
            ['beta@example.com', 'Beta'],
        ])

        # Act
        with app.test_client() as client:
            response = client.post(
                f'/{ns_name}/file/',
                data={'file': (workbook_stream, 'users.xlsx')},
                content_type='multipart/form-data'
            )

        # Assert
        self.assertEqual(200, response.status_code)
        self.assertEqual(2, mock_entity.create.call_count)
        mock_entity.update.assert_not_called()
        self.assertEqual({'created': 2, 'updated': 0}, response.json) # type: ignore

    def test_importing_updates_existing_rows_based_on_unique_fields(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = [
            TestBaseEntity(id=1, email='alpha@example.com', name='Alpha'),
        ]
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_import=True,
            file_unique_fields=['email'],
        )
        workbook_stream = self.build_workbook([
            ['email', 'name'],
            ['alpha@example.com', 'Alpha Updated'],
            ['gamma@example.com', 'Gamma'],
        ])

        # Act
        with app.test_client() as client:
            response = client.post(
                f'/{ns_name}/file/',
                data={'file': (workbook_stream, 'users.xlsx')},
                content_type='multipart/form-data'
            )

        # Assert
        self.assertEqual(200, response.status_code)
        mock_entity.update.assert_called_once_with(1, email='alpha@example.com', name='Alpha Updated')
        mock_entity.create.assert_called_once_with(email='gamma@example.com', name='Gamma')
        self.assertEqual({'created': 1, 'updated': 1}, response.json) # type: ignore

    def test_importing_returns_bad_request_when_file_missing(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_import=True,
        )

        # Act
        with app.test_client() as client:
            response = client.post(
                f'/{ns_name}/file/',
                data={},
                content_type='multipart/form-data'
            )

        # Assert
        self.assertEqual(400, response.status_code)
        self.assertEqual('Missing file upload', response.json['message']) # type: ignore

    def test_importing_registers_file_upload_parser(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_import=True,
        )

        file_resource = [r for r in namespace.resources if r.urls[0] == '/file/'][0]
        post_doc = file_resource.resource.post.__apidoc__

        self.assertIn('expect', post_doc)
        self.assertIsInstance(post_doc['expect'][0], RequestParser)
        parser = post_doc['expect'][0]
        self.assertEqual('file', parser.args[0].name)
        self.assertEqual('files', parser.args[0].location)

    def test_preview_returns_table_rows(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_preview=True,
        )
        workbook_stream = self.build_workbook([
            ['email', 'name'],
            ['alpha@example.com', 'Alpha'],
            ['beta@example.com', 'Beta'],
        ])

        # Act
        with app.test_client() as client:
            response = client.post(
                f'/{ns_name}/file/preview/',
                data={'file': (workbook_stream, 'users.xlsx')},
                content_type='multipart/form-data'
            )

        # Assert
        self.assertEqual(200, response.status_code)
        self.assertEqual({
            'headers': ['email', 'name'],
            'rows': [
                {'email': 'alpha@example.com', 'name': 'Alpha'},
                {'email': 'beta@example.com', 'name': 'Beta'},
            ]
        }, response.json) # type: ignore

    def test_preview_returns_bad_request_when_file_missing(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_preview=True,
        )

        # Act
        with app.test_client() as client:
            response = client.post(
                f'/{ns_name}/file/preview/',
                data={},
                content_type='multipart/form-data'
            )

        # Assert
        self.assertEqual(400, response.status_code)
        self.assertEqual('Missing file upload', response.json['message']) # type: ignore

    def test_exporting_without_headers_infers_from_entities(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = [
            {'id': 1, 'name': 'Alpha'},
            {'id': 2, 'name': 'Beta', 'extra': 'value'},
        ]
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_export=True,
            file_name='custom.xlsx'
        )

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/file/')

        # Assert
        self.assertEqual(200, response.status_code)
        self.assertIn('attachment; filename=custom.xlsx', response.headers['Content-Disposition'])

        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook.active

        header_values = [cell.value for cell in sheet[1]] # type: ignore
        first_row_values = [cell.value for cell in sheet[2]] # type: ignore
        second_row_values = [cell.value for cell in sheet[3]] # type: ignore

        self.assertEqual(['id', 'name', 'extra'], header_values)
        self.assertEqual([1, 'Alpha', None], first_row_values)
        self.assertEqual([2, 'Beta', 'value'], second_row_values)

    def test_exporting_with_no_results_returns_empty_sheet(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = []
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_export=True,
        )

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/file/')

        # Assert
        self.assertEqual(200, response.status_code)
        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook.active

        self.assertEqual(1, sheet.max_row) # type: ignore
        self.assertEqual(1, sheet.max_column) # type: ignore
        self.assertIsNone(sheet.cell(row=1, column=1).value) # type: ignore

    def test_exporting_returns_error_when_entity_raises_exception(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.side_effect = Exception({
            'status_code': 500,
            'message': 'Internal Error',
            'data': {},
        })
        build_resource(
            mock_entity,
            namespace,
            Model(ns_name),
            Model(ns_name),
            file_export=True,
        )

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/file/')

        # Assert
        self.assertEqual(500, response.status_code)
        self.assertEqual('Internal Error', response.json['message']) # type: ignore

    def test_getting_all_returns_ok_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/')

        # Assert
        self.assertEqual(200, response.status_code)

    def test_getting_all_data_sorted_by_key_returns_ok(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        mock_entity.get_all.return_value = [
            TestBaseEntity(id=2, name="B"),
            TestBaseEntity(id=1, name="A"),
            TestBaseEntity(id=3, name="C"),
        ]
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), sort_by="id")

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/')

        # Assert
        self.assertEqual(200, response.status_code)
        self.assertEqual(3, len(response.json)) # type: ignore

    def test_posting_returns_not_found_when_resource_is_not_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, *_ = self.setup_namespace(ns_name)
        
        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/')

        # Assert
        self.assertEqual(404, response.status_code)

    def test_posting_returns_created_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(201, response.status_code)
        mock_entity.create.assert_called_once_with(test_prop='test_value')

    def test_posting_raises_error_when_entity_raises_exception(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        headers={"Content-Type": "application/json"}
        mock_entity.create.side_effect = Exception({
            'status_code': 400,
            'message': 'Bad request',
            'data': {},
        })

        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/', headers=headers, json={})

        # Assert
        self.assertEqual(400, response.status_code)
        self.assertEqual('Bad request', response.json['message']) # type: ignore

    def test_getting_by_id_returns_not_found_when_resource_is_not_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, *_ = self.setup_namespace(ns_name)
        
        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/1')

        # Assert
        self.assertEqual(404, response.status_code)

    def test_getting_by_id_returns_ok_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(1, mock_entity.get_by_id.call_count)
        self.assertEqual(200, response.status_code)

    def test_getting_by_id_raises_error_when_entity_raises_exception(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        mock_entity.get_by_id.side_effect = Exception({
            'status_code': 400,
            'message': 'Bad request',
            'data': {},
        })

        # Act
        with app.test_client() as client:
            response = client.get(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(1, mock_entity.get_by_id.call_count)
        self.assertEqual(400, response.status_code)
        self.assertEqual('Bad request', response.json['message']) # type: ignore

    def test_putting_returns_not_found_when_resource_is_not_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, *_ = self.setup_namespace(ns_name)
        
        # Act
        with app.test_client() as client:
            response = client.put(f'/{ns_name}/1')

        # Assert
        self.assertEqual(404, response.status_code)

    def test_putting_returns_ok_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.put(f'/{ns_name}/1/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(1, mock_entity.update.call_count)
        self.assertEqual(200, response.status_code)

    def test_putting_raises_error_when_entity_raises_exception(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        headers={"Content-Type": "application/json"}
        mock_entity.update.side_effect = Exception({
            'status_code': 400,
            'message': 'Bad request',
            'data': {},
        })

        # Act
        with app.test_client() as client:
            response = client.put(f'/{ns_name}/1/', headers=headers, json={})

        # Assert
        self.assertEqual(1, mock_entity.update.call_count)
        self.assertEqual(400, response.status_code)
        self.assertEqual('Bad request', response.json['message']) # type: ignore

    def test_deleting_returns_not_found_when_resource_is_not_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, *_ = self.setup_namespace(ns_name)
        
        # Act
        with app.test_client() as client:
            response = client.delete(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(404, response.status_code)

    def test_deleting_returns_no_content_when_resource_is_defined(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))

        # Act
        with app.test_client() as client:
            response = client.delete(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(1, mock_entity.delete.call_count)
        self.assertEqual(204, response.status_code)

    def test_deleting_raises_error_when_entity_raises_exception(self, *_):
        # Arrange
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name))
        mock_entity.delete.side_effect = Exception({
            'status_code': 400,
            'message': 'Bad request',
            'data': {},
        })

        # Act
        with app.test_client() as client:
            response = client.delete(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(1, mock_entity.delete.call_count)
        self.assertEqual(400, response.status_code)
        self.assertEqual('Bad request', response.json['message']) # type: ignore

    def test_calling_event_on_post_when_defined(self, *_):
        # Arrange
        event = MagicMock()
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), on_post=event)
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(1, mock_entity.create.call_count)
        event.assert_called_once_with(mock_entity.create.return_value)
        self.assertEqual(201, response.status_code)

    def test_calling_event_on_put_when_defined(self, *_):
        # Arrange
        event = MagicMock()
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), on_put=event)
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.put(f'/{ns_name}/1/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(1, mock_entity.update.call_count)
        event.assert_called_once_with(mock_entity.update.return_value)
        self.assertEqual(200, response.status_code)

    def test_calling_event_on_delete_when_defined(self, *_):
        # Arrange
        event = MagicMock()
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), on_delete=event)

        # Act
        with app.test_client() as client:
            response = client.delete(f'/{ns_name}/1/')

        # Assert
        self.assertEqual(1, mock_entity.delete.call_count)
        event.assert_called_once()
        self.assertEqual(204, response.status_code)

    def test_calling_ruinous_event_on_post_raises_when_fallback_is_not_defined(self, *_):
        # Arrange
        event = MagicMock(side_effect=Exception)
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), on_post=event)
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(1, mock_entity.create.call_count)
        self.assertEqual(500, response.status_code)

    def test_calling_ruinous_event_on_post_calls_fallback_when_defined(self, *_):
        # Arrange
        event = MagicMock(side_effect=Exception)
        fallback = MagicMock()
        ns_name = 'testentity'
        app, namespace, mock_entity = self.setup_namespace(ns_name)
        build_resource(mock_entity, namespace, Model(ns_name), Model(ns_name), on_post=event, on_post_fail=fallback)
        headers={"Content-Type": "application/json"}

        # Act
        with app.test_client() as client:
            response = client.post(f'/{ns_name}/', headers=headers, json={'test_prop': 'test_value'})

        # Assert
        self.assertEqual(1, mock_entity.create.call_count)
        event.assert_called_once()
        fallback.assert_called_once()
        self.assertEqual(201, response.status_code)

    def test_building_dynamic_class_returns_expected_type(self, *_):
        # Arrange
        class TypeA:
            prop_a = "A"
        class TypeB:
            prop_b = "B"
        class TypeC:
            prop_c = "C"
        name = "DynamicResource"
        base_classes = [TypeB, TypeC]

        # Act
        result = build_dynamic_class(name, base_classes)

        # Assert
        self.assertIsInstance(result, type)
        self.assertEqual(result.__name__, name)
        self.assertEqual(result.__bases__, tuple(base_classes))
        self.assertNotIn("prop_a", result.__dict__)
        self.assertEqual(result.prop_b, "B")
        self.assertEqual(result.prop_c, "C")
