#!/usr/bin/env python3
"""
commit_report.py

Generate an Excel report comparing commits across multiple Git branches.

Features:
- Reads branch commit history using GitPython.
- Produces an .xlsx file with four worksheets:
    1. timeline_by_date      (chronological commit timeline)
    2. timeline_by_topology  (Git DAG / topo-order timeline)
    3. timeline_hybrid       (both date + topo indices, out-of-order flags)
    4. analytics_summary     (counts, unique/missing commits per branch, etc.)

- Branches are provided via:
    --branches       (comma-separated)
    --branches-file  (newline-separated list)

- Sheet names and colors are configurable via YAML:
    src/jps_release_audit_utils/conf/config.yaml

Default sheet names (overridable by YAML):
    timeline_by_date
    timeline_by_topology
    timeline_hybrid
    analytics_summary

Default colors (overridable by YAML):
    missing:      FFC7CE  (light red)
    all_present:  C6EFCE  (light green)
    out_of_order: FFEB9C  (light yellow)

Color semantics:
- Branch cells with "MISSING" => missing color.
- Branch cells when commit is in *all* branches => all_present color.
- Out-of-order commits (date earlier than previous topo commit's date):
    - timeline_by_topology: Date cell highlighted with out_of_order color.
    - timeline_hybrid: OutOfOrder cell highlighted with out_of_order color.

All sheets:
- Header row frozen (A2)
- Auto-filter enabled on header row
- Columns auto-sized

Usage (example):

    python -m jps_release_audit_utils.commit_report \
        --repo-path /path/to/repo \
        --branches develop,main,release/v5.8.0-rc \
        --output commit_report.xlsx

Or:

    python -m jps_release_audit_utils.commit_report \
        --repo-path /path/to/repo \
        --branches-file branches.txt \
        --output commit_report.xlsx

"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from datetime import datetime

import typer
from git import Repo
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import yaml


# ============================================================
# DEFAULT CONFIG
# ============================================================

# Default config path: src/jps_release_audit_utils/conf/config.yaml
DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent / "conf" / "config.yaml"

DEFAULT_SHEET_NAMES = {
    "timeline_by_date": "timeline_by_date",
    "timeline_by_topology": "timeline_by_topology",
    "timeline_hybrid": "timeline_hybrid",
    "analytics_summary": "analytics_summary",
}

DEFAULT_COLORS = {
    "missing": "FFC7CE",      # light red
    "all_present": "C6EFCE",  # light green
    "out_of_order": "FFEB9C", # light yellow
}


# ============================================================
# LOGGING
# ============================================================

logger = logging.getLogger("commit-report")


def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    logger.debug("Verbose logging enabled.")


# ============================================================
# CLI
# ============================================================

app = typer.Typer(add_completion=False)


# ============================================================
# CONFIG LOADER
# ============================================================

def load_config(
    config_path: Optional[Path],
    required: bool = False,
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Load sheet names and colors from a YAML config file.

    YAML example:

    sheets:
      timeline_by_date: "Timeline (by Date)"
      timeline_by_topology: "Timeline (Topology)"
      timeline_hybrid: "Timeline (Hybrid)"
      analytics_summary: "Analytics"

    colors:
      missing: "FFC7CE"
      all_present: "C6EFCE"
      out_of_order: "FFEB9C"
    """
    sheet_names = DEFAULT_SHEET_NAMES.copy()
    colors = DEFAULT_COLORS.copy()

    if config_path is None:
        logger.info("No config path supplied; using built-in defaults.")
        return sheet_names, colors

    if not config_path.is_file():
        if required:
            raise typer.Exit(f"ERROR: Config file not found: {config_path}")
        logger.info("Config file not found at %s; using built-in defaults.", config_path)
        return sheet_names, colors

    logger.info("Loading configuration from %s", config_path)

    with config_path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}

    sheets_cfg = data.get("sheets", {})
    for key, default_name in DEFAULT_SHEET_NAMES.items():
        if key in sheets_cfg and sheets_cfg[key]:
            sheet_names[key] = str(sheets_cfg[key])

    colors_cfg = data.get("colors", {})
    for key, default_color in DEFAULT_COLORS.items():
        if key in colors_cfg and colors_cfg[key]:
            colors[key] = str(colors_cfg[key])

    logger.debug("Effective sheet names: %s", sheet_names)
    logger.debug("Effective colors: %s", colors)

    return sheet_names, colors


# ============================================================
# HELPERS
# ============================================================

PR_REGEX = re.compile(r"#(\d+)")


def extract_pr_number(message: str) -> str:
    """Extract a PR number (#1234) from the commit message, if present."""
    match = PR_REGEX.search(message)
    return match.group(1) if match else ""


def load_branches(branches: Optional[str], branches_file: Optional[str]) -> List[str]:
    """
    Load branch list from:
        - --branches (comma-separated), or
        - --branches-file (newline-separated)

    Priority:
        1. branches_file
        2. branches
    """
    if branches_file:
        fp = Path(branches_file)
        if not fp.is_file():
            raise typer.Exit(f"ERROR: Branches file not found: {branches_file}")
        branch_list = [line.strip() for line in fp.read_text().splitlines() if line.strip()]
        logger.info("Loaded %d branches from file.", len(branch_list))
        return branch_list

    if branches:
        branch_list = [b.strip() for b in branches.split(",") if b.strip()]
        logger.info("Loaded branches from CLI: %s", branch_list)
        return branch_list

    raise typer.Exit("ERROR: Either --branches or --branches-file must be provided.")


# ============================================================
# GIT OPERATIONS
# ============================================================

def get_branch_commits(repo: Repo, branch_name: str) -> Dict[str, dict]:
    """
    Return commit_hash -> metadata dict for a single branch:

        {
            "datetime": datetime,
            "date": "YYYY-MM-DD",
            "message": str,
            "author": "Name <email>",
            "pr": str,
        }
    """
    logger.info("Loading commits for branch: %s", branch_name)

    if branch_name not in repo.branches:
        raise typer.Exit(f"ERROR: Branch '{branch_name}' does not exist locally.")

    commit_map: Dict[str, dict] = {}

    for commit in repo.iter_commits(branch_name):
        dt = datetime.fromtimestamp(commit.authored_date)
        date_str = dt.strftime("%Y-%m-%d")
        msg = commit.message.strip().replace("\n", " ")
        author = f"{commit.author.name} <{commit.author.email}>"
        pr_num = extract_pr_number(msg)

        commit_map[commit.hexsha] = {
            "datetime": dt,
            "date": date_str,
            "message": msg,
            "author": author,
            "pr": pr_num,
        }

    logger.info("  Found %d commits in %s", len(commit_map), branch_name)
    return commit_map


def get_topo_sorted_commits(repo: Repo, branches: List[str]) -> List[str]:
    """
    Use git rev-list --topo-order <branches...> to get commit hashes
    in true Git ancestry order.
    """
    logger.info("Computing topological order across branches: %s", branches)
    rev_output = repo.git.rev_list("--topo-order", *branches)
    hexshas = [line.strip() for line in rev_output.splitlines() if line.strip()]
    logger.info("Topological commit count: %d", len(hexshas))
    return hexshas


# ============================================================
# EXCEL HELPERS
# ============================================================

def autosize_columns(sheet) -> None:
    """Auto-adjust column widths based on content length."""
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 80)


def apply_frozen_and_filter(sheet) -> None:
    """Freeze header row and enable auto-filter."""
    sheet.freeze_panes = "A2"
    # Auto-filter over the full used range
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def write_sheet_by_date(
    wb: Workbook,
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 1: timeline_by_date
    Sorted by commit datetime ascending.
    Columns:
        Date, DateIndex, TopoIndex, CommitHash, Message, Author, PR_Number, <branches...>
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "Date",
        "DateIndex",
        "TopoIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
    ] + branches
    sheet.append(headers)

    # Sort commits by chronological date
    chronological = sorted(
        all_commits.items(),
        key=lambda kv: (kv[1]["datetime"], kv[0]),
    )

    for _, meta in chronological:
        # some commits might not have topo_index/date_index if something is odd; guard
        date_index = meta.get("date_index", "")
        topo_index = meta.get("topo_index", "")

        base_row = [
            meta["date"],
            date_index,
            topo_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
        ]

        # Branch cells + color logic
        row_values = list(base_row)
        missing_color = PatternFill(
            start_color=colors["missing"],
            end_color=colors["missing"],
            fill_type="solid",
        )
        all_present_color = PatternFill(
            start_color=colors["all_present"],
            end_color=colors["all_present"],
            fill_type="solid",
        )

        # branches start at column index:
        base_offset = len(base_row)  # 0-based for Python, 1-based in Excel
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append(meta["hash"])
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Apply colors to branch cells
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1  # 1-based
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_color
            elif all_present:
                cell.fill = all_present_color

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_by_topology(
    wb: Workbook,
    topo_commits: List[str],
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 2: timeline_by_topology
    Ordered by topo_index (Git ancestry).
    Columns:
        TopoIndex, Date, DateIndex, CommitHash, Message, Author, PR_Number, <branches...>

    out_of_order (date earlier than previous topo commit):
        Date cell highlighted with out_of_order color.
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "TopoIndex",
        "Date",
        "DateIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
    ] + branches
    sheet.append(headers)

    missing_fill = PatternFill(
        start_color=colors["missing"],
        end_color=colors["missing"],
        fill_type="solid",
    )
    all_present_fill = PatternFill(
        start_color=colors["all_present"],
        end_color=colors["all_present"],
        fill_type="solid",
    )
    out_of_order_fill = PatternFill(
        start_color=colors["out_of_order"],
        end_color=colors["out_of_order"],
        fill_type="solid",
    )

    for commit_hash in topo_commits:
        if commit_hash not in all_commits:
            continue
        meta = all_commits[commit_hash]

        topo_index = meta.get("topo_index", "")
        date_index = meta.get("date_index", "")

        base_row = [
            topo_index,
            meta["date"],
            date_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
        ]

        row_values = list(base_row)
        base_offset = len(base_row)
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append(meta["hash"])
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Apply branch cell colors
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_fill
            elif all_present:
                cell.fill = all_present_fill

        # Highlight out-of-order commits
        if meta.get("out_of_order"):
            # Date column is 2 (1-based)
            date_cell = sheet.cell(row=row_idx, column=2)
            date_cell.fill = out_of_order_fill

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_hybrid(
    wb: Workbook,
    topo_commits: List[str],
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 3: timeline_hybrid

    Also topo-order, but shows both topo_index and date_index,
    with explicit out-of-order flag and days-since-previous-topo.

    Columns:
        TopoIndex, Date, DateIndex, CommitHash, Message, Author, PR_Number,
        OutOfOrder, DaysSincePrevTopo, <branches...>
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "TopoIndex",
        "Date",
        "DateIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
        "OutOfOrder",
        "DaysSincePrevTopo",
    ] + branches
    sheet.append(headers)

    missing_fill = PatternFill(
        start_color=colors["missing"],
        end_color=colors["missing"],
        fill_type="solid",
    )
    all_present_fill = PatternFill(
        start_color=colors["all_present"],
        end_color=colors["all_present"],
        fill_type="solid",
    )
    out_of_order_fill = PatternFill(
        start_color=colors["out_of_order"],
        end_color=colors["out_of_order"],
        fill_type="solid",
    )

    prev_dt: Optional[datetime] = None

    for commit_hash in topo_commits:
        if commit_hash not in all_commits:
            continue
        meta = all_commits[commit_hash]

        topo_index = meta.get("topo_index", "")
        date_index = meta.get("date_index", "")

        dt = meta["datetime"]
        if prev_dt is None:
            days_since_prev = ""
        else:
            days_since_prev = (dt.date() - prev_dt.date()).days
        prev_dt = dt

        out_of_order_flag = "YES" if meta.get("out_of_order") else "NO"

        base_row = [
            topo_index,
            meta["date"],
            date_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
            out_of_order_flag,
            days_since_prev,
        ]

        row_values = list(base_row)
        base_offset = len(base_row)
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append(meta["hash"])
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Branch cell colors
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_fill
            elif all_present:
                cell.fill = all_present_fill

        # Highlight OutOfOrder cell if YES
        if meta.get("out_of_order"):
            # OutOfOrder column index is 8 (1-based)
            cell = sheet.cell(row=row_idx, column=8)
            cell.fill = out_of_order_fill

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_analytics(
    wb: Workbook,
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
) -> None:
    """
    Sheet 4: analytics_summary

    Simple summary metrics:
    - Total unique commits
    - Commits present in all branches
    - Commits with out_of_order flag
    - For each branch: total, unique-to-branch, missing-from-branch
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    # Basic metrics
    total_unique = len(all_commits)
    all_branches_set = set(branches)

    present_in_all = sum(
        1 for meta in all_commits.values()
        if meta["branches"] == all_branches_set
    )
    out_of_order_count = sum(
        1 for meta in all_commits.values()
        if meta.get("out_of_order")
    )

    sheet.append(["Metric", "Value"])
    sheet.append(["Total unique commits across all branches", total_unique])
    sheet.append(["Commits present in all branches", present_in_all])
    sheet.append(["Commits with out-of-order dates (vs topology)", out_of_order_count])
    sheet.append([])

    # Per-branch table
    sheet.append(["Branch", "TotalCommitsInBranch", "UniqueToBranch", "MissingFromBranch"])

    for branch in branches:
        total_in_branch = sum(
            1 for meta in all_commits.values()
            if branch in meta["branches"]
        )
        unique_to_branch = sum(
            1 for meta in all_commits.values()
            if meta["branches"] == {branch}
        )
        missing_from_branch = total_unique - total_in_branch

        sheet.append([branch, total_in_branch, unique_to_branch, missing_from_branch])

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


# ============================================================
# MAIN COMMAND
# ============================================================

@app.command()
def generate(
    repo_path: str = typer.Option(
        ".",
        help="Path to a local git repository (default: current working directory).",
    ),
    output: str = typer.Option(
        "commit_report.xlsx",
        help="Output Excel file path (default: commit_report.xlsx in CWD).",
    ),
    branches: Optional[str] = typer.Option(
        None,
        help="Comma-separated list of branches to inspect.",
    ),
    branches_file: Optional[str] = typer.Option(
        None,
        help="Path to file with newline-separated branches to inspect.",
    ),
    config: Optional[str] = typer.Option(
        None,
        help=(
            "Path to YAML configuration file. "
            "If not provided, defaults to conf/config.yaml within the package "
            "(if present), otherwise built-in defaults are used."
        ),
    ),
    verbose: bool = typer.Option(
        False,
        "--verbose",
        "-v",
        help="Enable verbose (DEBUG) logging.",
    ),
) -> None:
    """
    Generate an Excel report comparing commits across branches.
    """
    setup_logging(verbose)

    # Resolve config path
    if config is not None:
        config_path = Path(config).resolve()
        sheet_names, colors = load_config(config_path, required=True)
    else:
        sheet_names, colors = load_config(DEFAULT_CONFIG_PATH, required=False)

    branch_list = load_branches(branches, branches_file)
    logger.info("Using branches: %s", branch_list)

    repo = Repo(repo_path)
    if repo.bare:
        raise typer.Exit("ERROR: Provided path is not a valid git repository.")

    # Step 1: Collect commits per branch
    per_branch: Dict[str, Dict[str, dict]] = {}
    for branch in branch_list:
        per_branch[branch] = get_branch_commits(repo, branch)

    # Step 2: Merge into a unified all_commits map
    # all_commits[hash] = {
    #   "hash": ...,
    #   "datetime": ...,
    #   "date": ...,
    #   "message": ...,
    #   "author": ...,
    #   "pr": ...,
    #   "branches": set([...]),
    #   "topo_index": int,
    #   "date_index": int,
    #   "out_of_order": bool,
    # }
    all_commits: Dict[str, dict] = {}

    for branch, commit_map in per_branch.items():
        for commit_hash, meta in commit_map.items():
            if commit_hash not in all_commits:
                all_commits[commit_hash] = {
                    "hash": commit_hash,
                    "datetime": meta["datetime"],
                    "date": meta["date"],
                    "message": meta["message"],
                    "author": meta["author"],
                    "pr": meta["pr"],
                    "branches": set([branch]),
                }
            else:
                all_commits[commit_hash]["branches"].add(branch)

    logger.info("Total unique commits across all branches: %d", len(all_commits))

    # Step 3: Assign topo_index
    topo_commits = get_topo_sorted_commits(repo, branch_list)
    topo_idx = 1
    for commit_hash in topo_commits:
        if commit_hash in all_commits:
            all_commits[commit_hash]["topo_index"] = topo_idx
            topo_idx += 1

    # Step 4: Assign date_index based on chronological sort
    chronological = sorted(
        all_commits.items(),
        key=lambda kv: (kv[1]["datetime"], kv[0]),
    )
    for idx, (commit_hash, _) in enumerate(chronological, start=1):
        all_commits[commit_hash]["date_index"] = idx

    # Step 5: Determine out_of_order based on topo sequence
    prev_dt: Optional[datetime] = None
    for commit_hash in topo_commits:
        if commit_hash not in all_commits:
            continue
        meta = all_commits[commit_hash]
        dt = meta["datetime"]
        if prev_dt is None:
            meta["out_of_order"] = False
        else:
            meta["out_of_order"] = dt < prev_dt
        prev_dt = dt

    # Step 6: Write Excel workbook
    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    # Sheet 1: timeline_by_date
    write_sheet_by_date(
        wb=wb,
        all_commits=all_commits,
        branches=branch_list,
        sheet_name=sheet_names["timeline_by_date"],
        colors=colors,
    )

    # Sheet 2: timeline_by_topology
    write_sheet_by_topology(
        wb=wb,
        topo_commits=topo_commits,
        all_commits=all_commits,
        branches=branch_list,
        sheet_name=sheet_names["timeline_by_topology"],
        colors=colors,
    )

    # Sheet 3: timeline_hybrid
    write_sheet_hybrid(
        wb=wb,
        topo_commits=topo_commits,
        all_commits=all_commits,
        branches=branch_list,
        sheet_name=sheet_names["timeline_hybrid"],
        colors=colors,
    )

    # Sheet 4: analytics_summary
    write_sheet_analytics(
        wb=wb,
        all_commits=all_commits,
        branches=branch_list,
        sheet_name=sheet_names["analytics_summary"],
    )

    outpath = Path(output).resolve()
    wb.save(outpath)
    logger.info("Excel report written to: %s", outpath)


# ============================================================
if __name__ == "__main__":
    app()
