from loguru import logger
import openpyxl # kuradi kehv dokumentatsioon on asjal!!
import os

from dapu.agents.agent import AgentGeneric
from dapu.perks import split_task_id, calculate_file_content_hash, python_value_as_sql_string
from dapu.placeholder import Placeholder
from dapu.perks import excel_label_to_number, excel_number_to_label, sanityze, convert_to_iso_date, end_of_month, start_of_month, end_of_year, start_of_year
from dapu.textops import yaml_string_to_dict

class Agent(AgentGeneric):
    """
    - 
    do: exceltartu # custom agent.yaml may use word "excel" 
    source:
        name: "sample.xlsx" # Excel new format (openpyxl don't work with old format) 
        sheet: 1 # positsion of sheet in workbook (1 is first, humanly way)
        start_row: 2 # to eliminate header rows
    management: hash # delete
    delete: month # all, year, month, none, keys_file - do we need delete all data before loading, or periodic by range of excel existing dates 
    date_column: arve_algus_dt
    insert: simple # simple, key (key is not implemented)
    key:  #### not implemented yet, just idea... to allow more optimized key-based comparision instead of simple insert
        constraint: ak_harrastussport_4_uniq
        update_ts: dwh_updated_ts # võimalus träkkimiseks, veerg peab olema loodud
    deleted_keys: #### not implemented yet, just idea...
        format: excel # csv või text 
        file: "katse3deleted.xlsx"
        start_row: 2
        delimiter: "," # või ";" või "; " või " " -- csv jaoks
        cols: # list, kuna võti võib olla mitme-veeruline. vajame allikapoolset (st veerunumbrid excelis)
        - from: 3 # nt isikukood
        - from: 2 # nt arve number
    cols: # tulemtabeli veergude nimed ja kõik muu
    - name: isikukood # just text value
        from: a
    - name: isikukood_hashed
        from: a
        action: hash # some action before saving to database 
    - name: arve_algus # just text
        from: b
    - name: arve_algus_dt
        from: b
        action: to_date # transformed to ISO date string
    - name: summa_eur_dec # numeric(11,2)
        from: f
        action: to_dec # transformaed to numeric literal
    - name: dwh_imported_ts
        fixed: current_timestamp # insert ...(current_timestamp) -- as-is 
        # value: tere --> insert ..('tere') -- surround with apostrophes
    """
    myname: str = 'exceltartu'
    
    def do_action(self) -> bool:
        logger.info(f"start {self.myname} for {self.task_id}")
        route_code, schema_name, table_name = split_task_id(self.task_id)
        if 'source' not in self.action:
            logger.error(f'Missing source (excel file to import data from)');
            return False
        if 'cols' not in self.action:
            logger.error(f'Missing cols (listof columns of target table with instructions how them fullfil from excel)');
            return False
        profile: dict = self.context.get_profile(self.route_alias) or {}
        if not profile:
            logger.error(f"Missing profile {self.route_alias} (for task_id '{self.task_id}')")
            return False
        excel_dir = profile.get('path', None)
        if excel_dir is None:
            logger.error(f"Missing path in profile {self.route_alias} (for task_id '{self.task_id}')")
            return False
        if not os.path.exists(excel_dir):
            logger.error(f"Missing folder '{excel_dir}' (for task_id '{self.task_id}')")
            return False
        if 'name' not in self.action['source']:
            logger.error(f"Missing name of file (action>source>name) in loading definition (for task_id '{self.task_id}')")
            return False
        
        excel_file = os.path.join(excel_dir, self.action['source']['name']) # full name
        if not os.path.exists(excel_file):
            logger.info(f"No file '{excel_file}' (for task_id '{self.task_id}'), may be already consumed, end of pulling")
            return True # TRUE since no data since last loading, absolutelly fine 
        
        file_hash = ""
        file_management = self.action.get('management', 'hash') # hash (compare using last hash, do if different) vs delete (if exists then do)  
        if file_management == 'hash': # take last hash from registry 
            old_hash = self.context.get_task_hash_from_registry(self.task_id)
            file_hash = calculate_file_content_hash(excel_file)
            if old_hash == file_hash:
                return True # No need to load, ok
        
        delete_before = self.action.get('delete', 'month')
        delete_date_column = self.action.get('delete_date_column', None)
        sql_del = None
        periodic = False
        if delete_before == 'all':
            sql_del = f"""DELETE FROM {schema_name}.{table_name} WHERE TRUE"""
        if delete_before in ('day', 'month', 'year') and delete_date_column:
            periodic = True
        
        insert_mode = self.action.get('insert', 'simple')
        if insert_mode != 'simple':
            logger.error('Realiseerimata insert meetod, jääb katki')
            return False        
        
        sheet = self.get_sheet(excel_file, self.action['source'].get('sheet', 1)) # file and human understandable position of sheet (starts with 1)
        min_row, max_row = self.row_limits(sheet)
        if periodic: # if time-based deletion then we must read sheet once to find dates
            (period_start_dt, period_end_dt) = self.calculate_deletion_dates(sheet, min_row, max_row, delete_date_column)
            if period_start_dt is not None and period_end_dt is not None:
                sql_del = f"""DELETE FROM {schema_name}.{table_name} WHERE {delete_date_column} BETWEEN '{period_start_dt}' AND '{period_end_dt}' """

        if sql_del: # delete if is reason for that 
            self.context.target(sql_del, False) # FIXME: change logic -> load into temp table similar to target table, if succeeded than delete and insert
        self.load_sheet(sheet, min_row, max_row, schema_name, table_name, insert_mode)
        
        logger.debug(f'Fail imporditud {excel_file}')
        if file_management == 'hash':
            self.context.save_task_hash_to_registry(self.task_id, file_hash)
        else: # 'delete'
            # delete file
            os.remove(excel_file)
        logger.info(f"done {self.myname} for {self.task_id}")
        return True
             
    def load_sheet(self, sheet, min_row, max_row, schema_name, table_name, insert_mode):
        logger.debug(f'Laeme sisse read {min_row} .. {max_row}')
        rida = 0
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            # row on tuple tüüpi
            col_names = []
            col_values = []
            rida = rida + 1
            logger.debug(f"uus rida {rida}, analüüsime veerge def faili abil")
            for col in self.action['cols']:
                col_name = col['name'] # baasi tabeli veerunimi -- peab alati olema
                logger.debug(f"veerg: {col_name}")
                col_label_excel = col.get('from', None) # võib ka puududa? nt constant
                if col_label_excel:
                    logger.debug(f"veeru silt on: {col_label_excel}")
                    if not isinstance(col_label_excel, str):
                        logger.debug(type(col_label_excel))
                        col_label_excel = str(col_label_excel)
                    pos_excel = excel_label_to_number(col_label_excel) # 0-index
                    logger.debug(f"veeru indeks on: {pos_excel}")
                    logger.debug(f"väärtus on {row[pos_excel]}")
                    logger.debug(f"väärtuse tüüp on: {type(row[pos_excel])}")
                    col_value = sanityze(row[pos_excel], col.get('transform', None))
                    logger.debug(f"väärtus alguses: {col_value}")
                    # rega, et on olemas paar 
                    # (None -> NULL, väärtus -> 'väärtus') + ülakoma varjestus
                    if isinstance(col_value, str):
                        col_value = col_value.replace("'", "''")
                    #else:
                        #logger.debug('pole string')
                    col_values.append(python_value_as_sql_string(col_value))
                    #col_values.append('NULL' if col_value is None else f"'{col_value}'") # võiks veidi tüünida numbrite käitumist, aga PG saab ka nii hakkama
                    col_names.append(col_name)
            logger.debug('read läbi')        
            if insert_mode == 'simple':
                #sql_ins = f"""INSERT INTO {schema_name}.{table_name} ({', '.join(col_names)}) VALUES ({', '.join(map(str, col_values))})"""
                sql_ins = f"""INSERT INTO {schema_name}.{table_name} ({', '.join(col_names)}) VALUES ({', '.join(col_values)})"""
            else:
                sql_ins = None
            if sql_ins:
                logger.debug(f"  rea {rida} lisamine")
                self.context.target(sql_ins, False)
                logger.debug(f"  rida {rida} lisatud")         
        
    def calculate_deletion_dates(self, sheet, min_row, max_row, delete_date_column) -> tuple[str|None, str|None]:
        excel_col = None
        excel_col_action = None
        for col in self.action['cols']:
            if col.get('name', '') == delete_date_column:
                excel_col = col.get('from', None)
                excel_col_action = col.get('transform', None)
                break
        if excel_col is None:
            return (None, None)

        # on teada täht (tähed), leiame veeru jrk numbri (0-algusega)
        excel_pos = excel_label_to_number(excel_col) # failis utils.py
        
        cell_values = [] # list of ISO date strings (2024-01-23, 2024-11-06', ..) 
        # läbime selle ÜHE veeru kõigi ridadega (min_col ja max_col on see jrk/pos)
        for jrk, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=excel_pos, max_col=excel_pos, values_only=True), min_row):
            cell_value = sanityze(row[0], excel_col_action)
            if cell_value is not None:
                cell_values.append(cell_value) # kuupäev ISO kujul, et tekst oleks mõistlikult sorteeritav
            else:
                # kui kasvõi üks kuupäevaveerus olev väärtus pole kuupäev (või on tühi), siis me keeldume töötamast
                # jamasid saab liiga palju tulla sellest
                logger.debug(f'Kuupäevaveerus leiti mittesobiv väärtus. Rida {jrk} veerg {excel_col}')
                return (None, None)
        
        cell_values.sort() # sorteerime kasvavalt
        cell_min_value = cell_values[0] # ja võtame esimese ehk väikseima
        cell_values.sort(reverse=True) # sorteerime kahanevalt 
        cell_max_value = cell_values[0] # ja võtame esimese ehk suurima

        delete_before = self.action.get('delete', 'month')
        # leida min jaoks year/month 1.päev ja max jaoks viimane (iso formaadis)
        period_start_dt: str | None = None
        period_end_dt: str | None = None
        if delete_before == 'year':
            period_start_dt = start_of_year(cell_min_value)
            period_end_dt = end_of_year(cell_max_value) 
        if delete_before == 'month':
            period_start_dt = start_of_month(cell_min_value)
            period_end_dt = end_of_month(cell_max_value) 
        if delete_before == 'day':
            period_start_dt = cell_min_value
            period_end_dt = cell_max_value
        return (period_start_dt, period_end_dt)
                
    def get_sheet(self, excel_file, human_pos):
        """
        Return sheet object 
        """
        # lets find out needed sheet
        workbook = openpyxl.load_workbook(excel_file, data_only=True) # data_only
        
        # workbookil on olemas list lehtede nimedega (.sheetnames) ja workbook ise on dict, kus võtmeks on lehe nimi
        # list on positsionaalne (loodetavasti ikka sama järjekord, mis inimestel näha)
        # def-failis arendaja poolt määratud sheet 1 peaks ideeliselt olema sheetnames listi element 0
        sheet_names_list = workbook.sheetnames
        needed_name = sheet_names_list[human_pos - 1]
        sheet = workbook[needed_name]
        return sheet
        
    def row_limits(self, sheet) -> tuple[int, int]:
        min_row = self.action['source'].get('start_row', 1) # min row = 1 (iter_rows kasutab inimmumbreid)
        max_row = sheet.max_row
        return (min_row, max_row)
